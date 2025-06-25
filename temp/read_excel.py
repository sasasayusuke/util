#!/usr/bin/env python3
import subprocess
import sys
import json

# Install required packages
subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])

from openpyxl import load_workbook

def analyze_excel_format(filename):
    wb = load_workbook(filename, data_only=False)
    
    format_info = {
        "filename": filename,
        "sheets": [],
        "formatting_patterns": []
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "headers": [],
            "data_samples": [],
            "cell_formats": []
        }
        
        # Get headers (assuming first row)
        if sheet.max_row > 0:
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                if cell.value:
                    headers.append({
                        "column": col,
                        "value": str(cell.value),
                        "font": str(cell.font.name) if cell.font else None,
                        "bold": cell.font.bold if cell.font else False,
                        "fill_color": str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color else None
                    })
            sheet_info["headers"] = headers
        
        # Get sample data (rows 2-5)
        for row in range(2, min(6, sheet.max_row + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    row_data.append({
                        "column": col,
                        "value": str(cell.value),
                        "data_type": cell.data_type,
                        "number_format": cell.number_format
                    })
            if row_data:
                sheet_info["data_samples"].append({"row": row, "data": row_data})
        
        format_info["sheets"].append(sheet_info)
    
    return format_info

# Analyze sample.xlsx
try:
    result = analyze_excel_format("sample.xlsx")
    print(json.dumps(result, indent=2, ensure_ascii=False))
except Exception as e:
    print(f"Error: {str(e)}")