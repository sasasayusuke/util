from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer, range_copy_cell, relative_reference_copy, relative_reference_range_copy
from app.core.config import settings
from app.core.exceptions import ServiceError
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter


# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class toukeisuiihyoDateService(BaseService):
    """統計推移表(日単位)サービス"""

    def display(self, request, session) -> Dict[str, Any]:

        # 統計推移表(日単位)の印刷処理
        logger.info("【START】統計推移表(日単位)処理")

        storedname="usp_TK3600統計推移表_日単位"

        #367日以上の範囲を指定されている場合はエラーとする
        start_date = datetime.strptime(request.params['@iS集計日付'], '%Y/%m/%d')
        end_date = datetime.strptime(request.params['@iE集計日付'], '%Y/%m/%d')
        period_date = end_date - start_date
        logger.debug(str(period_date.days))
        if period_date.days >= 366:
            raise ServiceError('範囲が大きすぎます。（最大366日）')

        logger.info('取得開始')
        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=request.params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        logger.info('取得終了')
        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        if storedresults["output_values"]["@RetST"] == -1:
            raise ServiceError(storedresults["output_values"]["@RetMsg"])

        with get_excel_buffer() as buffer:
            wb = create_excel_object('Temp統計推移表_日単位.xlsx')

            # DBシートにデータを登録
            ws = wb['DB']

            # 顧客別　売上・原価
            row_num = 3
            for row_data in storedresults['results'][0]:
                ws.cell(row=row_num, column=1, value=row_data['営業担当CD'])
                ws.cell(row=row_num, column=2, value=row_data['番号'])
                ws.cell(row=row_num, column=3, value=row_data['集計CD'])
                ws.cell(row=row_num, column=4, value=row_data['年月日'])
                ws.cell(row=row_num, column=5, value=row_data['作業区分CD'])
                ws.cell(row=row_num, column=6, value=row_data['売上税抜金額'])
                ws.cell(row=row_num, column=7, value=row_data['原価税抜金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 工事　売上・原価
            row_num = 3
            for row_data in storedresults['results'][1]:
                ws.cell(row=row_num, column=10, value=row_data['工事担当CD'])
                ws.cell(row=row_num, column=11, value=row_data['番号'])
                ws.cell(row=row_num, column=12, value=row_data['集計CD'])
                ws.cell(row=row_num, column=13, value=row_data['年月日'])
                ws.cell(row=row_num, column=14, value=row_data['作業区分CD'])
                ws.cell(row=row_num, column=15, value=row_data['売上税抜金額'])
                ws.cell(row=row_num, column=16, value=row_data['原価税抜金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 入出庫
            row_num = 3
            for row_data in storedresults['results'][2]:
                ws.cell(row=row_num, column=25, value=row_data['集計CD'])
                ws.cell(row=row_num, column=26, value=row_data['年月日'])
                ws.cell(row=row_num, column=27, value=row_data['出庫金額'])
                ws.cell(row=row_num, column=28, value=row_data['入庫金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 三研　売上（円）
            row_num = 3
            for row_data in storedresults['results'][3]:
                ws.cell(row=row_num, column=31, value=row_data['年月'])
                ws.cell(row=row_num, column=32, value=row_data['売上税抜金額_円'])
                ws.cell(row=row_num, column=33, value=row_data['原価税抜金額_円'])
                ws.cell(row=row_num, column=34, value=row_data['人件一般費_円'])
                ws.cell(row=row_num, column=35, value=row_data['営業外費_円'])

                # データを書き込んだ後、次の行に移動
                row_num += 1

            # サービス　売上（円）
            row_num = 3
            for row_data in storedresults['results'][4]:
                ws.cell(row=row_num, column=67, value=row_data['年月'])
                ws.cell(row=row_num, column=68, value=row_data['売額_貨物自_三商'])
                ws.cell(row=row_num, column=69, value=row_data['原額_貨物自_三商'])
                ws.cell(row=row_num, column=70, value=row_data['売額_貨物自_店装'])
                ws.cell(row=row_num, column=71, value=row_data['原額_貨物自_店装'])
                ws.cell(row=row_num, column=72, value=row_data['売額_貨物自_サンプラ'])
                ws.cell(row=row_num, column=73, value=row_data['原額_貨物自_サンプラ'])
                ws.cell(row=row_num, column=74, value=row_data['売額_貨物自_三シャ'])
                ws.cell(row=row_num, column=75, value=row_data['原額_貨物自_三シャ'])
                ws.cell(row=row_num, column=76, value=row_data['売額_貨物自_トラン'])
                ws.cell(row=row_num, column=77, value=row_data['原額_貨物自_トラン'])
                ws.cell(row=row_num, column=78, value=row_data['売額_貨物自_その他'])
                ws.cell(row=row_num, column=79, value=row_data['原額_貨物自_その他'])
                ws.cell(row=row_num, column=80, value=row_data['売額_貨物自_事故'])
                ws.cell(row=row_num, column=81, value=row_data['原額_貨物自_事故'])
                ws.cell(row=row_num, column=82, value=row_data['売額_貨物庸_三商'])
                ws.cell(row=row_num, column=83, value=row_data['原額_貨物庸_三商'])
                ws.cell(row=row_num, column=84, value=row_data['売額_貨物庸_店装'])
                ws.cell(row=row_num, column=85, value=row_data['原額_貨物庸_店装'])
                ws.cell(row=row_num, column=86, value=row_data['売額_貨物庸_サンプラ'])
                ws.cell(row=row_num, column=87, value=row_data['原額_貨物庸_サンプラ'])
                ws.cell(row=row_num, column=88, value=row_data['売額_貨物庸_三シャ'])
                ws.cell(row=row_num, column=89, value=row_data['原額_貨物庸_三シャ'])
                ws.cell(row=row_num, column=90, value=row_data['売額_貨物庸_トラン'])
                ws.cell(row=row_num, column=91, value=row_data['原額_貨物庸_トラン'])
                ws.cell(row=row_num, column=92, value=row_data['売額_貨物庸_その他'])
                ws.cell(row=row_num, column=93, value=row_data['原額_貨物庸_その他'])
                ws.cell(row=row_num, column=94, value=row_data['売額_貨物庸_事故'])
                ws.cell(row=row_num, column=95, value=row_data['原額_貨物庸_事故'])
                ws.cell(row=row_num, column=96, value=row_data['売額_物流_ルート'])
                ws.cell(row=row_num, column=97, value=row_data['原額_物流_ルート'])
                ws.cell(row=row_num, column=98, value=row_data['売額_物流_丸和'])
                ws.cell(row=row_num, column=99, value=row_data['原額_物流_丸和'])
                ws.cell(row=row_num, column=100, value=row_data['売額_物流_その他'])
                ws.cell(row=row_num, column=101, value=row_data['原額_物流_その他'])
                ws.cell(row=row_num, column=102, value=row_data['売額_物流_事故'])
                ws.cell(row=row_num, column=103, value=row_data['原額_物流_事故'])
                ws.cell(row=row_num, column=104, value=row_data['人件一般費_貨物'])
                ws.cell(row=row_num, column=105, value=row_data['営業外費_貨物'])
                ws.cell(row=row_num, column=106, value=row_data['人件一般費_物流'])
                ws.cell(row=row_num, column=107, value=row_data['営業外費_物流'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 営業担当別経費
            row_num = 3
            for row_data in storedresults['results'][5]:
                ws.cell(row=row_num, column=110, value=row_data['業種区分_未使用'])
                ws.cell(row=row_num, column=111, value=row_data['部CD_未使用'])
                ws.cell(row=row_num, column=112, value=row_data['年月日'])
                ws.cell(row=row_num, column=113, value=row_data['人件費一般費'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 部署別　売上・原価
            row_num = 3
            for row_data in storedresults['results'][6]:
                ws.cell(row=row_num, column=117, value=row_data['部署CD'])
                ws.cell(row=row_num, column=118, value=row_data['年月日'])
                ws.cell(row=row_num, column=119, value=row_data['作業区分CD'])
                ws.cell(row=row_num, column=120, value=row_data['売上税抜金額'])
                ws.cell(row=row_num, column=121, value=row_data['原価税抜金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 商研シート関数追加
            ws = wb['商研']
            current_col = 9
            current_date = end_date
            # 列を追加
            for i in range(period_date.days + 1):
                if i != 0:
                    # 書式・数式を反映
                    relative_reference_range_copy(ws,ws,current_col-3,3,current_col-1,99,3,0)
                    range_copy_cell(ws,ws,current_col-3,2,current_col-1,99,3,0,False)
                    # 列幅を調整
                    ws.column_dimensions["{}".format(get_column_letter(current_col))].width = 10.1
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 1))].width = 5.7
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 2))].width = 10.1

                # 日付追加
                ws['{}2'.format(get_column_letter(current_col))].number_format = 'yyyy年m月d日'
                ws['{}2'.format(get_column_letter(current_col))].value = current_date
                current_col += 3
                current_date = current_date + relativedelta(days=-1)
            last_col = current_col-1
            last_col_alpha = get_column_letter(last_col)
            # 合計の関数を修正
            ws.cell(row=4, column=6, value="=SUM(I4:{}4)".format(last_col_alpha))
            ws.cell(row=5, column=6, value="=SUM(I5:{}5)".format(last_col_alpha))
            ws.cell(row=6, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I6:{t1}6)".format(t1=last_col_alpha))
            ws.cell(row=7, column=6, value="=SUM(I5:{}5)".format(last_col_alpha))
            ws.cell(row=8, column=6, value="=SUM(I5:{}5)".format(last_col_alpha))
            ws.cell(row=9, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I9:{t1}9)".format(t1=last_col_alpha))
            ws.cell(row=9, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I9:{t1}9)".format(t1=last_col_alpha))
            ws.cell(row=11, column=8, value="=SUM(I11:{}11)".format(last_col_alpha))
            for i in range(13, 19):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=20, column=8, value="=SUM(I20:{}20)".format(last_col_alpha))
            for i in range(22, 28):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=29, column=8, value="=SUM(I29:{}29)".format(last_col_alpha))
            for i in range(31, 100):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            
            ws.print_area = 'A1:{}107'.format(last_col_alpha)
            ws.sheet_view.view = 'pageBreakPreview'
            
            # 商研シートをアクティブにする
            wb.active = wb['商研']

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, None)

            logger.debug("Excel file generated successfully")

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
