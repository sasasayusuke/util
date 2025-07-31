from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer, range_copy_cell
from app.utils.service_utils import ClsTokuisaki
from app.core.config import settings
from app.core.exceptions import ServiceError
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill, Font

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class StoringListService(BaseService):
    """入庫リストサービス"""

    def display(self, request, session) -> StreamingResponse:

        # 入庫リストの印刷処理
        logger.info("【START】入庫リスト処理")

        storedname="usp_MT0600入庫リスト"

        params = {
            "@i入庫日":request.params["@i入庫日"],
            "@iS得意先CD":request.params["@iS得意先CD"],
            "@iE得意先CD":request.params["@iE得意先CD"],
            "@iS仕入先CD":request.params["@iS仕入先CD"],
            "@iE仕入先CD":request.params["@iE仕入先CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetDcnt":"INT", "@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            # wb = create_excel_object('Temp月ズレチェック表.xlsx')
            workbook = create_excel_object('Temp社内伝入庫リスト.xlsx')
            template_sheet = workbook['Template']

            # 初期化
            start_row = 7
            start_core_columns = 3
            end_core_columns = 14
            current_mitumori_number = None

            # 作成日を設定
            template_sheet["O1"] = f"作成日： {datetime.now().strftime('%Y/%m/%d')}"
            # 発注者を設定
            template_sheet["N4"] = request.params["AccountUserName"]

            # データのループ
            for result in storedresults['results'][0]:
                mitumori_number = result['見積番号']
                mitumori_title = result['見積件名']

                # 見積番号ごとに新しいシートを作成
                if mitumori_number != current_mitumori_number:
                    current_mitumori_number = mitumori_number
                    sheet_name = f"{mitumori_number}"
                    new_sheet = workbook.copy_worksheet(template_sheet)
                    ws = new_sheet
                    ws.title = sheet_name
                    ws.print_title_rows = '$1:$6'

                    # 改ページプレビューの設定
                    ws.sheet_view.view = 'pageBreakPreview'
                    ws.sheet_view.zoomScaleSheetLayoutView= 100

                    workbook.active = ws
                    current_row = start_row  # 行カウンタをリセット

                    # ヘッダー情報を設定
                    ws["D1"] = result["納入先名1"] + result["納入先名2"]
                    ws["D2"] = f"{mitumori_title if mitumori_title else ' '}  No.{mitumori_number if mitumori_number else ' '}"
                    ws["D4"] = result["入庫日"].strftime("%m/%d")
                    ws["A7"] = result["入庫日"]
                    if result["入出庫用CD"]:
                        ws["B7"] = f"*{result['入出庫用CD']}*"

                    # 行を調整
                    rec_cnt = result['Count']
                    if rec_cnt > 2:
                        # 不足している行を挿入し、書式をコピーする
                        ws.insert_rows(start_row + 1 , rec_cnt - 2)
                        original_height = ws.row_dimensions[start_row].height
                        for row in range(start_row + 1, start_row + rec_cnt):
                            ws.row_dimensions[row].height = original_height
                            for col in range(start_core_columns, end_core_columns + 1):
                                ws.cell(row=row, column=col)._style = copy(ws.cell(row=start_row, column=col)._style)

                    elif rec_cnt < 2:
                        # ダミー行が残ってしまうので削除
                        ws.delete_rows(start_row)

                # データ挿入
                ws.cell(row=current_row, column=3, value=current_row - start_row + 1)
                ws.cell(row=current_row, column=4, value=result['棚番名'])
                ws.cell(row=current_row, column=5, value=result['仕入先CD'])
                ws.cell(row=current_row, column=6, value=result['仕入先名'])
                ws.cell(row=current_row, column=7, value=result['製品NO'])
                ws.cell(row=current_row, column=8, value=result['仕様NO'])
                ws.cell(row=current_row, column=9, value=result['漢字名称'])
                ws.cell(row=current_row, column=10, value=result['W'])
                ws.cell(row=current_row, column=11, value=result['D'])
                ws.cell(row=current_row, column=12, value=result['H'])
                ws.cell(row=current_row, column=13, value=result['発注数'])

                current_row += 1

            # 不要なテンプレートシートの削除
            workbook.remove(template_sheet)

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, workbook, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            ) 
       
    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
