import io
import logging
from datetime import datetime
from typing import Dict, Any, List
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image
from fastapi.responses import StreamingResponse
from fastapi import HTTPException
from app.core.exceptions import ServiceError
from app.core.config import settings
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.string_utils import format_jp_date
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    delete_excel_object,
)
from decimal import Decimal

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class KeikakakuninhyouService(BaseService):
    def display(self, request, session):
        """経過確認表出力処理 (index)"""

        # JavaScriptから渡されたデータを取得
        data = request.json()  # ここで受け取ったデータを使う

        # データが文字列の場合、JSONとしてパースする
        if isinstance(data, str):
            try:
                import json
                data = json.loads(data)  # JSONとして読み込む
            except json.JSONDecodeError:
                raise ServiceError('無効なデータ形式です')

        # 'data1', 'data2', 'data3'などの辞書をリストに変換
        if isinstance(data, dict):
            data = [value for key, value in data.items()]

        # データ有無チェック
        if not data:
            raise ServiceError('該当データなし')

        # シート作成処理
        with get_excel_buffer() as buffer:
            # テンプレートファイル
            wb = create_excel_object('Template_経過確認表.xlsx')
            # テンプレートシートのコピー
            ws_template = wb['別紙_テンプレート']
            ws = wb.copy_worksheet(ws_template)
            # コピー先シート名
            ws.title = "経過確認表"
            # ページレイアウト設定の引き継ぎ
            ws.page_setup = copy(ws_template.page_setup)  # ページ設定
            ws.page_margins = copy(ws_template.page_margins)  # 余白設定
            ws.print_options = copy(ws_template.print_options)  # 印刷設定

            # 作成日をL1に挿入
            ws['L1'] = f"作成日： {self.format_date(datetime.now().strftime('%Y/%m/%d'))}"

            # データ挿入処理
            self.insert_data_to_sheet(ws, data)

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # 処理が終わったらテンプレートシートを削除
            del wb['別紙_テンプレート']

            # excel保存処理
            save_excel_to_buffer(buffer, wb, "sanwa55")
            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={'Content-Disposition': 'attachment; filename="sample.xlsx"'})

    def insert_data_to_sheet(self, ws, data: List[Dict[str, Any]]):
        """データをExcelシートに挿入する"""
        start_row = 3  # データ挿入を開始する行（3行目から開始）

        # 1行目の高さを取得して、それを後続の行に適用
        first_row_height = ws.row_dimensions[3].height
        if first_row_height:
            ws.row_dimensions[start_row].height = first_row_height  # 2行目の高さを1行目に合わせる

        # 枠線用のスタイル設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_border_left = Border(
            left=Side(style='thick')  # 左側の太線
        )

        thick_border_right = Border(
            right=Side(style='thick')  # 右側の太線
        )

        thick_border_bottom = Border(
            bottom=Side(style='thick')  # 下側の太線
        )

        # data が辞書形式で複数のデータが入っているので、ループ処理
        row_idx = start_row  # 挿入する行のインデックスを start_row で初期化
        for row_data in data:
            if isinstance(row_data, dict):
                for entry in row_data.values():  # 各data1, data2, data3を取り出す
                    # entry には '見積番号' や '種別' などの情報が含まれている
                    if isinstance(entry, dict):  # data1, data2, data3 の内容
                        # 各セルにデータを挿入
                        ws.cell(row=row_idx, column=1, value=int(entry.get("見積番号", "")))._style = copy(ws.cell(row=start_row, column=1)._style)
                        ws.cell(row=row_idx, column=2, value=entry.get("種別", ""))._style = copy(ws.cell(row=start_row, column=2)._style)
                        ws.cell(row=row_idx, column=3, value=entry.get("得意先", ""))._style = copy(ws.cell(row=start_row, column=3)._style)
                        ws.cell(row=row_idx, column=4, value=entry.get("納入先", ""))._style = copy(ws.cell(row=start_row, column=4)._style)
                        ws.cell(row=row_idx, column=5, value=entry.get("見積件名", ""))._style = copy(ws.cell(row=start_row, column=5)._style)
                        ws.cell(row=row_idx, column=6, value=self.format_currency((entry.get("見積金額", 0).replace(",",""))))._style = copy(ws.cell(row=start_row, column=6)._style)
                        ws.cell(row=row_idx, column=7, value=entry.get("Wリース区分", 0))._style = copy(ws.cell(row=start_row, column=7)._style)
                        ws.cell(row=row_idx, column=8, value=entry.get("W物件区分", 0))._style = copy(ws.cell(row=start_row, column=8)._style)
                        ws.cell(row=row_idx, column=9, value=self.format_date_datetype(entry.get("完了日", "")))._style = copy(ws.cell(row=start_row, column=9)._style)
                        ws.cell(row=row_idx, column=10, value=self.format_date_datetype(entry.get("請求書", "")))._style = copy(ws.cell(row=start_row, column=10)._style)
                        ws.cell(row=row_idx, column=11, value=self.format_date_datetype(entry.get("請求予定", "")))._style = copy(ws.cell(row=start_row, column=11)._style)
                        ws.cell(row=row_idx, column=12, value=entry.get("経過備考1", ""))._style = copy(ws.cell(row=start_row, column=12)._style)
                        ws.cell(row=row_idx, column=13, value=entry.get("経過備考2", ""))._style = copy(ws.cell(row=start_row, column=13)._style)

                        # ここで枠線を適用
                        for col in range(1, 14):  # 1列目から13列目までループ
                            cell = ws.cell(row=row_idx, column=col)

                            # 左端（1列目）のセルには左側のみ太線
                            if col == 1:
                                # 既存の細い枠線に加えて太線を設定
                                cell.border = Border(
                                    left=Side(style='thick'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                            # 右端（13列目）のセルには右側のみ太線
                            elif col == 13:
                                # 既存の細い枠線に加えて太線を設定
                                cell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thick'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                            else:
                                cell.border = thin_border

                        # 次の行に移動
                        row_idx += 1  

                        # 高さを継承（次の行にも同じ高さを設定） 
                        if row_idx <= len(data):  # 次の行がデータの範囲内である場合 
                            ws.row_dimensions[row_idx].height = first_row_height

        # 最後の行に下枠線だけ太線を適用
        last_row = row_idx - 1  # 最後のデータ行
        if last_row >= start_row:  # 空白行を避けるためのチェック
            for col in range(1, 14):  # 1列目から13列目までループ
                cell = ws.cell(row=last_row, column=col)
                # 既存の枠線に加えて下の枠線を太線に変更
                cell.border = Border(
                    left=Side(style='thin') if col != 1 else Side(style='thick'),
                    right=Side(style='thin') if col != 13 else Side(style='thick'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')  # 最後の行なので下側を太線に
                )

        # 最後の行が空白でないことを確認した後、行の高さの設定をします
        if row_idx > start_row:
            ws.row_dimensions[row_idx].height = first_row_height

    def format_date(self, date_value: str, date_format: str = "%Y/%m/%d") -> str:
        """日付をフォーマット"""
        if date_value:
            try:
                return datetime.strptime(date_value, "%Y/%m/%d").strftime(date_format)
            except ValueError:
                return date_value
        return ""

    def format_date_datetype(self, date_value: str):
        """日付をフォーマット"""
        if date_value:
            try:
                return datetime.strptime(date_value, "%Y/%m/%d")
            except ValueError:
                return date_value
        return ""

    def get_lease_type(self, value: Any) -> str:
        """リース区分の変換"""
        try:
            # 文字列として渡ってきた場合でも数値に変換してから処理
            value = int(value)
        except (ValueError, TypeError):
            pass  # 変換できなければそのまま文字列として扱う

        mapping = {1: "通常", 2: "リース"}
        return mapping.get(value, "")

    def get_property_type(self, value: Any) -> str:
        """物件区分の変換"""
        try:
            # 文字列として渡ってきた場合でも数値に変換してから処理
            value = int(value)
        except (ValueError, TypeError):
            pass  # 変換できなければそのまま文字列として扱う

        mapping = {1: "店舗MD戦略部", 2: "管財部", 3:"調剤企画部", 4:"建築部", 5:"総務部", 6:"商品部"}
        return mapping.get(value, "")
    
    def format_currency(self, value: Any) -> str:
        """通貨のフォーマット"""
        try:
            if value == "":
                value = 0
            return int(value)
        except (ValueError, TypeError):
            return str(value)

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
