import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
import os
from openpyxl.reader import excel
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.worksheet import Worksheet
from app.services.base_service import BaseService
from app.utils.service_utils import ClsTanto,ClsKaisya,ClsMitumoriH,ClsZeiritsu,ClsOutputRireki
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    execute_vba,
    set_footer_whitespace
)
from app.core.config import settings
from app.core.exceptions import ServiceError
from openpyxl.styles.borders import Border,Side
from openpyxl.styles import Alignment,Font
from openpyxl import load_workbook
from app.utils.string_utils import null_to_zero,convert_1_to_a,convert_rc_to_cell
from zipfile import ZipFile
import xml.etree.ElementTree as ET


# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class hattyusyoService(BaseService):
    """発注書サービス"""

    def display(self, request, session) -> StreamingResponse:

        # 発注書の印刷処理
        logger.info("【START】発注書")

        stored_name = ""
        if request.params["@format_category"] == "0":
            stored_name = "usp_HC0200発注書出力"
        elif request.params["@format_category"] == "1":
            stored_name = "usp_HC0201発注書出力_集計"
        else:
            raise ServiceError('書式区分が入力されていません')

        stored_params = {
            "@i見積番号":request.params["@i見積番号"],
            "@i開始CD":request.params["@i開始CD"],
            "@i終了CD":request.params["@i終了CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=stored_params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if len(storedresults["results"][0]) == 0 or len(storedresults['results'][1]) == 0:
            raise ServiceError('該当データなし')

        GrsHead = storedresults["results"][0]
        GrsDetail = storedresults["results"][1]

        with get_excel_buffer() as buffer:
            wb = create_excel_object('Template_発注書.xlsx')
            base_ws = wb["Template"]
            total_ws = wb["total_template"]

            header_fontSize = 10

            # 担当者
            base_ws['T6'].value = request.user
            base_ws.cell(row=6,column=20).font = Font(size=header_fontSize)
            base_ws.cell(row=6,column=20).alignment = Alignment(vertical='center',horizontal='left')

            # 表を何列拡張するか計算
            base_col_count = 7
            journal_count = len(GrsHead)
            if journal_count % base_col_count != 0:
                journal_count = ((journal_count//base_col_count)+1) *base_col_count

            # 表を拡張
            style_solid = Side(style='thin',color='000000')
            style_medium = Side(style='medium',color='000000')
            style_double = Side(style='double',color='000000')
            if journal_count != base_col_count:

                # 
                for i in range(1,journal_count-base_col_count+1):
                    base_ws.insert_cols(22)
                    base_ws.cell(row=9,column=22).border = Border(top=style_solid,bottom=style_double,left=style_solid,right=style_solid)
                    base_ws.cell(row=10,column=22).border = Border(top=style_solid,bottom=style_solid,left=style_solid,right=style_solid)
                    base_ws.cell(row=9,column=22)._style = copy(base_ws.cell(row=9, column=17)._style)
                    base_ws.cell(row=10,column=22)._style = copy(base_ws.cell(row=10, column=17)._style)
                # 
                for i in range(17,17+journal_count):
                    base_ws.column_dimensions[convert_1_to_a(i)].width=5.5
                
                # ヘッダーの文言とスタイル
                headers_base_col = 17
                for i in range(1,int(journal_count / base_col_count)):
                    col_num = headers_base_col + (base_col_count * i)
                    base_ws.cell(row=3,column=col_num,value="いつもお世話になっております。").font = Font(size=header_fontSize)
                    base_ws.cell(row=4,column=col_num,value="金額と納期確認して頂き返信をお願い致します。").font = Font(size=header_fontSize)

                    base_ws.cell(row=6,column=col_num,value="三和商研").font = Font(size=header_fontSize)
                    base_ws.cell(row=6,column=col_num).alignment = Alignment(vertical='center',horizontal='left')

                    base_ws.cell(row=6,column=col_num+3,value=request.user).font = Font(size=header_fontSize)
                    base_ws.cell(row=6,column=col_num+3).alignment = Alignment(vertical='center',horizontal='left')

            # 略称(列ごとに繰り返す)
            for i,journal in enumerate(GrsHead, start=17):
                if journal["略称"] == "":
                    base_ws.cell(row=9,column=i,value="")
                else:
                    base_ws.cell(row=9,column=i,value=journal["略称"])
                    
            # 合計を代入する関数
            def input_sum(ws,row,journal_count,style_medium):
                ws.merge_cells("J{0}:M{0}".format(row))
                ws["J{}".format(row)].number_format = "#,##0"
                ws.cell(row=row,column=10,value="=SUM(M{}:M{})".format(10,row-1)).font = Font(size=14)
                ws.cell(row=row,column=4,value="≪　合計　≫")._style = copy(total_ws.cell(row=1,column=1)._style)
                ws.cell(row=row,column=10).alignment = Alignment(vertical='center',horizontal='right')
                for i in range(1,17 + journal_count):
                    if i in [1,2,3] or i >= 16:
                        ws.cell(row=row,column=i).border = Border(top=style_medium,bottom=style_medium,left=style_solid,right=style_solid)
                    elif i == 10:
                        ws.cell(row=row,column=i).border = Border(top=style_medium,bottom=style_medium,left=style_solid)
                    else:
                        ws.cell(row=row,column=i).border = Border(top=style_medium,bottom=style_medium)
                
                # 保留（不要なのでコメントアウト）
                # from_cell = convert_rc_to_cell(row+2,2)
                # to_cell = convert_rc_to_cell(row+6,20)
                # ws.merge_cells("{}:{}".format(from_cell,to_cell))
                # for r in ws["{}:{}".format(from_cell,to_cell)]:
                #     for cell in r:
                #         cell.border = Border(top=style_medium,bottom=style_medium,left=style_medium,right=style_medium)
                
                # 印刷範囲
                to_cell = convert_rc_to_cell(row,journal_count+16)
                ws.print_area = "A1:{}".format(to_cell)
                return ws

            # 仕入先ごとにシート作成・入力
            old_suplier = ""
            old_delivery = ""
            ws:Worksheet
            row = 10
            total = 0
            base_row = 16
            first_flg = False
            for record in GrsDetail:

                #  or 配送先の場合は別シート
                if old_suplier != record["仕入先CD"] or old_delivery != record["配送先CD"]:
                    # 合計
                    if first_flg:
                        ws = input_sum(ws,row,journal_count,style_medium)

                    first_flg = True

                    old_suplier = record["仕入先CD"]
                    old_delivery = record["配送先CD"]

                    # シートコピー
                    ws = wb.copy_worksheet(base_ws)
                    ws.title = "{}-{}".format(old_suplier,old_delivery)
                    # 行を追加
                    row_count = record["仕入先件数"] + 1
                    if row_count % base_row != 0:
                        row_count = ((row_count // base_row) + 1) * base_row
                    start_row = 12
                    start_column = 1
                    end_row = start_row + row_count
                    end_column = 12 + journal_count -1
                    for i in range(1 , row_count):
                        range_copy_cell(ws,ws,start_column,start_row,end_column,start_row,0,i,True)
                    # 印刷設定
                    ws.print_title_cols = 'A:M'
                    ws.print_title_rows = '1:9'
                    ws.freeze_panes = 'A10'
                    # 改ページプレビューの設定
                    ws.sheet_view.view = 'pageBreakPreview'
                    ws.sheet_view.zoomScaleSheetLayoutView= 100
                    # フッター
                    ws.oddFooter.left.text = "※御不明な点は担当までお問い合わせください。"
                    ws.oddFooter.center.text = "&P / &Nページ"
                    ws.oddFooter.right.text = "株式会社 三 和 商 研\n TEL048-525-1171 FAX:048-525-0008"

                    # 書込み行をリセット
                    row = 10

                    # 配送先
                    ws['E3'].value = null_to_zero(record["配送先CD"],"")
                    ws['F3'].value = "{} {}".format(null_to_zero(record["配送先名1"],"") ,null_to_zero(record["配送先名2"],""))
                    ws['K3'].value = null_to_zero(record["納入先CD"],"")

                    if record["配送先CD"] == '01':
                        # 直送
                        ws['E4'].value = "〒{}".format(null_to_zero(record["郵便番号"],""))
                        ws['E5'].value = null_to_zero(record["住所1"],"")
                        ws['E6'].value = null_to_zero(record["住所2"],"")
                        ws['J4'].value = "TEL:{}".format(null_to_zero(record["納TEL"],""))
                    else:
                        ws['E4'].value = "〒{}".format(null_to_zero(record["配送郵便番号"],""))
                        ws['E5'].value = null_to_zero(record["配送住所1"],"")
                        ws['E6'].value = null_to_zero(record["配送住所2"],"")
                        ws['J4'].value = "TEL:{}".format(null_to_zero(record["配送電話番号"],""))

                    # 見積件名
                    ws['A8'].value = null_to_zero(record["見積件名"],"")
                    # 仕入先
                    ws['B3'].value = null_to_zero(record["仕入先CD"],"")
                    ws['B4'].value = null_to_zero(record["仕入先名1"],"")
                    ws['B5'].value = null_to_zero(record["仕入先名2"],"")

                    # 出力日
                    ws['K1'].value = "{}".format(datetime.date.today().strftime('%Y{0}%m{1}%d{2}').format('年','月','日出力分'))

                    # 発注番号
                    if record["統合見積番号"] == 0:
                        ws['F1'].value = null_to_zero(record["見積番号"],"")
                    else:
                        ws['F1'].value = null_to_zero(record["統合見積番号"],"")
                        # ↓ 新テンプレートではスペースがないため
                        # ws.merge_cells(start_row=2,start_column=6,end_row=2,end_column=7)
                        # ws['F2'].value = record["見積番号"]
                        # ws['F2'].fill = PatternFill(patternType='solid', fgColor='HC0C0C0')

                # ---------------------------------
                # 見積明細情報をセットします
                # ---------------------------------
                # 項目
                if record["見積区分"] in ["C","A","S"]:
                    ws.cell(row=row,column=1,value="").alignment = Alignment(vertical='center',horizontal='center',shrink_to_fit =True)
                else:
                    ws.cell(row=row,column=1,value=record["行番号"]).alignment = Alignment(vertical='center',horizontal='center',shrink_to_fit =True)

                # 製品No
                ws.cell(row=row,column=2,value=record["製品NO"])
                # 仕様No
                ws.cell(row=row,column=3,value=record["仕様NO"])

                # 品名
                ws.merge_cells(start_row=row,start_column=4,end_row=row,end_column=6)
                if record["見積区分"] in ["C","A","S"]:
                    ws.cell(row=row,column=4,value="\n  {}".format(record["漢字名称"])).alignment = Alignment(vertical='justify',horizontal='left',shrink_to_fit =True)
                else:
                    ws.cell(row=row,column=4,value="\n{}".format(record["漢字名称"])).alignment = Alignment(vertical='justify',horizontal='left',shrink_to_fit =True)
                
                W = ""
                D = ""
                H = ""# W
                if record['W'] != 0:
                    W = record['W']
                    
                # D
                if record['D'] != 0:
                    D = record['D']
                # else:
                #     if record["D1"] != 0 or record["D2"] != 0:
                #         D = "{}/{}".format(record['D1'],record['D2'])

                # H
                if record['H'] != 0:
                    H = format(record['H'])
                # else:
                #     if record["H1"] != 0 or record["H2"] != 0:
                #         H = "{}/{}".format(record['H1'],record['H2'])

                ws.cell(row=row,column=7,value=W)
                ws.cell(row=row,column=8,value=D)
                ws.cell(row=row,column=9,value=H)

                # 仕入単価
                if null_to_zero(request.params["@money_display"]) == '0':
                    if record["U区分"]=="U":
                        money = '{:,}'.format(record["仕入単価"])
                        ws.cell(row=row,column=10,value=money + '未')
                    else:
                        ws.cell(row=row,column=10,value=self.zero_to_emptyString(record["仕入単価"]))
                        ws.cell(row=row,column=10).number_format = '#,##0'
                    ws.cell(row=row,column=13,value=self.zero_to_emptyString(record["仕入金額"]))

                # 発注保留
                if record["U区分"] == "H":
                    ws.cell(row=row,column=13,value="発注保留")

                ws.cell(row=row,column=11,value=self.zero_to_emptyString(record["発注数"]))
                ws.cell(row=row,column=12,value=record["単位名"])
                ws.cell(row=row,column=14,value=self.zero_to_emptyString(record["客先在庫数"]))
                ws.cell(row=row,column=15,value=self.zero_to_emptyString(record["転用数"]))
                ws.cell(row=row,column=16,value=self.zero_to_emptyString(record["総数量"]))

                total = total + null_to_zero(record["仕入金額"])

                i = 12
                for i in range(1,len(GrsHead)+1):
                    ws.cell(row=row,column=i+16,value=self.zero_to_emptyString(record["仕分数{}".format(i)]))
                    ws.cell(row=row,column=i+16).number_format = '0'

                # スタイル
                for i in range(1,journal_count + 17):
                    ws.cell(row,i)._style=copy(ws.cell(10,i)._style)
                ws.row_dimensions[row].height = 30

                row = row + 1

            # 最終ページの合計
            ws = input_sum(ws,row,journal_count,style_medium)

            wb.remove(base_ws)
            wb.remove(total_ws)

            ws = wb.worksheets[0]
            wb.active = ws

            # 「余白に合わせて配置」をFalseに
            wb = set_footer_whitespace(wb)

            # 編集モードでリンクを張りたくない場合はコメント解除（処理時間+20秒）
            # wb = execute_vba(wb,'a','no_link_for_hattyusyo.xlsm',[])

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, 'sanwa55')

            # 履歴情報セット
            cOutputRireki = ClsOutputRireki()
            cOutputRireki.SetOutputLog(session,'発注書',request.params["@i見積番号"],request.params["@pc_name"])

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass

    def zero_to_emptyString(self,Number):
        if Number == 0 or Number == '0':
            return ''
        else:
            return Number