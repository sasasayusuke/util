from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse,JSONResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell_by_address,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from copy import copy
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font
from openpyxl.worksheet.pagebreak import Break

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class uriagemaeSeikyusyoService(BaseService):
    """請求書発行"""


    def display(self, request, session) -> Dict[str, Any]:

        # 請求書発行の印刷処理
        logger.info("【START】売上前請求書発行")

        if request.params["@i見積番号"] == "":
            raise ServiceError('見積番号が未入力です。')

        cMitumoriH = ClsMitumoriH(request.params["@i見積番号"])
        b_category = cMitumoriH.IsUCategory(session,'B')
        u_category = cMitumoriH.IsUCategory(session,'U')

        if b_category and u_category:
            raise ServiceError('原価・売価未確定項目があります。')
        if b_category:
            raise ServiceError('売価未確定項目があります。')
        if u_category:
            raise ServiceError('原価未確定項目があります。')


        # 担当者CD
        if request.params["@i表示担当"] == "":
            raise ServiceError('担当者CDが未入力です。')
        cTanto = ClsTanto()
        tanto_result = cTanto.GetbyID(session,request.params["@i表示担当"])
        if len(tanto_result) == 0:
            raise ServiceError('指定の担当者は存在しません。')

        # 請求日付
        cDates = ClsDates()
        cDates.GetbyId(session,"売掛月次更新日")
        gGetuDate = cDates.更新日付
        if request.params["@i請求日付"] == "":
            raise ServiceError('請求日付が未入力です。')

        # if datetime.strptime(request.params["@i請求日付"],"%Y-%m-%d") <= gGetuDate:
        #     raise ServiceError('更新済みの為、修正できません。')

        # 請求書発行日付
        if request.params["@i請求書発行日付"] == "":
            raise ServiceError('請求書発行履歴が未入力です。')

        self.download_kakutei(session,request.params["@i見積番号"],request.params["@i請求日付"])

        stored_results = self.download(session,request.params["@i見積番号"],request.params["@i請求書発行日付"])

        GrsHead  = stored_results["results"][0][0]
        GrsHeadU = stored_results["results"][1]
        GrsSheet = stored_results["results"][2]

        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_請求書.xlsx")
            ws = wb["Template"]
            copy_ws = wb["copy_sheet"]
            if (len(stored_results["results"][1]) == 1) and len(stored_results["results"][1]) + len(stored_results["results"][2]) <= 16:
                res = self.SeikyusyoMin(session,ws,copy_ws,GrsHead,GrsHeadU,GrsSheet,request.params)
            else:
                res = self.Seikyusyo(session,ws,copy_ws,GrsHead,GrsHeadU,GrsSheet,request.params)
            ws = res[0]
            row = res[1]

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # コピー用シートを削除
            wb.remove(copy_ws)

            # シート名の変更
            ws.title = '請求書'

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="O{}".format(row)
                )
                res_obj = converter.generate_pdf()
            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

                logger.debug("Excel file generated successfully")

            # 履歴情報セット
            cOutputRireki = ClsOutputRireki()
            cOutputRireki.SetOutputLog(session,'売上前請求書',request.params["@i見積番号"],request.params["@pc_name"])

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )


    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass

    def download_kakutei(self,session,estimate_no,billing_date) -> bool:
        res_bool = False

        stored_name = "usp_SE0803売上前請求作成処理"

        params = {
            "@i見積番号":estimate_no,
            "@i請求日付":billing_date
        }

        outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=params,outputparams=outputparams))

        logger.info('storedresults:{}'.format(storedresults['results']))

        result = storedresults["output_values"]
        if result["@RetST"] != 0:
            res_bool = False
            raise ServiceError("{}：{}".format(result["@RetST"],result["@RetMsg"]))
        else:
            res_bool = True

        return res_bool

    def download(self,session,estimate_no,billing_output_date):

        params = {
            "@i見積番号":estimate_no,
            "@i請求書発行日付":billing_output_date
        }

        stored_name = "usp_SE0804売上前請求書発行_客在"
        outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=params,outputparams=outputparams))

        result = storedresults["output_values"]
        if storedresults["output_values"]["@RetST"] != 0:
            raise ServiceError("{}：{}".format(result["@RetST"],result["@RetMsg"]))
        elif len(storedresults["results"][0]) == 0:
            raise ServiceError("該当データ無し")

        if len(storedresults["results"][1]) == 0:
            raise ServiceError("該当データ無し")

        if len(storedresults["results"][2]) == 0:
            raise ServiceError("該当データ無し")
        return storedresults


    def SeikyusyoMin(self,session,ws,copy_ws,GrsHead,GrsHeadU,GrsSheet,params):
        # ---------------------------------
        # 見積ヘッダー情報をセットします
        # ---------------------------------
        ws = self.SeikyusyoHeader(session,ws,GrsHead,params)


        # '---------------------------------
        # '見積シート内訳情報をセットします
        # '---------------------------------
        start_row = 21
        row = start_row
        sumFrom = start_row + 1
        siwake_no = ""

        for record in GrsSheet:

            if siwake_no != record["仕分番号"]:
                siwake_no = record["仕分番号"]
                ws.cell(row=row,column=1,value=record["仕分番号"])
                ws.cell(row=row,column=3,value=null_to_zero(record["仕分名称"],""))

                row = row + 1

            # 見積シート情報
            ws.cell(row=row,column=1,value=record["仕分番号"])
            if record["見積区分"] in ["C","A","S"]:
                ws.cell(row=row,column=2,value="")
                ws.cell(row=row,column=3,value="  {}".format(null_to_zero(record["漢字名称"],"")))
            else:
                ws.cell(row=row,column=2,value=record["行番号"])
                ws.cell(row=row,column=3,value="  {}".format(null_to_zero(record["漢字名称"],"")))
            ws.cell(row=row,column=5,value="{}{}".format(record["PC区分"],record["製品NO"]))
            ws.cell(row=row,column=6,value=record["仕様NO"])
            ws.cell(row=row,column=7,value=record["W"])

            if record["D"] == 0:
                if record['D1'] == 0 and record["D2"] == 0:
                    ws.cell(row=row,column=8,value="")
                else:
                    ws.cell(row=row,column=8,value="{}/{}".format(record["D1"],record["D2"]))
            else:
                ws.cell(row=row,column=8,value=record["D"])

            if record["H"] == 0:
                if record['H1'] == 0 and record["H2"] == 0:
                    ws.cell(row=row,column=9,value="")
                else:
                    ws.cell(row=row,column=9,value="{}/{}".format(record["H1"],record["H2"]))
            else:
                ws.cell(row=row,column=9,value=record["H"])

            ws.cell(row=row,column=10,value=record["売上数量"])
            ws.cell(row=row,column=11,value='=IF(MOD(J{0},1)=0,TEXT(J{0},"#,###"),TEXT(J{0},"#,##0.##"))'.format(row))
            ws.cell(row=row,column=12,value=record["単位名"])
            ws.cell(row=row,column=13,value=record["売上単価"])
            ws.cell(row=row,column=14,value='=IF(MOD(M{0},1)=0,TEXT(M{0},"#,###"),TEXT(M{0},"#,##0.##"))'.format(row))

            if record["見積区分"] in ["A","C","S"]:
                ws.cell(row=row,column=15,value=record["売上金額"] if record["売上金額"] != 0 else "")
            else:
                ws.cell(row=row,column=15,value=record["売上金額"])

            row = row + 1

        # 計算式をセット
        row = row + 1
        ws.cell(row=row,column=3,value="【合　　計】")
        ws["C{0}".format(row)].alignment = Alignment(horizontal='center')
        ws["D{0}".format(row)].alignment = Alignment(horizontal='center')
        ws.cell(row=row,column=15,value="=SUM(O{}:O{})".format(sumFrom,row-2))

        # 不要行を削除
        # for i in range(39,73):
        #     ws.delete_rows(39)

        return [ws,38]


    def Seikyusyo(self,session,ws,copy_ws,GrsHead,GrsHeadU,GrsSheet,params):
        # ---------------------------------
        # 見積ヘッダー情報をセットします
        # ---------------------------------
        ws = self.SeikyusyoHeader(session,ws,GrsHead,params)

        # '---------------------------------
        # '見積ヘッダー内訳情報をセットします
        # '---------------------------------
        start_row = 21
        row = start_row
        sumFrom = start_row + 1
        siwake_no = ""
        siwake_name = ""
        row_count_for_page = 1
        page_count = 1

        for record in GrsHeadU:

            if row == 38:
                ws.cell(row=row,column=3,value="次ページへ")
                ws["C{}".format(row)].alignment = Alignment(horizontal='center',vertical='center')

                range_copy_cell_by_address(copy_ws,ws,'A1','O35','A{}'.format(row+1),True)

                row = row + 4
                page_count = page_count + 1
                row_count_for_page = 1

            ws.cell(row=row,column=1,value=record["仕分番号"])
            ws.cell(row=row,column=3,value=null_to_zero(record["仕分名称"],""))
            ws.cell(row=row,column=15,value=record["売上金額"])

            row = row + 1

        if row == 38:
            ws.cell(row=row,column=3,value="次ページへ")
            ws["C{}".format(row)].alignment = Alignment(horizontal='center',vertical='center')

            range_copy_cell_by_address(copy_ws,ws,'A1','O35','A{}'.format(row+1),True)

            row = row + 3
            page_count = page_count + 1
            row_count_for_page = 1

        # 計算式をセット
        ws.cell(row=row+1,column=3,value="【合　　計】")
        ws["C{}".format(row+1)].alignment = Alignment(horizontal='center',vertical='center')
        ws.cell(row=row+1,column=15,value="=SUM(O21:O{})".format(row -1))

        # '---------------------------------
        # '見積シート内訳情報をセットします
        # '---------------------------------

        def insert_detail(ws,copy_ws,row):
            row = ((page_count-1) * 35) + 39
            range_copy_cell_by_address(copy_ws,ws,'A1','O35','A{}'.format(row),True)
            return [ws,row]


        for record in GrsSheet:
            if siwake_no != record["仕分番号"]:
                # 合計記載
                if siwake_no != "":
                    if row_count_for_page == 31 or row_count_for_page == 32:
                        res = insert_detail(ws,copy_ws,row)
                        ws= res[0]
                        row = res[1]
                        row = row + 3
                        page_count = page_count + 1
                        row_count_for_page = 1
                    # 計算式をセット
                    ws.cell(row+1,column=3,value='【{}小計】'.format(siwake_name))
                    ws["C{}".format(row+1)]._style = copy(copy_ws['Q1']._style)
                    ws.cell(row=row+1,column=15,value="=SUM(O{}:O{})".format(sumFrom,row-1))

                # 新頁
                res = insert_detail(ws,copy_ws,row)
                ws= res[0]
                row = res[1]
                row = row + 3
                page_count = page_count + 1
                row_count_for_page = 1
                ws.cell(row=row,column=1,value=record["仕分番号"])
                ws.cell(row=row,column=3,value=null_to_zero(record["仕分名称"],""))
                # ws.cell(row=row,column=15,value=record["売上金額"])

                siwake_no = record["仕分番号"]
                siwake_name = record["仕分名称"]
                sumFrom = row
                row = row + 1
                row_count_for_page = row_count_for_page + 1

            if row_count_for_page == 33:
                res = insert_detail(ws,copy_ws,row)
                ws= res[0]
                row = res[1]
                row = row + 3
                page_count = page_count + 1
                row_count_for_page = 1
            ws.cell(row=row,column=1,value=record["仕分番号"])
            if record["伝票区分"] in ['C','A','S']:
                ws.cell(row=row,column=2,value="")
                ws.cell(row=row,column=3,value="  {}".format(null_to_zero(record["漢字名称"],"")))
            else:
                ws.cell(row=row,column=2,value=record["行番号"])
                ws.cell(row=row,column=3,value=null_to_zero(record["漢字名称"],""))

            ws.cell(row=row,column=5,value="{}{}".format(record["PC区分"],record["製品NO"]))
            ws.cell(row=row,column=6,value=record["仕様NO"])

            ws.cell(row=row,column=7,value=record["W"])

            if record["D"] == 0:
                if record["D1"] == 0 and record["D2"] == 0:
                    ws.cell(row=row,column=8,value="")
                else:
                    ws.cell(row=row,column=8,value="{}/{}".format(record["D1"],record["D2"]))
            else:
                ws.cell(row=row,column=8,value=record["D"])

            if record["H"] == 0:
                if record["H1"] == 0 and record["H2"] == 0:
                    ws.cell(row=row,column=9,value="")
                else:
                    ws.cell(row=row,column=9,value="{}/{}".format(record["H1"],record["H2"]))
            else:
                ws.cell(row=row,column=9,value=record["H"])

            ws.cell(row=row,column=10,value=record["売上数量"])
            ws.cell(row=row,column=11,value='=IF(MOD(J{0},1)=0,TEXT(J{0},"#,###"),TEXT(J{0},"#,##0.##"))'.format(row))
            ws.cell(row=row,column=12,value=record["単位名"])
            ws.cell(row=row,column=13,value=record["売上単価"])
            ws.cell(row=row,column=14,value='=IF(MOD(M{0},1)=0,TEXT(M{0},"#,###"),TEXT(M{0},"#,##0.##"))'.format(row))

            if record["見積区分"] in ['A','C','S']:
                ws.cell(row=row,column=15,value=record["売上金額"] if record["売上金額"] != 0 else "")
            else:
                ws.cell(row=row,column=15,value=record["売上金額"])

            row = row + 1
            row_count_for_page = row_count_for_page + 1

        # 合計
        if row_count_for_page == 31 or row_count_for_page == 32:
            res = insert_detail(ws,copy_ws,row)
            ws= res[0]
            row = res[1]
            row = row + 3
            page_count = page_count + 1
            row_count_for_page = 1
        # 計算式をセット
        ws.cell(row+1,column=3,value='【{}小計】'.format(siwake_name))
        ws["C{}".format(row+1)]._style = copy(copy_ws['Q1']._style)
        ws.cell(row=row+1,column=15,value="=SUM(O{}:O{})".format(sumFrom,row-1))


        row = ((page_count-1) * 35) + 37
        return [ws,row]

    def SeikyusyoHeader(self,session,ws,GrsHead,params):
        # ---------------------------------
        # 見積ヘッダー情報をセットします
        # ---------------------------------
        ws['A1'].value = "№{}".format(GrsHead["見積番号"])
        ws['O1'].value = "{}".format(datetime.strftime(datetime.strptime(params["@i請求書発行日付"],"%Y-%m-%d"),'%Y{}%m{}%d{}'.format('年','月','日')))
        if GrsHead["得意先名2"] == "":
            ws['A3'].value = ""
            ws['A4'].value = GrsHead["得意先名1"]
        else:
            ws['A3'].value = GrsHead["得意先名1"]
            ws['A4'].value = GrsHead["得意先名2"]
        ws['D8'].value = GrsHead["合計金額"]

        cZeiritu = ClsZeiritsu()
        ws['A9'].value = "消費税等({}%)".format(cZeiritu.GetbyDate(session,GrsHead["請求日付"]))
        ws['D9'].value = GrsHead["外税額"]
        ws['D12'].value = GrsHead["見積件名"]

        if (GrsHead["納期S"] != None) and (GrsHead["納期E"] != None):
            ws["D14"].value = "{}～{}".format(
                datetime.strftime(GrsHead["納期S"],'%Y{}%m{}%d{}'.format('年','月','日')),
                datetime.strftime(GrsHead["納期E"],'%Y{}%m{}%d{}'.format('年','月','日')),
            )
        else:
            if GrsHead["納期S"] == None:
                ws['D14'].value = datetime.strftime(GrsHead["納期E"],'%Y{}%m{}%d{}'.format('年','月','日'))
            else:
                ws['D14'].value = datetime.strftime(GrsHead["納期S"],'%Y{}%m{}%d{}'.format('年','月','日'))

        if GrsHead["納入得意先CD"] in ['2401','8801']:
            ws['D15'].value = GrsHead["現場名"].replace(GrsHead["納入先CD"],"")
            ws['F15'].value = GrsHead["納入先CD"]
        else:
            ws['D15'].value = GrsHead["現場名"]
            ws['F15'].value = ""

        if GrsHead["支払条件"] == 0:
            ws['D16'].value = "別途御打ち合せによる"
        elif GrsHead["支払条件"] == 1:
            ws['D16'].value = GrsHead["支払条件その他"]

        ws['D17'].value = GrsHead["備考"]

        cTanto = ClsTanto()
        tanto_results = cTanto.GetbyID(session,params["@i表示担当"])
        if len(tanto_results[0]) != 0:
            ws['K15'].value = "担当：{}".format(tanto_results[0]["問い合わせ先"])
        else:
            ws['K15'].value = ""
        ws['K16'].value = ""

        if GrsHead["ウエルシア売場面積"] != 0:
            ws['A18'].value = "売 場 面 積 ："
            ws['D18'].value = "{}坪".format(GrsHead["ウエルシア売場面積"])
            tmp_side = Side(style='thin', color='000000')
            tmp_border = Border(bottom=tmp_side)
            for row in ws['A18:G18']:
                for cell in row:
                    cell.border = tmp_border

        cKaisya = ClsKaisya()
        kaisya_result = cKaisya.GetData(session)
        if len(kaisya_result) == 0:
            ws['K9'].value = ""
            ws['K10'].value = ""
            ws['K11'].value = ""
            ws['K12'].value = ""
        else:
            ws['K9'].value = "登録番号 {}".format(kaisya_result[0]["インボイス登録番号"])
            ws["K9"].font = Font(name='ＭＳ ゴシック')
            ws['K10'].value = "〒{}".format(kaisya_result[0]["郵便番号"])
            ws['K11'].value = kaisya_result[0]["住所1"]
            ws['K12'].value = "TEL {}(代)".format(kaisya_result[0]["電話番号"])
            ws['K13'].value = "FAX {}".format(kaisya_result[0]["FAX番号"])

        return ws

    def delete(self,session,request):
        estimate_no = request.params["@i見積番号"]
        select_query = "select 見積番号,請求日付 from TD売上前請求H where 見積番号 = {}".format(estimate_no)
        result = dict(SQLExecutor(session).execute_query(query=select_query))

        if len(result["results"]) == 0:
            raise ServiceError('指定の見積番号は存在しません。')

        cDates = ClsDates()
        cDates.GetbyId(session,"売掛月次更新日")
        gGetuDate = cDates.更新日付

        if result["results"][0]["請求日付"] <= gGetuDate:
            raise ServiceError('更新済みの為、修正できません。')

        delete_query = "delete from TD売上前請求H where 見積番号 = {}".format(estimate_no)
        result = dict(SQLExecutor(session).execute_query(query=delete_query))

        return JSONResponse(
            content={},
            headers={
                'Download-Skip': "yes"
            }
        )
