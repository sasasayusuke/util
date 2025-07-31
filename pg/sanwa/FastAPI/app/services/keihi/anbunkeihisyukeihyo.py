import datetime
from typing import Dict, Any
import io
import os
import logging
from copy import copy
from fastapi.responses import StreamingResponse
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.string_utils import encode_string, format_jp_date
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    execute_vba
)
from app.core.config import settings
from app.core.exceptions import ServiceError
import openpyxl

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class anbunKeihiSyukeiHyoService(BaseService):
    """按分経費集計表サービス"""

    def display(self, request, session) -> Dict[str, Any]:

        # 按分経費集計表の印刷処理
        logger.info("【START】按分経費集計表")

        storedname="usp_KD0800チーム別按分経費集計表出力V5"

        params = {
            "@iS集計日付":request.params["@iS集計日付"],
            "@iE集計日付":request.params["@iE集計日付"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetDcnt":"INT", "@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')



        with get_excel_buffer() as buffer:
            wb = create_excel_object('Tempチーム別按分経費一覧表v5.xlsx')

            ws = wb["Data"]

            # 作成日
            ws['G1'] = request.params["@集計年月"]

            # 初期位置指定
            start_row = 2
            current_row = 0

            for row_num, row_data in enumerate(storedresults['results'][0], start=start_row):
                ws.cell(row=row_num, column=1, value=row_data['集計CD'])._style = copy(ws.cell(row=start_row, column=1)._style)
                ws.cell(row=row_num, column=2, value=row_data['科目CD'])._style = copy(ws.cell(row=start_row, column=2)._style)
                ws.cell(row=row_num, column=3, value=row_data['科目名'])._style = copy(ws.cell(row=start_row, column=3)._style)
                ws.cell(row=row_num, column=4, value=row_data['按分金額'])._style = copy(ws.cell(row=start_row, column=4)._style)
                ws.cell(row=row_num, column=5, value=row_data['業種区分'])._style = copy(ws.cell(row=start_row, column=5)._style)
                ws.cell(row=row_num, column=6, value=row_data['業種名'])._style = copy(ws.cell(row=start_row, column=6)._style)

                current_row = row_num

                ws.row_dimensions[row_num].height = ws.row_dimensions[start_row].height

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            ws = wb["ピボット"]
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # ピボットテーブルの範囲変更
            # sheet = wb["ピボット"]
            # pivot_table = sheet.PivotTables("ピボットテーブル2")
            # pivot_table.set_range("A1:F{}".format(current_row))

            wb = execute_vba(wb,'a',"pivot_keisan.xlsm",['ピボット','ピボットテーブル2','Data',current_row])

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, 'osawa')

            filename = "按分"

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    # 'Content-Disposition': f'attachment; filename="{encode_string(filename)}.xlsx"'
                    'Content-Disposition': f'attachment; filename="test.xlsx"'
                }
            )




    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
