import io
import os
import logging
import copy
import html
import datetime
import tempfile
import openpyxl
import pdfkit
import uuid  # Excel/PDF出力に使用
from zipfile import ZipFile
import xml.etree.ElementTree as ET
from contextlib import contextmanager
import pythoncom
import win32com.client as win32
from openpyxl import Workbook, workbook, worksheet, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill, Alignment

from app.core.config import settings
from app.core.exceptions import ServiceError, ValidationError
from app.utils.string_utils import convert_cell_to_rc

from app.utils.db_utils import SQLExecutor  # 前回出力情報登録に使用
from openpyxl.worksheet.worksheet import Worksheet  # Excel/PDF出力に使用
from copy import copy  # Excel/PDF出力に使用
from fastapi.responses import StreamingResponse  # Excel/PDF出力に使用
from win32com.client import Dispatch  # Excel/PDF出力に使用
from io import BytesIO  # Excel/PDF出力に使用


# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)


@contextmanager
def get_excel_buffer():
    """Excel用のバッファを提供するコンテキストマネージャ"""
    buffer = io.BytesIO()
    try:
        yield buffer
    finally:
        buffer.close()


def create_excel_object(template_file: str = None):
    """
    Excelオブジェクトを作成する関数（xlsx形式のみ対応）

    Args:
        template_file (str, optional): テンプレートファイル名

    Returns:
        Excelオブジェクト
    """
    # テンプレートパスが指定されている場合、拡張子を確認
    if template_file:
        if not template_file.lower().endswith('.xlsx'):
            raise ValidationError("テンプレートファイルは 'xlsx' 形式である必要があります")
        template_path = os.path.join('./app/templates/xlsx', template_file)
        wb = load_workbook(template_path)
    else:
        wb = Workbook()

    return wb


def save_excel_to_buffer(buffer: io.BytesIO, wb: Workbook, password: str = None):
    """
    Excelオブジェクトをバッファに保存する関数
    引数passwordが入力されている場合、password付きで保存を試みるが、失敗した場合passwordなしで保存

    Args:
        wb: 保存するExcelオブジェクト
        buffer (io.BytesIO): 保存先のバッファ
        password (str, optional): パスワードを設定する

    Raises:
        ServiceError: Excel起動に関するエラー
        IOError: ファイル操作に関するエラー
    """
    logger.info(f"Excelファイルの保存処理を開始します:")

    tmp_file_path = None

    try:
        # 一時ファイルにExcelを保存
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            tmp_file_path = tmp_file.name
            try:
                wb.save(tmp_file_path)
            except Exception as e:
                raise ServiceError(f"Excelファイルの保存に失敗しました: {str(e)}")

        # パスワードの設定を試行
        if password:
            try:
                # COM環境の初期化
                pythoncom.CoInitialize()
                logger.info(f"Excel Applicationの起動を開始します。{settings.get_env()}")
                excel = win32.Dispatch("Excel.Application")
                excel.DisplayAlerts = False
                logger.info("Excel Applicationの起動が完了しました")
                try:
                    # ダイアログを無効化
                    excel.DisplayAlerts = False
                    wb_com = excel.Workbooks.Open(tmp_file_path)
                    wb_com.Password = password
                    wb_com.SaveAs(tmp_file_path, Password=password)
                    wb_com.Close()
                finally:
                    pass
                    # Excelアプリケーションの終了
                    # excel.Quit()
            except Exception as e:
                logger.warning(f"パスワード設定に失敗しました。パスワードなしで保存します: {str(e)}")
                try:
                    # パスワード設定に失敗した場合、パスワードなしで保存
                    wb_com = excel.Workbooks.Open(tmp_file_path)
                    wb_com.SaveAs(tmp_file_path)
                    wb_com.Close()
                except Exception as e:
                    raise ServiceError(f"パスワードなしでの保存にも失敗しました: {str(e)}")
            finally:
                # COM環境のクリーンアップ
                pythoncom.CoUninitialize()

        # バッファに一時ファイルから読み込み
        try:
            with open(tmp_file_path, "rb") as f:
                buffer.write(f.read())
            buffer.seek(0)
        except IOError as e:
            raise IOError(f"バッファへの書き込みに失敗しました: {str(e)}")

    finally:
        # 一時ファイルの削除
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.remove(tmp_file_path)
            except Exception as e:
                logger.warning(f"一時ファイルの削除に失敗しました: {str(e)}")


def range_copy_cell(
    src_sheet:  worksheet,
    dst_sheet:  worksheet,
    min_col:    int,
    min_row:    int,
    max_col:    int,
    max_row:    int,
    shift_col:  int,
    shift_row:  int,
    val_copy:   bool = True
) -> None:
    """
    Excelシートの指定範囲のセルをコピーする関数

    Args:
        src_sheet: コピー元のワークシート
        src_sheet: コピー先のワークシート
        min_col: コピー元の最小列番号
        min_row: コピー元の最小行番号
        max_col: コピー元の最大列番号
        max_row: コピー元の最大行番号
        shift_col: コピー先の列のオフセット
        shift_row: コピー先の行のオフセット
        val_copy: 値をコピーするか(デフォルト: True)
    """

    # コピー元の結合セルを深いコピーで保存
    merged_cells = list(src_sheet.merged_cells)

    # 指定範囲内の結合セルを解除
    cells_to_unmerge = [
        merged_cell
        for merged_cell in merged_cells
        if (min_row <= merged_cell.min_row <= max_row and min_col <= merged_cell.min_col <= max_col)
    ]
    for merged_cell in cells_to_unmerge:
        src_sheet.unmerge_cells(merged_cell.coord)

    # セルのコピー
    for col in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            # コピー元のセル座標を作成
            src_coord = f"{get_column_letter(col)}{row}"
            # コピー先のセル座標を作成
            dst_coord = f"{get_column_letter(col + shift_col)}{row + shift_row}"
            # セルの値をコピー
            if val_copy:
                if not isinstance(src_sheet[src_coord], MergedCell):
                    dst_sheet[dst_coord].value = src_sheet[src_coord].value
            # スタイルがある場合はコピー
            if src_sheet[src_coord].has_style:
                dst_sheet[dst_coord]._style = src_sheet[src_coord]._style

            dst_sheet.row_dimensions[row +
                                     shift_row].height = src_sheet.row_dimensions[row].height

    # セルを再結合
    for merged_cell in cells_to_unmerge:
        # コピー先の結合範囲を計算
        new_min_col = merged_cell.min_col + shift_col
        new_max_col = merged_cell.max_col + shift_col
        new_min_row = merged_cell.min_row + shift_row
        new_max_row = merged_cell.max_row + shift_row
        # コピー先の座標文字列を生成
        new_coord = f"{get_column_letter(new_min_col)}{new_min_row}:{get_column_letter(new_max_col)}{new_max_row}"
        # コピー先の位置を結合
        dst_sheet.merge_cells(new_coord)

        # コピー元の結合範囲を計算
        old_min_col = merged_cell.min_col
        old_max_col = merged_cell.max_col
        old_min_row = merged_cell.min_row
        old_max_row = merged_cell.max_row
        # コピー元の座標文字列を生成
        old_coord = f"{get_column_letter(old_min_col)}{old_min_row}:{get_column_letter(old_max_col)}{old_max_row}"
        # コピー元の位置を結合
        src_sheet.merge_cells(old_coord)


def range_copy_cell_by_address(
    src_sheet:      worksheet,
    dst_sheet:      worksheet,
    start_cell:     str,
    end_cell:       str,
    target_cell:    str,
    val_copy:       bool = True
) -> None:
    """
    Excelシートの指定範囲のセルをコピーする関数(セルアドレスで指定する)

    Args:
        src_sheet: コピー元のワークシート
        dst_sheet: コピー先のワークシート
        start_cell: コピー範囲の開始セル(例：'B3')
        end_cell: コピー範囲の終了セル(例：'D5')
        target_cell: コピー先の開始セル(例：'E1')
        val_copy: 値をコピーするか(デフォルト: True)
    """
    # 開始セルと終了セルを行と列の番号に変換
    start_row, start_col = convert_cell_to_rc(start_cell)
    end_row, end_col = convert_cell_to_rc(end_cell)
    target_row, target_col = convert_cell_to_rc(target_cell)

    # シフト量を計算
    shift_row = target_row - start_row
    shift_col = target_col - start_col

    # 既存の range_copy_cell 関数を呼び出し
    range_copy_cell(
        src_sheet=src_sheet,
        dst_sheet=dst_sheet,
        min_col=start_col,
        min_row=start_row,
        max_col=end_col,
        max_row=end_row,
        shift_col=shift_col,
        shift_row=shift_row,
        val_copy=val_copy
    )


def relative_reference_copy(
    src_sheet:  worksheet,
    dst_sheet:  worksheet,
    col:    int,
    row:    int,
    shift_col:  int,
    shift_row:  int
) -> None:
    """
    Excelシートの指定セルの数式を相対参照コピーする関数

    Args:
        src_sheet: コピー元のワークシート
        src_sheet: コピー先のワークシート
        col: コピー元の列番号
        row: コピー元の行番号
        shift_col: コピー先の列のオフセット
        shift_row: コピー先の行のオフセット
    """
    dst_sheet.cell(row=row+shift_row, column=col+shift_col, value=Translator(src_sheet['{get_col}{get_row}'.format(get_col=get_column_letter(col), get_row=row)].value, origin='{get_col}{get_row}'.format(
        get_col=get_column_letter(col), get_row=row)).translate_formula('{get_col}{get_row}'.format(get_row=row+shift_row, get_col=get_column_letter(col+shift_col))))


def relative_reference_range_copy(
    src_sheet:  worksheet,
    dst_sheet:  worksheet,
    min_col:    int,
    min_row:    int,
    max_col:    int,
    max_row:    int,
    shift_col:  int,
    shift_row:  int,
    val_copy:   bool = True
) -> None:
    """
    Excelシートの指定範囲のセルの数式をコピーする関数

    Args:
        src_sheet: コピー元のワークシート
        dst_sheet: コピー先のワークシート
        min_col: コピー元の最小列番号
        min_row: コピー元の最小行番号
        max_col: コピー元の最大列番号
        max_row: コピー元の最大行番号
        shift_col: コピー先の列のオフセット
        shift_row: コピー先の行のオフセット
        val_copy: 値をコピーするか(デフォルト: True)
    """

    # コピー元の結合セルを深いコピーで保存
    merged_cells = list(src_sheet.merged_cells)

    # 指定範囲内の結合セルを解除
    cells_to_unmerge = [
        merged_cell
        for merged_cell in merged_cells
        if (min_row <= merged_cell.min_row <= max_row and min_col <= merged_cell.min_col <= max_col)
    ]
    for merged_cell in cells_to_unmerge:
        src_sheet.unmerge_cells(merged_cell.coord)

    for col in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            dst_sheet.cell(row=row+shift_row, column=col+shift_col, value=Translator(src_sheet['{get_col}{get_row}'.format(get_col=get_column_letter(col), get_row=row)].value, origin='{get_col}{get_row}'.format(
                get_col=get_column_letter(col), get_row=row)).translate_formula('{get_col}{get_row}'.format(get_row=row+shift_row, get_col=get_column_letter(col+shift_col))))

    # セルを再結合
    for merged_cell in cells_to_unmerge:
        # コピー先の結合範囲を計算
        new_min_col = merged_cell.min_col + shift_col
        new_max_col = merged_cell.max_col + shift_col
        new_min_row = merged_cell.min_row + shift_row
        new_max_row = merged_cell.max_row + shift_row
        # コピー先の座標文字列を生成
        new_coord = f"{get_column_letter(new_min_col)}{new_min_row}:{get_column_letter(new_max_col)}{new_max_row}"
        # コピー先の位置を結合
        dst_sheet.merge_cells(new_coord)

        # コピー元の結合範囲を計算
        old_min_col = merged_cell.min_col
        old_max_col = merged_cell.max_col
        old_min_row = merged_cell.min_row
        old_max_row = merged_cell.max_row
        # コピー元の座標文字列を生成
        old_coord = f"{get_column_letter(old_min_col)}{old_min_row}:{get_column_letter(old_max_col)}{old_max_row}"
        # コピー元の位置を結合
        src_sheet.merge_cells(old_coord)


def delete_row_with_merged_ranges(sheet, idx) -> None:
    """
    「結合セル」を含む行を安全に削除する
        1. 削除対象行の結合セルの列範囲を取得
        2. 結合セルの解除
        3. 行の削除

    Args:
        sheet: 操作対象のワークシート
        idx: 削除する行のインデックス

    """

    # セル結合されている行を削除する
    begin_col, end_col = [], []
    # 削除対象行の結合セルの列情報を取得
    for n in sheet.merged_cells:
        if (n.min_row == idx and n.max_row == idx):
            begin_col.append(n.min_col)
            end_col.append(n.max_col)
    # 削除対象行のセル結合を解除
    for n in range(0, len(begin_col), 1):
        sheet.unmerge_cells(
            start_row=idx,
            start_column=begin_col[n],
            end_row=idx,
            end_column=end_col[n]
        )
    sheet.delete_rows(idx)
    for mcr in sheet.merged_cells:
        if idx < mcr.min_row:
            mcr.shift(row_shift=-1)
        elif idx <= mcr.max_row:
            mcr.shrink(bottom=1)


def insert_image(
    ws:         worksheet,
    img_file:   str,
    height:     int,
    width:      int,
    address:    str
) -> None:
    """
    ワークシートに画像を挿入する関数

    Args:
        ws:         挿入したいワークシート
        img_file:   画像ファイル名
        height:     画像の高さ
        width:      画像の幅
        address:    挿入するセル番地
    """
    img_path = os.path.join('./app/templates/img', img_file)
    img = Image(img_path)
    img.width, img.height = width, height
    ws.add_image(img, address)


def execute_vba(
    wb: Workbook,
    function_name: str,
    vba_filename: str,
    params: list
):
    """
    外部ファイルのマクロを実行する関数

    Args:
        wb              :openpyxlのワークブックオブジェクト
        function_name   :マクロの関数名
        vba_filename    :マクロが登録されているファイル名
        params          :マクロに渡す引数

    return:
        openpyxlのワークブックオブジェクト
    """
    # excelのライセンスがない場合はopenpyxlのwbオブジェクトを返却
    if not settings.EXCEL_LICENSE:
        return wb
    # COM環境の初期化
    pythoncom.CoInitialize()

    try:
        # 1.ワークブックをtmpフォルダに保存する
        dt = datetime.datetime.now()
        # tmpファイル名を一意にするためにミリ秒を付与
        millisecound = int(dt.timestamp())*1000
        filename = "temp{}.xlsx".format(millisecound)
        excel_path = os.path.join(os.getenv('TEMP'), filename)
        wb.save(excel_path)

        # excel起動
        logger.info(f"Excel Applicationの起動を開始します。{settings.get_env()}")
        excel = win32.Dispatch("Excel.Application")
        excel.visible = False
        excel.DisplayAlerts = False
        logger.info("Excel Applicationの起動が完了しました")

        # excelファイルを開く
        target_wb = excel.Workbooks.Open(excel_path)
        vba_wb = excel.Workbooks.Open(os.path.abspath(
            './app/templates/xlsm/{}'.format(vba_filename)))

        # マクロ実行
        tmp_params = [filename, *params]
        excel.Application.Run("{}!{}".format(
            vba_wb.Name, function_name), *tmp_params)

        # 3．保存して閉じる
        target_wb.Close(SaveChanges=True)
        vba_wb.Close(SaveChanges=False)

        # 4. openpyxlでロードしてwbを返す
        res_wb = openpyxl.load_workbook(excel_path)
        return res_wb

    finally:
        # excel.Quit()
        if os.path.exists(excel_path):
            os.remove(excel_path)


def set_footer_whitespace(wb: Workbook):
    """
    Excelブックの全シートのフッターの「余白に合わせて配置」と「拡大縮小に合わせて表示」する関数
    XMLを直接編集してフッターの設定

    Args:
        wb  :openpyxlのワークブックオブジェクト

    Returns:
        設定を変更したワークブックオブジェクト
    """
    try:
        sheets_count = len(wb.sheetnames)
        dt = datetime.datetime.now()
        millisecound = int(dt.timestamp())*1000
        filename = "temp{}.xlsx".format(millisecound)
        excel_path = os.path.join(os.getenv('TEMP'), filename)
        wb.save(excel_path)

        with ZipFile(excel_path, "a") as zf:
            for i in range(1, sheets_count+1):
                with zf.open("xl/worksheets/sheet{}.xml".format(i)) as sheet:
                    tree = ET.parse(sheet)
                    root = tree.getroot()

                    for headerFooter in root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}headerFooter"):
                        headerFooter.set("alignWithMargins", "0")
                        headerFooter.set("scaleWithDoc", "0")

                    # ファイルに書き戻し
                    with zf.open("xl/worksheets/sheet{}.xml".format(i), "w") as sheet:
                        sheet.write(ET.tostring(root))

        res_wb = load_workbook(excel_path)
        return res_wb
    finally:
        if os.path.exists(excel_path):
            os.remove(excel_path)


class ExcelPDFConverter:
    def __init__(self, wb: Workbook, ws: worksheet, start_cell: str, end_cell: str):
        """
        ExcelワークシートからPDFを生成するコンバーター

        Args:
            ws: 対象のワークシート
            start_cell: 開始セル（例: "A1"）
            end_cell: 終了セル（例: "X30"）
        """
        self.wb = wb
        self.ws = ws
        self.start_cell = start_cell
        self.end_cell = end_cell
        self.start_row, self.start_col = convert_cell_to_rc(self.start_cell)
        self.end_row, self.end_col = convert_cell_to_rc(self.end_cell)
        self.licence = settings.EXCEL_LICENSE

    def _generate_html(self) -> str:
        """HTMLテーブルを生成"""

        # スタイル定義
        style = """
            <style>
                table { border-collapse: collapse; width: 100%; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                tr:nth-child(even) { background-color: #f2f2f2; }
            </style>
        """

        # テーブル生成
        html_content = ["<table>"]

        for row in range(self.start_row, self.end_row + 1):
            html_content.append("<tr>")
            for col in range(self.start_col, self.end_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell_value = html.escape(
                    str(cell.value)) if cell.value is not None else "&nbsp;"

                # セルの書式設定を反映
                style_attrs = []
                if cell.font.bold:
                    style_attrs.append("font-weight: bold")
                if cell.alignment.horizontal:
                    style_attrs.append(
                        f"text-align: {cell.alignment.horizontal}")

                style_str = f' style="{"; ".join(style_attrs)}"' if style_attrs else ""
                html_content.append(f"<td{style_str}>{cell_value}</td>")
            html_content.append("</tr>")

        html_content.append("</table>")

        return f"{style}{''.join(html_content)}"

    def generate_pdf(self) -> bytes:
        """PDFを生成"""
        logger.info("PDFファイルの作成処理を開始します")

        # ライセンスあり
        if self.licence:
            # COM初期化
            pythoncom.CoInitialize()

            # Excelアプリケーションを起動
            logger.info(f"Excel Applicationの起動を開始します。{settings.get_env()}")
            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False
            logger.info("Excel Applicationの起動が完了しました")

            excel_path = None
            pdf_path = None
            try:
                # 一時ExcelファイルおよびPDFファイルを作成
                excel_path = os.path.join(os.getenv('TEMP'), 'temp.xlsx')
                pdf_path = os.path.join(os.getenv('TEMP'), 'output.pdf')

                self.wb.save(excel_path)
                wb = excel.Workbooks.Open(excel_path)
                ws = wb.Worksheets[self.ws.title]
                ws.PageSetup.PrintArea = f"{self.start_cell}:{self.end_cell}"
                ws.ExportAsFixedFormat(0, pdf_path)
                wb.Close(False)

                with open(pdf_path, 'rb') as f:
                    pdf = f.read()

            finally:
                # excel.Quit()
                for path in [excel_path, pdf_path]:
                    if path and os.path.exists(path):
                        try:
                            os.remove(path)
                        except:
                            pass
                # COM解放
                pythoncom.CoUninitialize()
        else:
            # 一度HTMLへ変換
            html_content = self._generate_html()

            # Windows環境でのwkhtmltopdfのパス指定
            config = pdfkit.configuration(
                wkhtmltopdf=settings.WKHTMLTOPDF_PATH)

            # PDF設定
            options = {
                'page-size': 'A4',
                'margin-top': '20mm',
                'margin-right': '20mm',
                'margin-bottom': '20mm',
                'margin-left': '20mm',
                'encoding': 'UTF-8',
            }

            # HTMLからPDFを生成（configurationを指定）
            pdf = pdfkit.from_string(
                html_content, False, options=options, configuration=config)
        return pdf


def delete_excel_object(ws, start_row, start_col, end_row, end_col):
    """
    指定した範囲のセルを削除（クリア）する関数

    Parameters:
    ws (Worksheet): 対象のワークシート
    start_row (int): 範囲の開始行番号
    start_col (int): 範囲の開始列番号（1ベース）
    end_row (int): 範囲の終了行番号
    end_col (int): 範囲の終了列番号（1ベース）
    """

    # 結合されているセルの解除処理（範囲内のみ）
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        if (
            min_row >= start_row and max_row <= end_row and
            min_col >= start_col and max_col <= end_col
        ):
            # 指定範囲内に完全に含まれる結合セルのみ解除
            ws.unmerge_cells(str(merged_range))

    # 範囲内のセルをクリア
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None  # セルの値をクリア
            cell.fill = PatternFill(fill_type=None)  # セルの背景色をクリア
            cell.border = None  # セルのボーダーをクリア
            cell.font = None  # セルのフォントをクリア
            cell.alignment = None  # セルの配置をクリア


def copy_images_to_new_sheet(template_sheet, new_sheet):
    """
    テンプレートシートの画像を新しいシートにコピーする関数。

    Parameters:
    template_sheet (Worksheet): 画像をコピー元となるテンプレートのワークシート。
    new_sheet (Worksheet): 画像をコピー先となる新しいワークシート。

    Returns:
    None
        新しいシートに画像を挿入する処理のみを行い、特に返り値はありません。
    """
    for image in template_sheet._images:  # テンプレートシート内の画像をループ
        img = Image(image.ref)  # Imageオブジェクトを作成
        img.anchor = image.anchor  # 画像の位置情報を維持
        new_sheet.add_image(img)  # 新しいシートに画像を挿入


def register_history(session, list_name, estimate_no, login_id):
    """
    履歴情報を登録する共通関数。

    指定された履歴情報（リスト名、見積番号、PC名）を使用して、
    `usp_見積出力履歴更新` ストアプロシージャを実行し、履歴をデータベースに登録します。

    Parameters:
    session (Session): データベースのセッション。ストアプロシージャを実行するためのセッションオブジェクト。
    list_name (str): 履歴に登録するリストの名前。履歴の対象となるリストを識別するための名前。
    estimate_no (str): 見積番号。履歴に関連付ける見積もりの番号。
    login_id (str): ログインID。履歴を登録したユーザーを識別するためのID。

    Returns:
    None
        履歴登録が成功した場合、特に返り値はありません。
        ストアプロシージャの実行結果は、データベースに反映されます。

    Raises:
    ServiceError: ストアプロシージャの実行に失敗した場合や、エラーメッセージが返された場合に発生するエラー。
    """
    storedname = "usp_見積出力履歴更新"

    # パラメータの設定
    params = {
        "@iListName": list_name,
        "@iNumber": estimate_no,
        "@iPCName": login_id,
    }

    # 履歴更新ストアプロシージャの実行
    SQLExecutor(session).execute_stored_procedure(
        storedname=storedname, params=params, outputparams={"@RetST": "INT", "@RetMsg": "VARCHAR(255)"})


def copy_sheet_settings(template_sheet: Worksheet, new_sheet: Worksheet):
    """ページ設定、ヘッダー・フッター、印刷タイトル、余白、セルスタイルを完全に引き継ぐ"""

    # ページ設定の引き継ぎ
    new_sheet.page_setup = copy(template_sheet.page_setup)
    new_sheet.page_margins = copy(template_sheet.page_margins)
    new_sheet.print_options = copy(template_sheet.print_options)

    # 具体的なページ設定の引き継ぎ
    new_sheet.page_setup.orientation = copy(
        template_sheet.page_setup.orientation)
    new_sheet.page_setup.paperSize = copy(template_sheet.page_setup.paperSize)
    new_sheet.page_setup.fitToHeight = copy(
        template_sheet.page_setup.fitToHeight)
    new_sheet.page_setup.fitToWidth = copy(
        template_sheet.page_setup.fitToWidth)
    new_sheet.page_setup.scale = copy(template_sheet.page_setup.scale)
    new_sheet.page_setup.draft = copy(template_sheet.page_setup.draft)
    new_sheet.page_setup.horizontalDpi = copy(
        template_sheet.page_setup.horizontalDpi)
    new_sheet.page_setup.verticalDpi = copy(
        template_sheet.page_setup.verticalDpi)

    # 余白の詳細設定
    new_sheet.page_margins.left = copy(template_sheet.page_margins.left)
    new_sheet.page_margins.right = copy(template_sheet.page_margins.right)
    new_sheet.page_margins.top = copy(template_sheet.page_margins.top)
    new_sheet.page_margins.bottom = copy(template_sheet.page_margins.bottom)
    new_sheet.page_margins.header = copy(template_sheet.page_margins.header)
    new_sheet.page_margins.footer = copy(template_sheet.page_margins.footer)

    # ヘッダー・フッターの引き継ぎ
    if hasattr(template_sheet, "oddHeader") and hasattr(template_sheet, "oddFooter"):
        new_sheet.oddHeader.left.text = template_sheet.oddHeader.left.text
        new_sheet.oddHeader.center.text = template_sheet.oddHeader.center.text
        new_sheet.oddHeader.right.text = template_sheet.oddHeader.right.text

        new_sheet.oddFooter.left.text = template_sheet.oddFooter.left.text
        new_sheet.oddFooter.center.text = template_sheet.oddFooter.center.text
        new_sheet.oddFooter.right.text = template_sheet.oddFooter.right.text

    # 印刷タイトルの引き継ぎ
    if template_sheet.print_title_rows:
        new_sheet.print_title_rows = template_sheet.print_title_rows
    if template_sheet.print_title_cols:
        new_sheet.print_title_cols = template_sheet.print_title_cols

    # セルスタイルの引き継ぎ（必要に応じて）
    for row in template_sheet.iter_rows():
        for cell in row:
            if cell.has_style:
                target_cell = new_sheet[cell.coordinate]
                target_cell.alignment = copy(cell.alignment)
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.number_format = copy(cell.number_format)

    # セルの高さと列の幅の引き継ぎ
    for row_idx in range(1, template_sheet.max_row + 1):
        new_sheet.row_dimensions[row_idx].height = template_sheet.row_dimensions[row_idx].height

    for col_idx in template_sheet.column_dimensions:
        new_sheet.column_dimensions[col_idx].width = template_sheet.column_dimensions[col_idx].width


def file_generate_response(buffer, wb, output_format):
    """ExcelまたはPDFレスポンスを生成"""
    try:
        if output_format == 'excel':
            # Excelレスポンスを生成
            save_excel_to_buffer(buffer, wb, None)
            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

        elif output_format == 'pdf':
            # PDFレスポンスを生成
            tmp_file_path = os.path.join(
                tempfile.gettempdir(), f"temp_{uuid.uuid4().hex}.xlsx")
            pdf_file_path = tmp_file_path.replace('.xlsx', '.pdf')

            try:
                # Excelファイルを保存
                wb.save(tmp_file_path)

                # COM初期化とExcelアプリケーションの起動
                pythoncom.CoInitialize()
                excel = Dispatch("Excel.Application")
                excel.Visible = False

                # try:
                workbook = excel.Workbooks.Open(tmp_file_path)
                workbook.ExportAsFixedFormat(0, pdf_file_path)
                workbook.Close(False)
                # finally:
                #     excel.Quit()

                # PDFをレスポンスとして返却
                with open(pdf_file_path, 'rb') as pdf_file:
                    pdf_content = pdf_file.read()

                return StreamingResponse(BytesIO(pdf_content), media_type="application/pdf", headers={
                    'Content-Disposition': 'attachment; filename="output.pdf"'
                })

            finally:
                # 一時ファイルの削除
                for file_path in [tmp_file_path, pdf_file_path]:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                pythoncom.CoUninitialize()

        else:
            raise ValueError(
                "Invalid output format specified. Use 'excel' or 'pdf'.")

    except Exception as e:
        logger.error(f"レスポンス生成エラー: {e}")
        raise ServiceError("レスポンスの生成に失敗しました")
