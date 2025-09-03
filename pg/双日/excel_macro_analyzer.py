#!/usr/bin/env python3
"""
Excel マクロ分析ツール
双日ライフワンプロジェクトのExcelファイルからマクロとVBAコードを分析します。
"""

import os
import sys
import pandas as pd
from oletools import olevba
from oletools.olevba import VBA_Parser
import openpyxl
from openpyxl import load_workbook
import xlrd
from pathlib import Path
import json
import re

class ExcelMacroAnalyzer:
    def __init__(self, base_path):
        self.base_path = Path(base_path)
        self.results = {}
        
    def analyze_file(self, file_path):
        """指定されたExcelファイルを分析"""
        file_path = Path(file_path)
        result = {
            'file_name': file_path.name,
            'file_path': str(file_path),
            'file_size': file_path.stat().st_size if file_path.exists() else 0,
            'has_macros': False,
            'vba_modules': [],
            'worksheet_count': 0,
            'worksheet_names': [],
            'error': None
        }
        
        try:
            print(f"\n=== 分析中: {file_path.name} ===")
            
            # VBAマクロの検出
            if file_path.suffix.lower() in ['.xls', '.xlsm', '.xlsx']:
                result.update(self._analyze_vba_macros(file_path))
                result.update(self._analyze_worksheet_structure(file_path))
                
        except Exception as e:
            result['error'] = str(e)
            print(f"エラー: {e}")
            
        return result
    
    def _analyze_vba_macros(self, file_path):
        """VBAマクロの分析"""
        result = {
            'has_macros': False,
            'vba_modules': [],
            'macro_functions': [],
            'macro_complexity': 'None'
        }
        
        try:
            # oletoolsを使用してVBAマクロを分析
            vba_parser = VBA_Parser(str(file_path))
            
            if vba_parser.detect_vba_macros():
                result['has_macros'] = True
                print(f"✓ VBAマクロが検出されました")
                
                # 各VBAモジュールの分析
                for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                    if vba_code:
                        module_info = {
                            'module_name': vba_filename,
                            'code_lines': len(vba_code.split('\n')),
                            'functions': self._extract_vba_functions(vba_code),
                            'code_snippet': vba_code[:500] + '...' if len(vba_code) > 500 else vba_code
                        }
                        result['vba_modules'].append(module_info)
                        result['macro_functions'].extend(module_info['functions'])
                
                # 複雑性評価
                total_lines = sum(module['code_lines'] for module in result['vba_modules'])
                if total_lines > 100:
                    result['macro_complexity'] = 'High'
                elif total_lines > 50:
                    result['macro_complexity'] = 'Medium'
                else:
                    result['macro_complexity'] = 'Low'
                    
            else:
                print("✗ VBAマクロは検出されませんでした")
                
            vba_parser.close()
            
        except Exception as e:
            print(f"VBA分析エラー: {e}")
            
        return result
    
    def _extract_vba_functions(self, vba_code):
        """VBAコードから関数/サブルーチンを抽出"""
        functions = []
        
        # 関数とサブルーチンのパターン
        patterns = [
            r'(?i)^\s*(?:public\s+|private\s+)?(?:sub|function)\s+(\w+)\s*\(',
            r'(?i)^\s*(?:public\s+|private\s+)?(?:sub|function)\s+(\w+)\s*$'
        ]
        
        lines = vba_code.split('\n')
        for line in lines:
            for pattern in patterns:
                match = re.search(pattern, line)
                if match:
                    func_name = match.group(1)
                    if func_name not in functions:
                        functions.append(func_name)
        
        return functions
    
    def _analyze_worksheet_structure(self, file_path):
        """ワークシートの構造を分析"""
        result = {
            'worksheet_count': 0,
            'worksheet_names': [],
            'data_summary': {}
        }
        
        try:
            if file_path.suffix.lower() == '.xlsx':
                # openpyxlを使用(.xlsx)
                wb = load_workbook(str(file_path), data_only=True)
                result['worksheet_names'] = wb.sheetnames
                result['worksheet_count'] = len(wb.sheetnames)
                
                # 各シートの基本情報
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    result['data_summary'][sheet_name] = {
                        'max_row': ws.max_row,
                        'max_column': ws.max_column,
                        'has_data': ws.max_row > 1 or ws.max_column > 1
                    }
                
            elif file_path.suffix.lower() == '.xls':
                # xlrdを使用(.xls)
                wb = xlrd.open_workbook(str(file_path))
                result['worksheet_names'] = wb.sheet_names()
                result['worksheet_count'] = len(wb.sheet_names())
                
                # 各シートの基本情報
                for sheet_name in wb.sheet_names():
                    ws = wb.sheet_by_name(sheet_name)
                    result['data_summary'][sheet_name] = {
                        'max_row': ws.nrows,
                        'max_column': ws.ncols,
                        'has_data': ws.nrows > 1 or ws.ncols > 1
                    }
                    
        except Exception as e:
            print(f"ワークシート分析エラー: {e}")
            
        return result
    
    def generate_report(self, results):
        """分析結果のレポートを生成"""
        print("\n" + "="*80)
        print("Excel マクロ分析レポート")
        print("="*80)
        
        macro_files = []
        non_macro_files = []
        
        for file_name, result in results.items():
            if result.get('has_macros', False):
                macro_files.append(result)
            else:
                non_macro_files.append(result)
        
        print(f"\n📊 分析結果サマリー:")
        print(f"  - 総ファイル数: {len(results)}")
        print(f"  - マクロ有りファイル: {len(macro_files)}")
        print(f"  - マクロ無しファイル: {len(non_macro_files)}")
        
        if macro_files:
            print(f"\n🔍 マクロが含まれるファイル:")
            for result in macro_files:
                print(f"\n  📄 {result['file_name']}")
                print(f"     - ファイルサイズ: {result['file_size']:,} bytes")
                print(f"     - ワークシート数: {result['worksheet_count']}")
                print(f"     - VBAモジュール数: {len(result['vba_modules'])}")
                print(f"     - マクロ複雑度: {result['macro_complexity']}")
                
                if result['macro_functions']:
                    print(f"     - 関数/サブルーチン: {', '.join(result['macro_functions'])}")
                
                print(f"     - ワークシート: {', '.join(result['worksheet_names'])}")
                
                # VBAコードの詳細
                for module in result['vba_modules']:
                    print(f"\n       📋 モジュール: {module['module_name']}")
                    print(f"          - コード行数: {module['code_lines']}")
                    print(f"          - 関数数: {len(module['functions'])}")
                    if module['functions']:
                        print(f"          - 関数一覧: {', '.join(module['functions'])}")
                    
                    # コードの一部を表示
                    if module['code_snippet']:
                        print(f"          - コード抜粋:")
                        code_lines = module['code_snippet'].split('\n')[:10]
                        for i, line in enumerate(code_lines):
                            if line.strip():
                                print(f"            {i+1:2d}: {line}")
        
        if non_macro_files:
            print(f"\n📋 マクロが含まれないファイル:")
            for result in non_macro_files:
                print(f"  - {result['file_name']} (ワークシート: {result['worksheet_count']})")
        
        return {
            'macro_files': macro_files,
            'non_macro_files': non_macro_files,
            'total_files': len(results)
        }

def main():
    # 分析対象ファイルのリスト
    target_files = [
        '01_受領資料/管理変動費(202501).xls',
        '01_受領資料/管理変動費(202412).xls',
        '01_受領資料/管理変動費(202401).xls',
        '01_受領資料/計算用管理変動費Master.xls',
        '01_受領資料/MO報告光熱費.xls',
        '01_受領資料/CSVデータ/検針結果出力ファイル1_20250101_20250131.xlsx'
    ]
    
    base_path = '/home/sdt_op/projects/util/pg/soujitsu'
    analyzer = ExcelMacroAnalyzer(base_path)
    
    results = {}
    
    for target_file in target_files:
        file_path = Path(base_path) / target_file
        if file_path.exists():
            results[target_file] = analyzer.analyze_file(file_path)
        else:
            print(f"ファイルが見つかりません: {file_path}")
    
    # レポート生成
    report = analyzer.generate_report(results)
    
    # 結果をJSONファイルに保存
    output_file = Path(base_path) / 'excel_macro_analysis_results.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 詳細結果を保存しました: {output_file}")
    
    return results

if __name__ == "__main__":
    main()