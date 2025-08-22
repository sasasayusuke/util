from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image
)
from app.utils.service_utils import ClsTanto,ClsKaisya
from openpyxl.drawing.image import Image

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class simamuraTyokunoDenpyoSeikyusyoService(BaseService):
    """しまむら直納伝票請求書"""


    def display(self, request, session) -> Dict[str, Any]:

        # しまむら直納伝票請求書の印刷処理
        logger.info("【START】しまむら直納伝票請求書")

        storedname="usp_SE1600SM請求書発行_直納伝票"

        # params作成
        params = {
            "@iS請求日付":request.params["@iS請求日付"],
            "@iE請求日付":request.params["@iE請求日付"],
            "@i得意先CD":request.params["@i得意先CD"],
            "@i伝票種類":request.params["@i伝票種類"],
            "@iSM内容区分":request.params["@iSM内容区分"],
            "@i請求書発行日付":request.params["@i請求書発行日付"],
            "@i送料区分":request.params["@i送料区分"],
            "@i並び順区分":request.params["@i並び順区分"]
        }

        outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params,outputparams=outputparams))

        if len(storedresults["results"][0]) == 0 or len(storedresults['results'][1]) == 0:
            raise ServiceError('該当データなし')

        # 請求期間を日付オブジェクトに変換
        date_start = datetime.strptime(request.params['@iS請求日付'],'%Y-%m-%d')
        date_end = datetime.strptime(request.params['@iE請求日付'],'%Y-%m-%d')


        with get_excel_buffer() as buffer:
            wb = create_excel_object('Template_請求書_SM直納.xlsx')
            ws = wb["Template"]
            copy_ws = wb["Sheet1"]

            insert_image(ws,'三和ロゴ.png',90,300,'F5')

            # ヘッダー
            header_data = storedresults['results'][0][0]
            # 作成日
            ws['H1'].value = "作成日： {}".format(date_end.strftime('%Y{0}%m{1}%d{2}').format(*'年月日'))
            # 得意先
            if header_data["得意先名2"] == "":
                ws['A3'].value = ""
                ws['A4'].value = header_data["得意先名1"]
            else:
                ws['A3'].value = header_data["得意先名1"]
                ws['A4'].value = header_data["得意先名2"]
            # 見積金額
            ws['B8'].value = header_data['合計金額']
            # 消費税
            ws['B9'].value = header_data['外税額']
            # 見積件名
            if request.params['@i伝票種類'] == '2':
                ws['B12'].value = "直納伝票 {} 御請求分".format(date_end.strftime('%Y{0}%m{1}').format(*'年月'))
            elif request.params['@i伝票種類'] == '3':
                ws['B12'].value = "mail発注書(消耗品) {} 御請求分".format(date_end.strftime('%Y{0}%m{1}').format(*'年月'))
            elif request.params['@i伝票種類'] == '4':
                ws['B12'].value = "mail発注書(直納) {} 御請求分".format(date_end.strftime('%Y{0}%m{1}').format(*'年月'))
            # 送料
            if request.params['@i送料区分'] == '0':
                ws['B11'].value = ''
            elif request.params['@i送料区分'] == '1':
                ws['B11'].value = '送料'
            # 納期
            ws['B14'].value = "{} ～ {}".format(date_start.strftime('%Y{0}%m{1}%d{2}').format(*'年月日'),date_end.strftime('%Y{0}%m{1}%d{2}').format(*'年月日'))
            # 受渡地
            ws['B15'].value = "{} 様 各店".format(header_data["得意先名1"])
            # 支払条件
            ws['B16'].value = "{}日締　従来通り".format(date_end.strftime('%d'))
            # 備考
            ws['B17'].value = ""

            # 担当者
            cTanto = ClsTanto().GetbyID(session,header_data["担当者CD"])
            if len(cTanto) == 1:
                if cTanto["問い合わせ先"] != "":
                    ws["F16"].value = "担当：{}".format(cTanto["問い合わせ先"])
            else :
                ws['F16'].value = ""
            # E-mail
            ws["F17"].value = ""
            del cTanto # メモリ開放

            # インボイス番号
            cKaisya = ClsKaisya().GetData(session)
            if len(cKaisya) == 1 and cKaisya[0]["インボイス登録番号"] != "":
                ws['F9'].value = "登録番号 {}".format(cKaisya[0]["インボイス登録番号"])
            else:
                ws['F9'].value = ""

            # 書込行数
            row = 21

            # 頁カウント
            page_count = 1
            # ページごとの行カウント
            row_count_for_page = 1

            for row_data in storedresults['results'][1]:

                # 書き込み行
                if row == 55:
                    range_copy_cell_by_address(copy_ws,ws,'A1','H52','A{}'.format(row),True)
                    row = 58
                    page_count = page_count + 1
                    row_count_for_page = 1
                elif page_count > 1 and row_count_for_page == 50:
                    # コピー
                    # if page_count > 1:
                    range_copy_cell_by_address(copy_ws,ws,'A1','H52','A{}'.format(row),True)
                    # 頁と行の更新
                    page_count = page_count + 1
                    row_count_for_page = 1
                    row = row + 3


                if request.params["@i伝票種類"] == "2":
                    ws.cell(row=row, column=1, value=row_data['他社伝票番号'])    # 他社伝票番号
                ws.cell(row=row, column=2, value=row_data['納入先CD'])    # 納入先CD
                ws.cell(row=row, column=4, value=row_data['税抜金額'])   # 税抜金額
                ws.cell(row=row, column=5, value=row_data['見積番号'])    # 見積番号

                row = row + 1
                row_count_for_page = row_count_for_page + 1

            # 合計
            if row == 54 or row == 53:
                row = 58
                page_count = page_count + 1
                row_count_for_page = 1
            elif page_count > 1 and row_count_for_page == 49 or row_count_for_page == 50:
                # コピー
                range_copy_cell_by_address(copy_ws,ws,'A1','H52','A{}'.format(row),True)
                # 頁と行の更新
                page_count = page_count + 1
                row_count_for_page = 1
                row = row + 3
            ws.cell(row = row+1,column=1,value="【合　　計】")
            ws.cell(row = row+1,column=4,value="=SUM(D{}:D{})".format('21',row-1))


            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # コピー用シートを削除
            wb.remove(copy_ws)

            # シート名の変更
            ws.title = '請求書'

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )




    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass

# select MH.見積番号,USH.請求日付
# from TD見積 as MH
# inner join TD売上請求H as USH
# on MH.見積番号 = USH.見積番号
# where MH.得意先CD='3201'
# and MH.伝票種類=2
# and MH.SM内容区分=0
# order by MH.見積番号 DESC
