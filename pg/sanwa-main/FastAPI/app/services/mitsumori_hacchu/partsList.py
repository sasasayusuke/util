from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer, range_copy_cell
from app.utils.service_utils import ClsMitumoriH
from app.core.config import settings
from app.core.exceptions import ServiceError
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill, Font
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    get_column_letter,
    ExcelPDFConverter
)
from openpyxl.formatting.rule import FormulaRule

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class PartsListService(BaseService):
    """部材リストサービス"""

    def display(self, request, session) -> StreamingResponse:

        # 部材リストの印刷処理
        logger.info("【START】部材リスト処理")

        listName = "部材リスト"

        storedname="usp_MT0400部材リスト出力"

        params = {
            "@i見積番号":request.params["EstimateNo"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        header_results = storedresults['results'][0]
        results = storedresults['results'][1]

        if len(results) < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            workbook = create_excel_object('Template_部材リスト.xlsx')
            ws = workbook['Template']

            # 初期化
            start_row = 8
            current_row = start_row  # データ書き込み開始行
            init_flag = True

            # 表を何列拡張するか計算
            base_haeder_count = 18
            header_count = len(header_results)

            # if header_count % 3 != 0:
            #     header_count = ((header_count // 3)+1) * 3

            for i in range(1,header_count+1):
                range_copy_cell_by_address(ws,ws,'S7','S8','{}7'.format(get_column_letter(i+base_haeder_count)),False)
                logger.info('{}7'.format(get_column_letter(i+base_haeder_count)))

            # 表を拡張
            for i,header_result in enumerate(header_results, start=19):
                ws.cell(row=7, column=i, value=header_result['仕分番号'] if header_result['略称'] == "" else header_result['略称'])

            # データのループ
            for result in results:
                
                # 初回レコードの場合
                if init_flag:
                    # 先頭レコードからヘッダー情報を設定
                    ws.cell(row=3, column=3, value=result["見積番号"])
                    ws.cell(row=3, column=5, value=result["見積件名"])
                    ws.cell(row=4, column=3, value=result["得意先名1"])
                    ws.cell(row=4, column=7, value=datetime.now().strftime('%Y年%m月%d日'))
                    ws.cell(row=5, column=3, value=result["現場名"])
                    ws.cell(row=5, column=7, value=result["担当者名"])
                    init_flag = False
                else:
                    # データ行の書式をコピー
                    range_copy_cell_by_address(ws,ws,'A8','{}8'.format(get_column_letter(base_haeder_count+header_count)),'A{}'.format(current_row),False)
                
                # データ行を設定
                if (result["見積区分"] == "C") or (result["見積区分"] == "A") or (result["見積区分"] == "S"):
                    ws.cell(row=current_row, column=1, value="")
                else:
                    ws.cell(row=current_row, column=1, value=result["行番号"])
                ws.cell(row=current_row, column=2, value=result["製品NO"])
                ws.cell(row=current_row, column=3, value=result["仕様NO"])
                # 見積区分が C or A or S の場合は名称の先頭に半角スペースを2つ付ける
                if (result["見積区分"] == "C") or (result["見積区分"] == "A") or (result["見積区分"] == "S"):
                    ws.cell(row=current_row, column=4, value="  " + result["漢字名称"])
                else:
                    ws.cell(row=current_row, column=4, value=result["漢字名称"])
                if result["W"]:
                    ws.cell(row=current_row, column=6, value=result["W"])
                if result["D"]:
                    ws.cell(row=current_row, column=7, value=result["D"])
                if result["H"]:
                    ws.cell(row=current_row, column=8, value=result["H"])
                ws.cell(row=current_row, column=9, value=result["明細備考"])
                ws.cell(row=current_row, column=10, value=result["仕入先CD"])
                ws.cell(row=current_row, column=11, value=result["仕入先名"])
                ws.cell(row=current_row, column=12, value=result["配送先CD"])
                ws.cell(row=current_row, column=13, value=result["配送先名"])
                # 使用量合計-客先在庫数-転用の関数を設定
                ws.cell(row=current_row, column=14, value=f'=IF(ISERROR(R{current_row}-P{current_row}-Q{current_row}),"",R{current_row}-P{current_row}-Q{current_row})')
                ws.cell(row=current_row, column=15, value=result["単位名"])
                ws.cell(row=current_row, column=16, value=result["客先在庫数"] if result["客先在庫数"] !=0 else "")
                ws.cell(row=current_row, column=17, value=result["転用数"] if result["転用数"] !=0 else "")
                # 仕分数1+仕分数2の関数を設定
                ws.cell(row=current_row, column=18, value=f'=IF(SUM(S{current_row}:{get_column_letter(base_haeder_count+header_count)}{current_row})=0,0,SUM(S{current_row}:{get_column_letter(base_haeder_count+header_count)}{current_row}))')
                if '仕分数1' in result:
                    ws.cell(row=current_row, column=19, value=result["仕分数1"] if result["仕分数1"] !=0 else "")
                if '仕分数2' in result:
                    ws.cell(row=current_row, column=20, value=result["仕分数2"] if result["仕分数2"] !=0 else "")
                if '仕分数3' in result:
                    ws.cell(row=current_row, column=21, value=result["仕分数3"] if result["仕分数3"] !=0 else "")
                if '仕分数4' in result:
                    ws.cell(row=current_row, column=22, value=result["仕分数4"] if result["仕分数4"] !=0 else "")
                if '仕分数5' in result:
                    ws.cell(row=current_row, column=23, value=result["仕分数5"] if result["仕分数5"] !=0 else "")
                if '仕分数6' in result:
                    ws.cell(row=current_row, column=24, value=result["仕分数6"] if result["仕分数6"] !=0 else "")
                if '仕分数7' in result:
                    ws.cell(row=current_row, column=25, value=result["仕分数7"] if result["仕分数7"] !=0 else "")
                if '仕分数8' in result:
                    ws.cell(row=current_row, column=26, value=result["仕分数8"] if result["仕分数8"] !=0 else "")
                if '仕分数9' in result:
                    ws.cell(row=current_row, column=27, value=result["仕分数9"] if result["仕分数9"] !=0 else "")
                if '仕分数10' in result:
                    ws.cell(row=current_row, column=28, value=result["仕分数10"] if result["仕分数10"] !=0 else "")
                if '仕分数11' in result:
                    ws.cell(row=current_row, column=29, value=result["仕分数11"] if result["仕分数11"] !=0 else "")
                if '仕分数12' in result:
                    ws.cell(row=current_row, column=30, value=result["仕分数12"] if result["仕分数12"] !=0 else "")
                if '仕分数13' in result:
                    ws.cell(row=current_row, column=31, value=result["仕分数13"] if result["仕分数13"] !=0 else "")
                if '仕分数14' in result:
                    ws.cell(row=current_row, column=32, value=result["仕分数14"] if result["仕分数14"] !=0 else "")
                if '仕分数15' in result:
                    ws.cell(row=current_row, column=33, value=result["仕分数15"] if result["仕分数15"] !=0 else "")
                if '仕分数16' in result:
                    ws.cell(row=current_row, column=34, value=result["仕分数16"] if result["仕分数16"] !=0 else "")
                if '仕分数17' in result:
                    ws.cell(row=current_row, column=35, value=result["仕分数17"] if result["仕分数17"] !=0 else "")
                if '仕分数18' in result:
                    ws.cell(row=current_row, column=36, value=result["仕分数18"] if result["仕分数18"] !=0 else "")
                if '仕分数19' in result:
                    ws.cell(row=current_row, column=37, value=result["仕分数19"] if result["仕分数19"] !=0 else "")
                if '仕分数20' in result:
                    ws.cell(row=current_row, column=38, value=result["仕分数20"] if result["仕分数20"] !=0 else "")
                if '仕分数21' in result:
                    ws.cell(row=current_row, column=39, value=result["仕分数21"] if result["仕分数21"] !=0 else "")
                if '仕分数22' in result:
                    ws.cell(row=current_row, column=40, value=result["仕分数22"] if result["仕分数22"] !=0 else "")
                if '仕分数23' in result:
                    ws.cell(row=current_row, column=41, value=result["仕分数23"] if result["仕分数23"] !=0 else "")
                if '仕分数24' in result:
                    ws.cell(row=current_row, column=42, value=result["仕分数24"] if result["仕分数24"] !=0 else "")
                if '仕分数25' in result:
                    ws.cell(row=current_row, column=43, value=result["仕分数25"] if result["仕分数25"] !=0 else "")
                if '仕分数26' in result:
                    ws.cell(row=current_row, column=44, value=result["仕分数26"] if result["仕分数26"] !=0 else "")
                if '仕分数27' in result:
                    ws.cell(row=current_row, column=45, value=result["仕分数27"] if result["仕分数27"] !=0 else "")
                if '仕分数28' in result:
                    ws.cell(row=current_row, column=46, value=result["仕分数28"] if result["仕分数28"] !=0 else "")
                if '仕分数29' in result:
                    ws.cell(row=current_row, column=47, value=result["仕分数29"] if result["仕分数29"] !=0 else "")
                if '仕分数30' in result:
                    ws.cell(row=current_row, column=48, value=result["仕分数30"] if result["仕分数30"] !=0 else "")

                #書式設定
                fillcolor = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
                rule1 = FormulaRule(formula =[f'=$J${current_row}="3150"'], fill = fillcolor)
                ws.conditional_formatting.add(f'A{current_row}:M{current_row}', rule1)
                rule2 = FormulaRule(formula =[f'=$L${current_row}="03"'], fill = fillcolor)
                ws.conditional_formatting.add(f'A{current_row}:M{current_row}', rule2)

                current_row += 1

            # 履歴情報をセット
            storedname="usp_見積出力履歴更新"

            params = {
                "@iListName":listName,
                "@iNumber":request.params["EstimateNo"],
                "@iPCName":request.params["LoginId"],
            }

            # STORED実行
            storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

            # # 不要なテンプレートシートの削除
            # workbook.remove(template_sheet)

            ws.title = listName

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, workbook, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            ) 
       
    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
