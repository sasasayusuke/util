from datetime import datetime
from typing import Dict, Any
import io
import logging
from copy import copy
from fastapi.responses import StreamingResponse
from app.core.exceptions import ServiceError
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer
from app.utils.string_utils import NullToZero
from app.core.config import settings
from openpyxl.utils import get_column_letter

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class SyukkalistService(BaseService):
    """出荷リストサービス"""

    def display(self, request, session) -> StreamingResponse:
        """印刷処理を行うメソッド"""

        # 出荷リストの印刷処理
        logger.info("【START】出荷リスト出力処理")

        # SQL実行　見積情報のチェック
        query = "SELECT 見積番号, 見積件名, 得意先名1, 得意先名2, 現場名, 納期S, 納期E From TD見積 WHERE 見積番号 = {}".format(request.params['@i見積番号'])
        sqlresult=dict(SQLExecutor(session).execute_query(query=query))

        # STORED実行
        storedname="usp_MT1200出荷リスト抽出_梱包出荷用"
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=request.params))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            # Excelオブジェクトの作成
            wb = create_excel_object('Temp出荷リスト_梱包出荷用.xlsx')
            ws = wb.active

            # ヘッダー情報追加
            # 納品日
            ws['E1'] = sqlresult['results'][0]['納期S'].strftime("%m/%d（%a）")
            # 見積No
            ws['C3'] = storedresults['results'][0][0]['見積番号']
            # 見積名称
            ws['C4'] = storedresults['results'][0][0]['見積件名']
            # 送り先
            ws['H2'] = storedresults['results'][0][0]['現場名']
            # 店番
            ws['N2'] = storedresults['results'][0][0]['納入先CD']
            # 郵便番号
            ws['I3'] = storedresults['results'][0][0]['郵便番号']
            # TEL
            ws['L3'] = "TEL:　　{}".format(storedresults['results'][0][0]['納TEL'])
            # 担当
            ws['O3'] = request.user
            # 住所1
            ws['H4'] = storedresults['results'][0][0]['住所1']
            # 住所2
            ws['H5'] = storedresults['results'][0][0]['住所2']


            # 初期位置指定
            start_row = 8

            for row_num, row_data in enumerate(storedresults['results'][0], start=start_row):
                logger.info("列番号:{}".format(row_num))
                # 項目
                if row_data['見積区分'] ==  "C" or  row_data['見積区分'] == "A" or  row_data['見積区分'] == "S":
                    ws.cell(row=row_num, column=1, value="")._style = copy(ws.cell(row=start_row, column=1)._style)
                else:
                    ws.cell(row=row_num, column=1, value=row_data['行番号'])._style = copy(ws.cell(row=start_row, column=1)._style)
                # 棚番
                ws.cell(row=row_num, column=2, value=row_data['棚番名'])._style = copy(ws.cell(row=start_row, column=2)._style)
                # 製品No
                ws.cell(row=row_num, column=3, value=row_data['製品NO'])._style = copy(ws.cell(row=start_row, column=3)._style)
                # 仕様No
                ws.cell(row=row_num, column=4, value=row_data['仕様NO'])._style = copy(ws.cell(row=start_row, column=4)._style)
                # 名称
                if row_data['見積区分'] ==  "C" or  row_data['見積区分'] == "A" or  row_data['見積区分'] == "S":
                    ws.cell(row=row_num, column=5, value="  {}".format(row_data['漢字名称']))._style = copy(ws.cell(row=start_row, column=5)._style)
                else:
                    ws.cell(row=row_num, column=5, value=row_data['漢字名称'])._style = copy(ws.cell(row=start_row, column=5)._style)
                # W
                ws.cell(row=row_num, column=6, value=row_data['W'] if row_data["W"] != 0 else "")._style = copy(ws.cell(row=start_row, column=6)._style)
                # D
                if row_data['D'] == 0:
                    if row_data['D1'] == 0 and row_data['D2'] == 0:
                        ws.cell(row=row_num, column=7, value="")._style = copy(ws.cell(row=start_row, column=7)._style)
                    else:
                        ws.cell(row=row_num, column=7, value="{D1}/{D2}".format(D1=row_data['D1'],D2=row_data['D2']))._style = copy(ws.cell(row=start_row, column=7)._style)
                else:
                    ws.cell(row=row_num, column=7, value=row_data['D'])._style = copy(ws.cell(row=start_row, column=7)._style)
                # H
                if row_data['H'] == 0:
                    if row_data['H1'] == 0 and row_data['H2'] == 0:
                        ws.cell(row=row_num, column=8, value="")._style = copy(ws.cell(row=start_row, column=8)._style)
                    else:
                        ws.cell(row=row_num, column=8, value="{H1}/{H2}".format(H1=row_data['H1'],H2=row_data['H2']))._style = copy(ws.cell(row=start_row, column=8)._style)
                else:
                    ws.cell(row=row_num, column=8, value=row_data['H'])._style = copy(ws.cell(row=start_row, column=8)._style)
                # 発注数
                ws.cell(row=row_num, column=9, value=row_data['発注数'])._style = copy(ws.cell(row=start_row, column=9)._style)
                # 単位名
                ws.cell(row=row_num, column=10, value=row_data['単位名'])._style = copy(ws.cell(row=start_row, column=10)._style)
                # 仕入先
                if row_data['仕入先CD'] == "3150" and row_data['配送先CD'] == "02":
                    ws.cell(row=row_num, column=11, value="{siiresaki_name}-{haisousaki_name}".format(siiresaki_name=row_data['仕入先名'],haisousaki_name=row_data['配送先名']))._style = copy(ws.cell(row=start_row, column=11)._style)
                else:
                    ws.cell(row=row_num, column=11, value=row_data['仕入先名'])._style = copy(ws.cell(row=start_row, column=11)._style)
                # 備考
                ws.cell(row=row_num, column=12, value=row_data['明細備考'])._style = copy(ws.cell(row=start_row, column=12)._style)
                ws.cell(row=row_num, column=13)._style = copy(ws.cell(row=start_row, column=13)._style)
                # 入庫日
                ws.cell(row=row_num, column=14, value=row_data['入庫日'])._style = copy(ws.cell(row=start_row, column=14)._style)
                # チェック
                ws.cell(row=row_num, column=15)._style = copy(ws.cell(row=start_row, column=15)._style)

                # 高さ調整
                ws.row_dimensions[row_num].height = ws.row_dimensions[start_row].height

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # 印刷範囲の設定
            ws.print_area = 'A1:O{}'.format(row_num)

            # シート名の変更
            ws.title = '出荷リスト'

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
