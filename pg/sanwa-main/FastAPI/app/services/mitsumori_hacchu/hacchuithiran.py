import logging
from datetime import datetime
from typing import Dict, Any
from collections import defaultdict
from openpyxl.styles import Border, Side
from openpyxl.styles.fonts import Font
from app.core.exceptions import ServiceError
from app.core.config import settings
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    copy_sheet_settings,
    file_generate_response,
    range_copy_cell_by_address
)

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class HacchuithiranService(BaseService):
    def display(self, request, session):
        """発注一覧表出力処理"""
        logger.info("【START】発注一覧表出力処理")

        stored_name = "usp_HC0300発注一覧表"

        # リクエストパラメータを取得
        params = {
            "@i見積番号": request.params.get('@i見積番号'),
            "@is仕入先CD": int(request.params.get('@is仕入先CD', '0') or '1'),
            "@ie仕入先CD": int(request.params.get('@ie仕入先CD', '0') or '9999'),
        }

        # 出力形式の取得 (デフォルトはExcel)
        output_format = request.params.get('format', 'excel').lower()
        logger.debug(f"出力形式: {output_format}")

        # デバッグ用: パラメータの確認
        logger.debug(f"Stored procedure parameters: {params}")

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(
            storedname=stored_name, params=params))

        # デバッグ用: 結果の中身を確認
        logger.debug(f"Stored procedure raw results: {storedresults}")

        # 結果が空の場合はエラーを投げる
        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')
        
        # Excelファイルの作成とレスポンスの生成
        with get_excel_buffer() as buffer:
            # Excelオブジェクトの作成
            wb = create_excel_object('Template_発注一覧表.xlsx')
            ts = wb["Copy"]

            # グルーピングしてシートを作成
            self.create_sheets_by_supplier(wb, storedresults["results"], ts)

            # 出力形式の判定しファイルを出力
            return file_generate_response(buffer, wb, output_format)
        

    def create_sheets_by_supplier(self, wb, results, ts):
        """仕入先CDごとにシートを作成"""
        # データを仕入先CDでグルーピング
        grouped_data = defaultdict(list)
        for item in results[0]:
            grouped_data[item["仕入先CD"]].append(item)

        # 各仕入先ごとにシートを作成
        for supplier_cd, items in grouped_data.items():
            supplier_name = items[0]["仕入先名1"]
            sheet_name = supplier_name[:31]  # シート名は31文字以内に制限

            # テンプレートシートのコピー
            template_sheet = wb['Template']
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name

            # ページ設定の引き継ぎ
            copy_sheet_settings(template_sheet, new_sheet)

            # ヘッダーにデータ挿入
            self.insert_header_data(new_sheet, items[0])
            # 表にデータ挿入処理
            self.populate_data_to_sheet(new_sheet, items, ts)

        # 処理が終わったらテンプレートシートを削除
        del wb['Template']
        del wb['Copy']

    def insert_header_data(self, ws, first_item):
        """ヘッダー情報をシートに挿入"""
        ws["Q1"].value = datetime.now().strftime('%Y{0}%m{1}%d{2}').format(*'年月日')  # 作成日
        ws["C2"].value = first_item["見積件名"]  # 物件名
        ws["C3"].value = '[' + str(first_item['納入先CD']) + '] ' + first_item['納入先名1'] + first_item['納入先名2']  # 納入先情報
        ws["M2"].value = '[' + str(first_item["仕入先CD"]) + ']'  # 仕入先CD
        ws["M3"].value = first_item["仕入先名1"]  # 仕入先名
        ws["T2"].value = '[' + str(first_item["見積番号"]) + ']'  # 見積番号
        ws["T3"].value = first_item["見積日付"].strftime('%Y{0}%m{1}%d{2}').format(*'年月日')  # 見積日付

    def populate_data_to_sheet(self, ws, items, ts):
        """シートにデータを挿入"""
        start_row = 7  # 表のデータ挿入開始行
        row_height = ws.row_dimensions[start_row].height  # 開始行の高さを取得
        thin_border = Border(bottom=Side(style='dashed'))  # 波線の枠線を定義
        total_amount = 0  # 合計金額を計算

        # 表データの挿入
        for i, item in enumerate(items, start=start_row):
            # 詳細行をコピー
            range_copy_cell_by_address(ts,ws,'A1','U1','A{}'.format(i),True)
            self.insert_row_data(ws, i, item)

            # 行の高さを継承
            ws.row_dimensions[i].height = row_height

            # 波線の枠線を追加
            for col in range(1, 22):  # A列(1)からU列(21)まで
                cell = ws.cell(row=i, column=col)
                cell.border = thin_border

            # 合計金額を計算
            total_amount += item["仕入金額"] if item["仕入金額"] != 0 else 0

        # 合計金額の挿入
        total_row = start_row + len(items)
        ws[f"F{total_row}"].value = "＊仕入先計＊"  # 合計ラベルを設定
        ws[f"F{total_row}"].font = Font(bold = True)  # 合計ラベルを太文字化
        ws[f"U{total_row}"].value = total_amount  # 合計金額を設定
        ws[f"U{total_row}"].number_format = '#,##0'  # 数値フォーマットを設定

        # 合計金額の行に上下の線を引く（U列まで）
        self.add_border_to_total_row(ws, total_row, height=25)  # 高さ25を指定

    def add_border_to_total_row(self, ws, total_row, height=None):
        """合計金額の行に上下の線を引く"""

        # 行の高さを指定
        if height is not None:
            ws.row_dimensions[total_row].height = height

        border_style = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # U列までの範囲に線を引く
        for col in range(1, 22):  # F列(6)からU列(21)まで
            cell = ws.cell(row=total_row, column=col)
            cell.border = border_style

    def insert_row_data(self, ws, row, item):
        """シートの表部分にデータを挿入"""
        # 行番号
        ws[f"A{row}"].value = item["行番号"] if item["行番号"] != 0 else ""

        # 製品No
        ws[f"B{row}"].value = item["製品NO"] if item["製品NO"] != 0 else ""

        # 仕様No
        ws[f"D{row}"].value = item["仕様NO"] if item["仕様NO"] != 0 else ""

        # 品名
        ws[f"F{row}"].value = item["漢字名称"]

        # サイズ (W×D×H)
        size_str = ''
        if item["W"] > 0:
            size_str = str(item["W"]) + 'W'
        if item["D"] > 0:
            if size_str != '':
                size_str += '×'
            size_str += str(item["D"]) + 'D'
        else:
            if not (item["D1"] == 0 and item["D2"] == 0):
                if size_str != '':
                    size_str += '×'
                size_str += str(item["D1"]) + '/' + str(item["D2"]) + 'D'
        if item["H"] > 0:
            if size_str != '':
                size_str += '×'
            size_str += str(item["H"]) + 'H'
        else:
            if not (item["H1"] == 0 and item["H2"] == 0):
                if size_str != '':
                    size_str += '×'
                size_str += str(item["H1"]) + '/' + str(item["H2"]) + 'H'

        ws[f"K{row}"].value = size_str

        # 発注数
        ws[f"Q{row}"].value = item["発注数"] if item["発注数"] != 0 else ""

        # 単位名
        ws[f"R{row}"].value = item["単位名"]

        # 仕入単価: U区分が"U"の場合に "未" を付与
        if item["U区分"] == "U":
            ws[f"T{row}"].value = f'{item["仕入単価"]} 未'
        else:
            ws[f"T{row}"].value = item["仕入単価"] if item["仕入単価"] != 0 else ""

        # 仕入金額
        ws[f"U{row}"].value = item["仕入金額"] if item["仕入金額"] != 0 else ""

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
