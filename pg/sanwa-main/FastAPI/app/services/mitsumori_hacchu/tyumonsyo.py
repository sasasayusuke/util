import datetime
from typing import Dict, Any
import io
import logging
from copy import copy
from fastapi.responses import StreamingResponse
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.string_utils import null_to_zero
from app.utils.service_utils import ClsMitumoriH, ClsTanto, ClsOutputRireki
from app.utils.string_utils import format_jp_date
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    ExcelPDFConverter,
    range_copy_cell
)
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class tyumonsyoService(BaseService):
    """注文書サービス"""

    def display(self, request, session) -> Dict[str, Any]:

        # 注文書の印刷処理
        logger.info("【START】注文書処理")

        storedname1 = "usp_MT0300見積書出力"
        storedname2 = "usp_MT0301見積書出力_客在"
        storedname3 = "usp_MT0302見積書出力_部材計"
        storedname4 = "usp_MT0303見積書出力_部材計_客在"

        if request.params['@i相見積区分'] == "0" or request.params['@i相見積区分'] == "1":
            storedname = storedname1
        elif request.params['@i相見積区分'] == "2":
            storedname = storedname2
        elif request.params['@i相見積区分'] == "3":
            storedname = storedname3
        elif request.params['@i相見積区分'] == "4":
            storedname = storedname4

        params = {
            "@i見積番号": request.params['@i見積番号']
        }

        outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams=outputparams))

        storedoutput = storedresults["output_values"]
        if storedoutput["@RetST"] != 0:
            raise ServiceError("{} : {}".format(storedoutput["@RetST"],storedoutput["@RetMsg"]))

        if len(storedresults["results"][0]) == 0 or len(storedresults["results"][1]) == 0 or len(storedresults["results"][2]) == 0:
            raise ServiceError("該当データ無し")

        cMitumoriH = ClsMitumoriH(request.params["@i見積番号"])

        getbyid =  cMitumoriH.GetbyID(session)

        u_category = cMitumoriH.IsUCategory(session,'B')

        if u_category:
            raise ServiceError('売価未確定項目があります。')

        if getbyid:
            if cMitumoriH.原価率 >= 100:
                raise ServiceError('原価率が100%に達しています。')

        header = storedresults['results'][0][0]
        header_utiwake = storedresults['results'][1]
        sheet_utiwake = storedresults['results'][2]

        meisai_min = False

        if len(header_utiwake) == 1 and len(header_utiwake + sheet_utiwake) <= 17:
            meisai_min = True

        with get_excel_buffer() as buffer:

            wb = create_excel_object('Template_注文書.xlsx')

            ws = wb['Template']

            # 見積ヘッダー情報
            # 見積番号
            ws['A1'] = "№{}".format(header['見積番号'])
            # 作成日
            if header['見積日出力'] == 1:
                ws['O1'] = format_jp_date(header['見積日付'])
            # 得意先名
            if not header['得意先名2']:
                ws['K7'] = ""
                ws['K8'] = header['得意先名1']
            else:
                ws['K7'] = header['得意先名1']
                ws['K8'] = header['得意先名2']

            # 御見積金額
            ws['D8'] = header['合計金額'] + header['出精値引']
            # 消費税額
            ws['D9'] = header['外税額']
            # 件名
            ws['D12'] = header['見積件名']
            # 納期
            if header['納期表示'] == 0:
                if header['納期S'] == None and header['納期E'] == None:
                    ws['D13'] = ""
                elif header['納期S'] and header['納期E']:
                    ws['D13'] = f"{format_jp_date(header['納期S'])}～{format_jp_date(header['納期E'])}"
                elif header['納期S'] == None:
                    ws['D13'] = format_jp_date(header['納期E'])
                else:
                    ws['D13'] = format_jp_date(header['納期S'])

            elif header['納期表示'] == 1:
                ws['D13'] =  "別途御打ち合せによる"
            elif header['納期表示'] == 2:
                ws['D13'] =  header['納期その他']

            # 受渡地 店番
            if header['納入得意先CD'] == "2401" or header['納入得意先CD'] == "8001":
                ws['D14'] =  str(header['現場名']).replace(header['納入先CD'], "")
                ws['G14'] =  header['納入先CD']
            else:
                ws['D14'] =  header['現場名']
                ws['G14'] =  ""

            # 御支払条件
            if header['支払条件'] == 0:
                ws['D15'] =  "別途御打ち合せによる"
            elif header['支払条件'] == 1:
                ws['D15'] = header['支払条件その他']

            # 見積有効期限
            if header['有効期限'] != 0:
                ws['D16'] = "{}日".format(header['有効期限'])

            # 備考
            ws['D17'].value = header['備考']

            # 注文者名 会社担当者
            # tanto_result = ClsTanto.GetbyID(self, session, header['担当者CD'])
            # ws['K16'] = "注文者名：{}".format(tanto_result[0]['問い合わせ先'])

            # ウエルシア専用
            if header['ウエルシア売場面積'] != 0:
                ws['A18'] = "売 場 面 積 ："
                ws['D18'] = "{}坪".format(header['ウエルシア売場面積'])
                side = Side(style='thin', color='000000')
                welcia_border = Border(top=side, bottom=side)
                for column in range(1,8):
                    ws.cell(row=18, column=column).border = welcia_border

            if meisai_min == True:
            
                # 見積シート情報
                start_row = 21

                row_num = start_row

                wSiwaNo = None

                utiwake_ws = wb['Template_utiwake']

                for row_data in sheet_utiwake:
                    if wSiwaNo != row_data['仕分名称']:
                        for range_col in range(1,16):
                            ws.cell(row=row_num, column=range_col)._style = copy(ws.cell(row=start_row, column=range_col)._style)
                        ws.cell(row=row_num, column=1, value=row_data['仕分番号'])
                        ws.cell(row=row_num, column=3, value=row_data['仕分名称'])

                        wSiwaNo = row_data['仕分名称']
                        row_num += 1
                    # SEC
                    ws.cell(row=row_num, column=1, value=row_data['仕分番号'])._style = copy(ws.cell(row=start_row, column=1)._style)
                    # 38行以降の3列と4列目をセル結合
                    if row_num > 39:
                        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
                    if row_data['見積区分'] == "C" or row_data['見積区分'] == "A" or row_data["見積区分"] == "S":
                        # 項目
                        ws.cell(row=row_num, column=2, value="")._style = copy(ws.cell(row=start_row, column=2)._style)
                        # 品名
                        ws.cell(row=row_num, column=3, value="  {}".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=3)._style)
                        ws.cell(row=row_num, column=4)._style = copy(ws.cell(row=start_row, column=4)._style)
                        # ws.cell(row=row_num, column=17, value="  ".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=17)._style)
                    else:
                        # 項目
                        ws.cell(row=row_num, column=2, value=row_data['行番号'])._style = copy(ws.cell(row=start_row, column=2)._style)

                        # 品名
                        ws.cell(row=row_num, column=3, value="  {}".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=3)._style)
                        ws.cell(row=row_num, column=4)._style = copy(ws.cell(row=start_row, column=4)._style)
                        # ws.cell(row=row_num, column=17, value="  ".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=17)._style)
                    ws.cell(row=row_num, column=5, value="{}{}".format(row_data['PC区分'], row_data['製品NO']))._style = copy(ws.cell(row=start_row, column=5)._style)
                    ws.cell(row=row_num, column=6, value=row_data['仕様NO'])._style = copy(ws.cell(row=start_row, column=6)._style)
                    ws.cell(row=row_num, column=7, value=row_data['W'])._style = copy(ws.cell(row=start_row, column=7)._style)

                    if row_data['D'] == 0:
                        if row_data['D1'] == 0 and row_data['D2'] == 0:
                            ws.cell(row=row_num, column=8, value="")._style = copy(ws.cell(row=start_row, column=8)._style)
                        else:
                            ws.cell(row=row_num, column=8, value="{}/{}".format(row_data['D1'], row_data['D2']))._style = copy(ws.cell(row=start_row, column=8)._style)
                    else:
                        ws.cell(row=row_num, column=8, value=row_data['D'])._style = copy(ws.cell(row=start_row, column=8)._style)

                    if row_data['H'] == 0:
                        if row_data['H1'] == 0 and row_data['H2'] == 0:
                            ws.cell(row=row_num, column=9, value="")._style = copy(ws.cell(row=start_row, column=9)._style)
                        else:
                            ws.cell(row=row_num, column=9, value="{}/{}".format(row_data['H1'], row_data['H2']))._style = copy(ws.cell(row=start_row, column=9)._style)
                    else:
                        ws.cell(row=row_num, column=9, value=row_data['H'])._style = copy(ws.cell(row=start_row, column=9)._style)

                    # 数量_元
                    ws.cell(row=row_num, column=10, value=row_data['数量'])._style = copy(ws.cell(row=start_row, column=10)._style)
                    # 数量
                    ws.cell(row=row_num, column=11, value='=IF(MOD(J{row_num},1)=0,TEXT(J{row_num},"#,###"),TEXT(J{row_num},"#,##0.##"))'.format(row_num=row_num))._style = copy(ws.cell(row=start_row, column=11)._style)
                    # 単位
                    ws.cell(row=row_num, column=12, value=row_data['単位名'])._style = copy(ws.cell(row=start_row, column=12)._style)
                    # 単価_元 金額
                    if request.params['@i相見積区分'] != "1":
                        ws.cell(row=row_num, column=13, value=row_data['売上単価'])._style = copy(ws.cell(row=start_row, column=13)._style)
                        if row_data['見積区分'] != "A" and row_data['見積区分'] != "C" and row_data['見積区分'] != "S":
                            ws.cell(row=row_num, column=15, value=row_data['売上金額'])._style = copy(ws.cell(row=start_row, column=15)._style)
                        else:
                            ws.cell(row=row_num, column=15, value="" if row_data['売上金額'] == 0 else row_data['売上金額'])._style = copy(ws.cell(row=start_row, column=15)._style)
                    else:
                        ws.cell(row=row_num, column=13, value="")._style = copy(ws.cell(row=start_row, column=13)._style)
                        ws.cell(row=row_num, column=15, value="")._style = copy(ws.cell(row=start_row, column=15)._style)
                    # 単価
                    ws.cell(row=row_num, column=14, value='=IF(MOD(M{row_num},1)=0,TEXT(M{row_num},"#,###"),TEXT(M{row_num},"#,##0.##"))'.format(row_num=row_num))._style = copy(ws.cell(row=start_row, column=14)._style)

                    row_num += 1

                # 合計追加
                range_copy_cell(ws,ws,1,21,15,22,0,row_num - start_row,False)
                ws.cell(row=row_num+1, column=3, value="【合　　計】")._style = copy(ws.cell(row=start_row, column=15)._style)
                ws.cell(row=row_num+1, column=3).alignment = Alignment(horizontal='center',vertical='center')
                ws.cell(row=row_num+1, column=15, value="=SUM(O22:O{})".format(row_num-1))._style = copy(ws.cell(row=start_row, column=15)._style)

                ws.title = "見積積算書"

                # 改ページプレビューの設定
                ws.sheet_view.view = 'pageBreakPreview'
                ws.sheet_view.zoomScaleSheetLayoutView= 100

                # 履歴情報をセット
                storedname="usp_見積出力履歴更新"

                params = {
                    "@iListName":'注文書',
                    "@iNumber":request.params["@i見積番号"],
                    "@iPCName":request.user,
                }

                # STORED実行
                storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

                # 印刷範囲を設定
                ws.print_area = 'A1:O39'

                wb.remove(utiwake_ws)

                # pdf、excel判定
                if request.params['type'] == 'pdf':
                    
                    # PDFファイルの返却
                    converter = ExcelPDFConverter(
                        wb=wb,
                        ws=ws,
                        start_cell="A1",
                        end_cell="O39".format(row_num+1)
                    )

                    pdf_content = converter.generate_pdf()

                    return StreamingResponse(
                        io.BytesIO(pdf_content),
                        media_type="application/pdf",
                        headers={
                            "Content-Disposition": f'attachment; filename="sample.pdf"'
                        }
                    )

                else:

                    # Excelオブジェクトの保存
                    save_excel_to_buffer(buffer, wb, None)

                    return StreamingResponse(
                        io.BytesIO(buffer.getvalue()),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={
                            'Content-Disposition': 'attachment; filename="sample.xlsx"'
                        }
                    )
                
            else:

                start_row = 21

                row_num = start_row

                first_utiwake_table = False

                utiwake_count = 0

                utiwake_ws = wb['Template_utiwake']

                # 中央表示
                center_bottom = Alignment(horizontal='center',vertical='bottom')

                # 明細内訳
                for row_data in header_utiwake:
                    # SEC
                    ws.cell(row=row_num, column=1, value=row_data['仕分番号'])._style = copy(ws.cell(row=start_row, column=1)._style)
                    # 品名
                    ws.cell(row=row_num, column=3, value=row_data['仕分名称'])._style = copy(ws.cell(row=start_row, column=3)._style)
                    # 単価_元 金額
                    if request.params['@i相見積区分'] != "1":
                        ws.cell(row=row_num, column=15, value=row_data['売上金額'])._style = copy(ws.cell(row=start_row, column=15)._style)
                    else:
                        ws.cell(row=row_num, column=15, value="")._style = copy(ws.cell(row=start_row, column=15)._style)

                    row_num += 1
                    utiwake_count += 1

                    if utiwake_count == 17 and first_utiwake_table == False:
                        ws.cell(row=row_num+1, column=3, value="次ページへ")
                        range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num+1,True)
                        utiwake_count = 0
                        first_utiwake_table = True
                        row_num += 5

                    if utiwake_count == 32:
                        ws.cell(row=row_num+1, column=3, value="次ページへ")
                        range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num+1,True)
                        utiwake_count = 0
                        row_num += 5

                # 内訳合計
                ws['C'+ str(row_num+1)].alignment = center_bottom
                ws.cell(row=row_num+1, column=3, value="【合　　計】")
                ws.cell(row=row_num+1, column=15, value="=SUM(O21:O{})".format(row_num))

                if first_utiwake_table == True:

                    # 表行数との差を計算
                    row_diff = 33 - utiwake_count

                    # 明細用の表を追加
                    range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num+row_diff,True)

                    # 明細行の調整
                    row_num = row_num + row_diff + 4

                else:
                    # 表行数との差を計算
                    row_diff = 18 - utiwake_count

                    # 明細用の表を追加
                    range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num+row_diff,True)

                    # 明細行の調整
                    row_num = row_num + row_diff + 4

                # 見積シート情報
                wSiwaNo = None

                first_meisai_table = False

                # 明細カウント
                meisai_count = 0

                for row_data in sheet_utiwake:
                    if wSiwaNo != row_data['仕分名称'] and first_meisai_table == False:
                        for range_col in range(1,16):
                            ws.cell(row=row_num, column=range_col)._style = copy(ws.cell(row=start_row, column=range_col)._style)
                        ws.cell(row=row_num, column=1, value=row_data['仕分番号'])
                        ws.cell(row=row_num, column=3, value=row_data['仕分名称'])

                        wSiwaNo = row_data['仕分名称']

                        first_meisai_row = row_num

                        meisai_count += 1
                        row_num += 1

                        first_meisai_table = True

                    if wSiwaNo != row_data['仕分名称'] and first_meisai_table == True:

                        # 小計の追加
                        # 表に2行以上空きがない場合、次の表に
                        if meisai_count == 33:
                            range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num,True)
                            meisai_count = 0
                            row_num += 4

                            ws['C'+ str(row_num+1)].alignment = center_bottom
                            ws.cell(row=row_num+1, column=3, value="【{}小計】".format(wSiwaNo))
                            ws.cell(row=row_num+1, column=15, value="=SUM(O{}:O{})".format(first_meisai_row,row_num))

                        else:
                            ws['C'+ str(row_num+1)].alignment = center_bottom
                            ws.cell(row=row_num+1, column=3, value="【{}小計】".format(wSiwaNo))
                            ws.cell(row=row_num+1, column=15, value="=SUM(O{}:O{})".format(first_meisai_row,row_num))


                        # 仕分名称が変更されたら表を移動
                        # 表行数との差を計算
                        row_diff = 33 - meisai_count

                        # 表を追加
                        range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num+row_diff,True)
                        meisai_count = 0

                        row_num += 4 + row_diff
                        
                        for range_col in range(1,16):
                            ws.cell(row=row_num, column=range_col)._style = copy(ws.cell(row=start_row, column=range_col)._style)
                        ws.cell(row=row_num, column=1, value=row_data['仕分番号'])
                        ws.cell(row=row_num, column=3, value=row_data['仕分名称'])

                        wSiwaNo = row_data['仕分名称']
                        first_meisai_row = row_num
                        meisai_count += 1
                        row_num += 1
                        
                    # SEC
                    ws.cell(row=row_num, column=1, value=row_data['仕分番号'])._style = copy(ws.cell(row=start_row, column=1)._style)
                    # 38行以降の3列と4列目をセル結合
                    if row_num > 39:
                        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
                    if row_data['見積区分'] == "C" or row_data['見積区分'] == "A" or row_data["見積区分"] == "S":
                        # 項目
                        ws.cell(row=row_num, column=2, value="")._style = copy(ws.cell(row=start_row, column=2)._style)
                        # 品名
                        ws.cell(row=row_num, column=3, value="  {}".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=3)._style)
                        ws.cell(row=row_num, column=4)._style = copy(ws.cell(row=start_row, column=4)._style)
                        # ws.cell(row=row_num, column=17, value="  ".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=17)._style)
                    else:
                        # 項目
                        ws.cell(row=row_num, column=2, value=row_data['行番号'])._style = copy(ws.cell(row=start_row, column=2)._style)

                        # 品名
                        ws.cell(row=row_num, column=3, value="  {}".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=3)._style)
                        ws.cell(row=row_num, column=4)._style = copy(ws.cell(row=start_row, column=4)._style)
                        # ws.cell(row=row_num, column=17, value="  ".format(null_to_zero(row_data['漢字名称'], "")))._style = copy(ws.cell(row=start_row, column=17)._style)
                    ws.cell(row=row_num, column=5, value="{}{}".format(row_data['PC区分'], row_data['製品NO']))._style = copy(ws.cell(row=start_row, column=5)._style)
                    ws.cell(row=row_num, column=6, value=row_data['仕様NO'])._style = copy(ws.cell(row=start_row, column=6)._style)
                    ws.cell(row=row_num, column=7, value=row_data['W'])._style = copy(ws.cell(row=start_row, column=7)._style)

                    if row_data['D'] == 0:
                        if row_data['D1'] == 0 and row_data['D2'] == 0:
                            ws.cell(row=row_num, column=8, value="")._style = copy(ws.cell(row=start_row, column=8)._style)
                        else:
                            ws.cell(row=row_num, column=8, value="{}/{}".format(row_data['D1'], row_data['D2']))._style = copy(ws.cell(row=start_row, column=8)._style)
                    else:
                        ws.cell(row=row_num, column=8, value=row_data['D'])._style = copy(ws.cell(row=start_row, column=8)._style)

                    if row_data['H'] == 0:
                        if row_data['H1'] == 0 and row_data['H2'] == 0:
                            ws.cell(row=row_num, column=9, value="")._style = copy(ws.cell(row=start_row, column=9)._style)
                        else:
                            ws.cell(row=row_num, column=9, value="{}/{}".format(row_data['H1'], row_data['H2']))._style = copy(ws.cell(row=start_row, column=9)._style)
                    else:
                        ws.cell(row=row_num, column=9, value=row_data['H'])._style = copy(ws.cell(row=start_row, column=9)._style)

                    # 数量_元
                    ws.cell(row=row_num, column=10, value=row_data['数量'])._style = copy(ws.cell(row=start_row, column=10)._style)
                    # 数量
                    ws.cell(row=row_num, column=11, value='=IF(MOD(J{row_num},1)=0,TEXT(J{row_num},"#,###"),TEXT(J{row_num},"#,##0.##"))'.format(row_num=row_num))._style = copy(ws.cell(row=start_row, column=11)._style)
                    # 単位
                    ws.cell(row=row_num, column=12, value=row_data['単位名'])._style = copy(ws.cell(row=start_row, column=12)._style)
                    # 単価_元 金額
                    if request.params['@i相見積区分'] != "1":
                        ws.cell(row=row_num, column=13, value=row_data['売上単価'])._style = copy(ws.cell(row=start_row, column=13)._style)
                        if row_data['見積区分'] != "A" and row_data['見積区分'] != "C" and row_data['見積区分'] != "S":
                            ws.cell(row=row_num, column=15, value=row_data['売上金額'])._style = copy(ws.cell(row=start_row, column=15)._style)
                        else:
                            ws.cell(row=row_num, column=15, value="" if row_data['売上金額'] == 0 else row_data['売上金額'])._style = copy(ws.cell(row=start_row, column=15)._style)
                    else:
                        ws.cell(row=row_num, column=13, value="")._style = copy(ws.cell(row=start_row, column=13)._style)
                        ws.cell(row=row_num, column=15, value="")._style = copy(ws.cell(row=start_row, column=15)._style)
                    # 単価
                    ws.cell(row=row_num, column=14, value='=IF(MOD(M{row_num},1)=0,TEXT(M{row_num},"#,###"),TEXT(M{row_num},"#,##0.##"))'.format(row_num=row_num))._style = copy(ws.cell(row=start_row, column=14)._style)

                    meisai_count += 1

                    row_num += 1

                    if meisai_count == 34:
                        range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num-1,True)
                        meisai_count = 0
                        row_num += 3

                # 最終小計
                if meisai_count == 33:
                    range_copy_cell(utiwake_ws,ws,1,1,15,37,0,row_num,True)
                    meisai_count = 0
                    row_num += 4

                    ws['C'+ str(row_num+1)].alignment = center_bottom
                    ws.cell(row=row_num+1, column=3, value="【{}小計】".format(wSiwaNo))
                    ws.cell(row=row_num+1, column=15, value="=SUM(O{}:O{})".format(first_meisai_row,row_num))
                else:
                    ws['C'+ str(row_num+1)].alignment = center_bottom
                    ws.cell(row=row_num+1, column=3, value="【{}小計】".format(wSiwaNo))
                    ws.cell(row=row_num+1, column=15, value="=SUM(O{}:O{})".format(first_meisai_row,row_num-1))


                ws.title = "見積積算書"

                # 表行数との差を計算
                row_diff = 33 - meisai_count

                # 履歴情報セット
                # cOutputRireki = ClsOutputRireki
                # cOutputRireki.SetOutputLog(self,session,'注文書',request.params["@i見積番号"])

                # 履歴情報をセット
                storedname="usp_見積出力履歴更新"

                params = {
                    "@iListName":'注文書',
                    "@iNumber":request.params["@i見積番号"],
                    "@iPCName":request.user,
                }

                # STORED実行
                storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

                wb.remove(utiwake_ws)

                # pdf、excel判定
                if request.params['type'] == 'pdf':
                    # PDFファイルの返却
                    converter = ExcelPDFConverter(
                        wb=wb,
                        ws=ws,
                        start_cell="A1",
                        end_cell="O{}".format(row_num+row_diff)
                    )

                    pdf_content = converter.generate_pdf()

                    return StreamingResponse(
                        io.BytesIO(pdf_content),
                        media_type="application/pdf",
                        headers={
                            "Content-Disposition": f'attachment; filename="sample.pdf"'
                        }
                    )

                else:

                    # 印刷範囲を設定
                    ws.print_area = 'A1:O{}'.format(row_num+row_diff)

                    # Excelオブジェクトの保存
                    save_excel_to_buffer(buffer, wb, None)

                    return StreamingResponse(
                        io.BytesIO(buffer.getvalue()),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={
                            'Content-Disposition': 'attachment; filename="sample.xlsx"'
                        }
                    )


    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
