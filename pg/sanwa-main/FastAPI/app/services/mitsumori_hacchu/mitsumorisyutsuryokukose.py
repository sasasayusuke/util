import logging
import math
import re
from typing import Dict, Any
from copy import copy
from operator import itemgetter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from app.core.exceptions import ServiceError
from app.core.config import settings
from app.services.base_service import BaseService
from itertools import groupby
from app.utils.db_utils import SQLExecutor
from app.utils.string_utils import format_jp_date,null_to_zero
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    range_copy_cell,
    delete_excel_object,
    copy_images_to_new_sheet,
    register_history,
    copy_sheet_settings,
    file_generate_response,
    range_copy_cell_by_address
)
from app.utils.service_utils import ClsDates, ClsMitumoriH,ClsTanto
from decimal import Decimal

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)


class MitsumoriSyutsuryokuKoseService(BaseService):
    def display(self, request, session):
        """ 見積書出力（コーセー専用）処理 (index) """
        logger.info("【START】見積書出力（コーセー専用）処理")


        # チェック
        cMitumoriH = ClsMitumoriH(request.params['@i見積番号'])
        if cMitumoriH.IsUCategory(session,'B') == True:
            raise ServiceError('売価未確定項目があります。')
        

        # 使用するストアドプロシージャ名
        stored_name = "usp_MT0304見積書出力_客在_コーセー"

        # ストアプロシージャから取得為の必須条件
        params = {
            "@i見積番号": request.params['@i見積番号'],
            "@i見積番号明細部": request.params['@i明細見積']
        }
        # 出力形式の取得 (デフォルトはExcel)
        output_format = request.params.get('format', 'excel').lower()

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(
            storedname=stored_name, params=params, outputparams={"@RetST": "INT", "@RetMsg": "VARCHAR(255)"}))

        # データ有無チェック
        if len(storedresults["results"][0]) == 0 or len(storedresults["results"][1]) == 0 or len(storedresults["results"][2]) == 0:
            raise ServiceError("該当データなし")

        # シート作成処理
        with get_excel_buffer() as buffer:
            # テンプレートファイル
            wb = create_excel_object('Template_見積書(コーセー専用).xlsx')
            # テンプレートシートのコピー
            new_sheet = wb['別紙_テンプレート']
            # new_sheet = wb.copy_worksheet(template_sheet)
            # コピー先シート名
            new_sheet.title = "見積書"
            # ページ設定の引き継ぎ
            # copy_sheet_settings(template_sheet, new_sheet)
            # テンプレートシートの画像を新しいシートにコピー
            # copy_images_to_new_sheet(template_sheet, new_sheet)
            # 改ページプレビューの設定を追加
            new_sheet.sheet_view.view = 'pageBreakPreview'
            new_sheet.sheet_view.zoomScaleSheetLayoutView = 100
            # 明細用シート
            copy_ws = wb['copy_sheet']
            # データ量を判定して出力処理を振り分け
            if storedresults["count"] < 17:  # データ数が16より少ない場合
                self.populate_data_to_sheet(
                    session,new_sheet,copy_ws, storedresults['results'], request, True)
            else:
                self.populate_data_to_sheet(
                    session,new_sheet,copy_ws, storedresults['results'], request, False)
            # 処理が終わったらテンプレートシートを削除
            # del wb['別紙_テンプレート']
            del wb["copy_sheet"]

            # 履歴登録処理を呼び出す
            register_history(
                session, request.params["ListName"], request.params["EstimateNo"], request.params["LoginId"])

            # 出力形式の判定しファイルを出力
            return file_generate_response(buffer, wb, output_format)

    def populate_data_to_sheet(self, session,ws,copy_ws, array, request, min_mode=False):
        """データ 挿入処理"""
        # arrayの中身を確認
        logger.info(f"arrayの内容: {array}")

        # 固定データの挿入
        self.insert_static_data_to_sheet(session,ws, array, request)

        # データ数によるだし分け
        if min_mode:
            logger.info("処理状況確認 明細少ない")

            # データの挿入処理
            self.insert_third_array_data(ws,copy_ws, array[2], 21, 1,request)

            # 不要分の表を削除
            delete_excel_object(ws, 40, 1, 73, 15)
            ws.print_area = "A1:O36"

        else:
            logger.info("処理状況確認 明細多い")

            # 第二配列のデータ挿入
            current_row = self.insert_second_array_data(ws,copy_ws, array[1], 21, 1)

            # 第三配列のデータ挿入
            self.insert_third_array_group_data(ws,copy_ws, array, current_row, 1,request)

    def insert_static_data_to_sheet(self,session, ws, array, request):
        """シートの固定表記部分にデータをに挿入"""
        # 取得配列定義
        data = array[0][0]

        # 見積番号
        ws["A1"].value = "No" + str(data.get("見積番号", ""))

        # 見積日付
        見積日出力 = data.get("見積日出力", None)
        見積日付 = data.get("見積日付", None)
        if 見積日出力 == 1:  # 見積日出力が1の場合
            ws["O1"].value = format_jp_date(見積日付)
        

        # 得意先名
        得意先名1 = data.get("得意先名1", '')
        得意先名2 = data.get("得意先名2", '')
        if 得意先名2:  # 得意先名2がある場合
            ws["A3"].value = 得意先名1
            ws["A4"].value = 得意先名2
        else:  # 得意先名2がない場合
            ws["A4"].value = 得意先名1

        ws["E4"].value = "殿"

        # 御見積金額
        合計金額 = float(data.get("合計金額", 0))
        出精値引 = float(data.get("出精値引", 0))
        外税額 = float(data.get("外税額", 0))
        ws["D8"].value = 合計金額 + 出精値引
        ws["D9"].value = 外税額
        ws['D10'] = 合計金額 + 出精値引 + 外税額

        # 見積件名
        ws["D12"].value = data.get("見積件名", "")

        # 納期表示
        納期表示 = data.get("納期表示", None)
        納期S = data.get("納期S", None)
        納期E = data.get("納期E", None)
        納期その他 = data.get("納期その他", "")
        if 納期表示 == 0:
            if 納期S is None and 納期E is None:
                ws["D13"].value = ""
            elif 納期S is not None and 納期E is not None:
                ws["D13"].value = f"{format_jp_date(納期S)}～{format_jp_date(納期E)}"
            elif 納期S is None:
                ws["D13"].value = format_jp_date(納期E)
            else:
                ws["D13"].value = format_jp_date(納期S)
        elif 納期表示 == 1:
            ws["D13"].value = "別途御打ち合せによる"
        elif 納期表示 == 2:
            ws["D13"].value = 納期その他

        # 受渡地
        受渡地 = None
        店番 = None
        納入得意先CD = data.get("納入得意先CD", "")
        現場名 = data.get("現場名", "")
        納入先CD =  data.get("納入先CD", "")

        # 納入得意先CDに応じて処理を変更
        if 納入得意先CD in ["2401", "8001"]:
            # 現場名から納入得意先CDを削除
            受渡地 = re.sub(納入先CD, '', 現場名).strip()
            店番 = 納入先CD
            
        else:
            # 通常の処理
            受渡地 = 現場名
            店番 = None

        ws["D14"].value = 受渡地
        ws["G14"].value = 店番
        
        # 支払条件
        支払条件 = data.get("支払条件", None)
        支払条件その他 = data.get("支払条件その他", "")
        if 支払条件 == 0:
            ws["D15"].value = "別途御打ち合せによる"
        elif 支払条件 == 1:
            ws["D15"].value = 支払条件その他
        else:
            ws["D15"].value = ""

        # 有効期限
        有効期限 = data.get("有効期限", None)
        if 有効期限:
            ws["D16"].value = f"{有効期限}日"

        # 備考
        備考 = data.get("備考", "")
        ws["D17"].value = 備考

        # 担当者・問い合わせ先
        cTanto = ClsTanto()
        tanto_result = cTanto.GetbyID(session,request.params["表示担当"])
        if len(tanto_result) == 0:
            担当者名 = ""
        else:
            担当者名 = "問い合わせ先：{}".format(tanto_result[0]["問い合わせ先"])
        ws["K16"].value = 担当者名

        # ウエルシア売場面積
        sales_area = data.get("ウエルシア売場面積", 0)
        if sales_area:
            ws["A11"].value = "売 場 面 積 ："
            ws["D11"].value = f"{sales_area}坪"
            border = Border(left=Side(style="thin"), right=Side(
                style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            for col in range(1, 8):
                ws[f"{get_column_letter(col)}18"].border = border

        # クレーム区分設定
        claim_type = data.get("クレーム区分", 0)
        if claim_type == 1:
            # ws[""].value = "クレーム" #挿入場所不明
            pass

    def insert_second_array_data(self, ws,copy_ws, second_array, start_row, start_col):
        """
        第2配列のデータをExcelシートにループで配置し、合計を挿入する関数

        Parameters:
        ws (Worksheet): 対象のワークシート
        second_array (list): 第2配列のデータ
        start_row (int): 開始行番号（1ベース）
        start_col (int): 開始列番号（1ベース）
        """
        current_row = start_row  # データ挿入の開始行

        for item in second_array:
            if current_row == 37:
                ws.cell(row=current_row,column=3,value="次ページへ").alignment = Alignment(horizontal="center")
                # ws.cell(row=current_row,column=16,value="次ページへ").alignment = Alignment(horizontal="center")
                range_copy_cell_by_address(copy_ws,ws,'A1','O35','A38',True)
                current_row = 41
                continue
            # 各データを列に挿入
            ws.cell(row=current_row, column=start_col).value = item.get(
                "仕分番号", "")  # 列1のデータ
            ws.cell(row=current_row, column=start_col +
                    2).value = item.get("仕分名称", "")  # 列2のデータ
            ws.cell(row=current_row, column=start_col +
                    14).value = item.get("売上金額", "")  # 列3のデータ
            # ws.cell(row=current_row, column=start_col + 3).value = item.get("列4", "")  # 列4のデータ

            # 次の行に進む
            current_row += 1

        if current_row == 37:
            ws.cell(row=current_row,column=3,value="次ページへ").alignment = Alignment(horizontal="center")
            # ws.cell(row=current_row,column=16,value="次ページへ").alignment = Alignment(horizontal="center")
            current_row = 41

        # 合計ラベルの挿入
        total_row = current_row + 1  # 最終行の次に1行空けて挿入
        total_cell = ws.cell(row=total_row, column=start_col + 2)  # 合計セルを指定
        total_cell.value = "【合　　計】"  # 合計ラベル
        total_cell.alignment = Alignment(
            horizontal='center', vertical='center')  # 水平方向、垂直方向両方センタリング

        # 列14の合計を計算して挿入（`column=start_col + 14` の合計）
        total_col_idx = start_col + 14
        total_col = 0  # 合計を初期化
        for row_idx in range(start_row, current_row):
            cell_value = ws.cell(row=row_idx, column=total_col_idx).value
            if isinstance(cell_value, (int, float)):
                total_col += cell_value
        ws.cell(row=total_row, column=total_col_idx).value = total_col

        return current_row

    def insert_third_array_data(self, ws, copy_ws,third_array, start_row, start_col,request):
        """
        第3配列のデータをExcelシートにループで配置し、1行目に「仕分名称」を挿入し、2行目からデータを差し込む関数。
        行番号でソートも行う。

        Parameters:
        ws (Worksheet): 対象のワークシート
        third_array (list): 第3配列のデータ
        start_row (int): 開始行番号（1ベース）
        start_col (int): 開始列番号（1ベース）
        """
        # third_arrayの要素が辞書であることを前提にソート
        sorted_third_array = sorted(
            [item for item in third_array if isinstance(item, dict)],
            key=lambda x: x.get("行番号", 0)
        )

        current_row = start_row + 1  # 2行目からデータを挿入するため、開始行を1つ進める

        # 1行目に「仕分名称」を挿入
        if sorted_third_array:
            ws.cell(row=start_row, column=start_col).value = self.get_value_or_empty(sorted_third_array[0], "仕分番号")  # 仕分名称
            ws.cell(row=start_row, column=start_col +                    2).value = self.get_value_or_empty(sorted_third_array[0], "仕分名称")  # 仕分名称

        # ソートしたデータをループで挿入
        for item in sorted_third_array:
            # 1行分のデータを挿入するためにinsert_row_data関数を利用
            self.insert_row_data(ws, current_row, start_col, item,request,copy_ws)

            # 次の行に進む
            current_row += 1

        # 合計ラベルの挿入
        total_row = current_row + 1  # 最終行の次に1行空けて挿入
        total_cell = ws.cell(row=total_row, column=start_col + 2)  # 合計セルを指定
        total_cell.value = "【合　　計】"  # 合計ラベル
        total_cell.alignment = Alignment(
            horizontal='center', vertical='center')  # 水平方向、垂直方向両方センタリング

        # 売上金額列の合計を計算して挿入（`column=start_col + 14` の合計）
        total_col_idx = start_col + 14
        total_col = Decimal("0")  # 合計を初期化
        for row_idx in range(start_row + 1, current_row):  # 2行目から合計を計算
            cell_value = ws.cell(row=row_idx, column=total_col_idx).value
            total_col += self.get_numeric_value(cell_value, "売上金額")
        ws.cell(row=total_row, column=total_col_idx).value = total_col

        # 0のセルを空にする処理（合計後に行う）
        # for row_idx in range(start_row + 1, current_row):
        #     for col_idx in range(start_col, start_col + 15):  # 必要な範囲の列をループ
        #         cell_value = ws.cell(row=row_idx, column=col_idx).value
        #         if self.is_blank(cell_value):
        #             ws.cell(row=row_idx, column=col_idx).value = ""  # 空白処理

    # 
    
    def insert_third_array_group_data(self, ws, copy_ws,array, start_row, start_col,request):
        """
        第3配列のデータをExcelシートにグループごとに挿入し、合計を計算して挿入する関数。

        Parameters:
            ws (Worksheet): 対象のワークシート
            array (list): 第3配列のデータ
            start_row (int): 開始行番号（1ベース）
            start_col (int): 開始列番号（1ベース）
        """
        # 第3配列のデータを取得
        material_details = array[2] if len(array) > 2 else []

        table_height = 35  # 表全体の高さ（ヘッダー行を含む）
        max_rows_per_table = 32  # 表に挿入可能な最大行数（データ行のみ）
        siwake_no = ""
        siwake_name = ""
        sum_from = 0
        page_count = math.ceil((start_row - 38) / table_height)+1
        row = ((page_count-1) * table_height ) + 38
        row_count_for_page = 1

        # 明細欄を追加する関数
        def insert_detail(ws,copy_ws,row):
            row = ((page_count-1) * table_height) + 38
            range_copy_cell_by_address(copy_ws,ws,'A1','O35','A{}'.format(row),True)
            ws.print_area = "A1:O{}".format(row + 34)
            return row
        
        for record in material_details:

            # 仕分け番号の区切り（改行）
            if record["仕分番号"] != siwake_no:
                # 合計記載
                if siwake_no != "":
                    # 合計を入力
                    if row_count_for_page == 32 or row_count_for_page == 33:
                        row = insert_detail(ws,copy_ws,row)
                        row = row + 2
                        page_count = page_count + 1
                        row_count_for_page = 1
                    # 計算式をセット
                    ws.cell(row=row+1,column=3,value='【{}小計】'.format(siwake_name)).alignment = Alignment(horizontal="center")
                    ws["C{}".format(row+1)]._style = copy(copy_ws['Q1']._style)
                    ws.cell(row=row+1,column=15,value="=SUM(O{}:O{})".format(sum_from,row-1))

                # 新ページ追加
                row = insert_detail(ws,copy_ws,row)
                row += 3
                page_count += 1
                row_count_for_page = 1
                ws.cell(row=row,column=1,value=record["仕分番号"])
                ws.cell(row=row,column=3,value=null_to_zero(record["仕分名称"],"")).alignment = Alignment(horizontal='left',vertical='center')
                ws["C{}".format(row)]._style = copy(copy_ws['Q2']._style)
                siwake_no = record["仕分番号"]
                siwake_name = record["仕分名称"]
                row += 1
                row_count_for_page += 1
                sum_from = row
            
            # ページ最下行まで来たら新ページ作成
            if(row_count_for_page > max_rows_per_table):
                row = insert_detail(ws,copy_ws,row)
                row += 3
                row_count_for_page = 1
                page_count += 1
                ws["C{}".format(row)]._style = copy(copy_ws['Q2']._style)

            self.insert_row_data(ws,row,start_col,record,request,copy_ws)
            row += 1
            row_count_for_page += 1

        # 最後のページの合計
        if row_count_for_page == 33 or row_count_for_page == 32:
            row = insert_detail(ws,copy_ws,row)
            row = row + 2
            page_count = page_count + 1
            row_count_for_page = 1
        # 計算式をセット
        ws.cell(row=row+1,column=3,value='【{}小計】'.format(siwake_name)).alignment = Alignment(horizontal="center")
        ws["C{}".format(row+1)]._style = copy(copy_ws['Q1']._style)
        ws.cell(row=row+1,column=15,value="=SUM(O{}:O{})".format(sum_from,row-1))

    def copy_cell(self, ws, counter, start_row):
        """
        テンプレート表を特定位置から指定回数繰り返し配置する関数。

        Parameters:
            ws (Worksheet): データを挿入するシート
            counter (int): 表を配置する回数
            start_row (int) : 最初の表を配置する行
        """
        # コピー元テンプレートを開く
        wb = create_excel_object('Template_見積書.xlsx')
        src_sheet = wb['別紙_テンプレート']

        # コピー範囲の定義（テンプレートの表）
        min_col = 1  # A列
        min_row = 39  # 開始行
        max_col = 15  # R列
        max_row = 73  # 終了行
        table_height = max_row - min_row  # 表の高さを計算

        # 初期位置（配置開始セルを計算）
        current_row = start_row  # 最初の表を配置する行
        shift_col = 0  # 列は固定でA列から開始

        # 指定された回数分コピーして配置
        for i in range(counter):
            # コピー範囲を計算
            shift_row = current_row - min_row  # コピー先の行位置
            range_copy_cell(
                src_sheet,
                ws,
                min_col,
                min_row,
                max_col,
                max_row,
                shift_col,
                shift_row,
                True,
            )
            # 次の表の開始位置を計算
            current_row += table_height + 1  # 表の高さ + 空行1行分

    def insert_row_data(self, ws, row, col, item,request,copy_ws):
        """1行分のデータをExcelシートに挿入"""
        # 変数定義
        getQuoteKubun = self.get_value_or_empty(item, '見積区分')
        
        #仕分番号
        self.write_cell(ws, row, col, self.get_value_or_empty(item, "仕分番号"))

        #行番号
        行番号 = self.get_value_or_empty(item, '行番号')
        if getQuoteKubun not in ['C', 'A', 'S']:
            self.write_cell(ws, row, col + 1, 行番号)
        else:
            self.write_cell(ws, row, col + 1, "")

        # 品名
        漢字名称 = self.get_value_or_empty(item, '漢字名称')
        if getQuoteKubun in ['C', 'A', 'S']:
            productName = f'  {漢字名称}'
        else:
            productName = 漢字名称
        self.write_cell(ws, row, col + 2, productName)
        ws["C{}".format(row)]._style = copy(copy_ws['Q2']._style)
        # self.write_cell(ws, row, col + 2, productName, alignment=Alignment(horizontal='left'))
        ws.cell(row=row,column=3).alignment = Alignment(vertical='center',horizontal='left',shrink_to_fit =True)
        
        # 製品NO
        self.write_cell(ws, row, col + 4, self.get_value_or_empty(item, "製品NO"))
        # 仕様NO
        self.write_cell(ws, row, col + 5, self.get_value_or_empty(item, "仕様NO"))
        # サイズ
        self.write_cell(ws, row, col + 6, self.get_value_or_empty(item, "W"))

        d = None
        getD = self.get_value_or_empty(item, "D")
        getD1 = self.get_value_or_empty(item, "D1")
        getD2 = self.get_value_or_empty(item, "D2")
        if getD:
            d = getD
        elif getD1 and getD2:
            d = f'{getD1}/{getD2}'
        self.write_cell(ws, row, col + 7, d)

        h = None
        getH = self.get_value_or_empty(item, "H")
        getH1 = self.get_value_or_empty(item, "H1")
        getH2 = self.get_value_or_empty(item, "H2")
        if getH:
            h = getH
        elif getH1 and getH2:
            h = f'{getH1}/{getH2}'
        self.write_cell(ws, row, col + 8, h)

        # 数量（隠し行）
        self.write_cell(ws, row, col + 9, self.get_value_or_empty(item, "数量"))
        # 数量
        self.write_cell(ws, row, col + 10, '=IF(MOD(J{0},1)=0,TEXT(J{0},"#,###"),TEXT(J{0},"#,##0.##"))'.format(row))
        # 単位
        self.write_cell(ws, row, col + 11, self.get_value_or_empty(item, "単位名"))
        # 売上単価（隠し行）
        self.write_cell(ws, row, col + 12, self.get_value_or_empty(item, "売上単価"))
        # 売り上げ単価
        self.write_cell(ws, row, col + 13, '=IF(MOD(M{0},1)=0,TEXT(M{0},"#,###"),TEXT(M{0},"#,##0.##"))'.format(row))


        # 売上金額
        price = None
        getQuantity = self.get_numeric_value(item, "数量")
        getSalesUnitPrice = self.get_numeric_value(item, "売上単価")
        getSalesFraction = self.get_numeric_value(item, "売上端数")

        base_value = getQuantity * getSalesUnitPrice
        if getSalesFraction == 0:  # 四捨五入
            if getQuoteKubun not in ['C', 'A', 'S']:
                price = round(base_value, 0)
            elif round(base_value, 0):
                price = round(base_value, 0)
        elif getSalesFraction == 1:  # 切り上げ
            if getQuoteKubun not in ['C', 'A', 'S']:
                price = math.ceil(base_value)
            elif math.ceil(base_value):
                price = math.ceil(base_value)
        elif getSalesFraction == 2:  # 切り下げ
            if getQuoteKubun not in ['C', 'A', 'S']:
                price = math.floor(base_value)
            elif math.floor(base_value):
                price = math.floor(base_value)

        # セルに金額を記入
        self.write_cell(ws, row, col + 14, price)
        # sales_value = self.get_numeric_value(item, "売上金額")
        # self.write_cell(ws, row, col + 14, "" if sales_value == 0 else sales_value)

    def get_numeric_value(self, item, key):
        """数値データを安全に取得"""
        value = self.get_value_or_empty(item, key)
        return Decimal(value) if isinstance(value, (int, float, Decimal)) else Decimal("0")

    def write_cell(self, ws, row, col, value, alignment=None):
        """
        Excel セルに値を書き込むユーティリティ関数。
        値が 0 の場合は空文字を挿入する。

        Parameters:
        ws (Worksheet): 対象のワークシート
        row (int): 行番号（1ベース）
        col (int): 列番号（1ベース）
        value (any): 書き込む値
        alignment (Alignment): セルのアライメント（オプション）
        """
        cell = ws.cell(row=row, column=col)
        cell.value = "" if value == 0 else value

        # alignmentが指定されていれば、新しいインスタンスを作成して設定
        if alignment:
            cell.alignment = Alignment(
                horizontal=alignment.horizontal, vertical=alignment.vertical)

    def is_blank(self, value):
        """
        文字列が空または空白（スペース、タブ、改行、連続スペース）だけの場合はTrueを返す。
        """
        if isinstance(value, str):
            return not value.strip()  # 空文字かスペースだけならTrueを返す
        return value is None or value == "" or value == 0  # 0 や None を空白とみなす

    def get_value_or_empty(self, item, key):
        """
        itemが辞書の場合、keyで指定された値を取得し、存在しない場合は空文字を返す。
        itemが辞書でない場合は、そのままの値を返す。

        Parameters:
        item (any): 値（辞書型もしくはその他の型）
        key  (str): 辞書のキー

        Returns:
        str or any: 取得した値、もしくは空文字
        """
        if isinstance(item, dict):  # itemが辞書型の場合
            return item.get(key, "")
        else:
            # 辞書型でない場合、そのまま値を返す
            return item if item is not None else ""

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass

