import io
import logging
from datetime import datetime
from typing import Dict, Any, List
from copy import copy
from openpyxl import load_workbook
from fastapi.responses import StreamingResponse
from fastapi import HTTPException
from app.core.exceptions import ServiceError
from app.core.config import settings
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.string_utils import format_jp_date
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
)
from openpyxl.worksheet.pagebreak import Break

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class MitsumoriIraisyoService(BaseService):
    def display(self, request, session):
        """見積依頼書出力処理"""
        logger.info("【START】見積依頼書出力処理")

        storedname = "usp_MT0900見積依頼書出力"

        mitsumori_no = {
            "@i見積番号": request.params.get('@i見積番号'),
            "@i開始CD": int(request.params.get('@i開始CD', '0') or '1'),
            "@i終了CD": int(request.params.get('@i終了CD', '0') or '9999'),
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(
            storedname=storedname, params=mitsumori_no))

        # 結果が空の場合はエラーを投げる
        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            # Excelオブジェクトの作成
            wb = create_excel_object('Temp見積依頼書.xlsx')

            # 「別紙_テンプレート」というシートをコピー
            template_sheet = wb['別紙_テンプレート']

            # 仕入先CD-配送先CDの組み合わせでシートを作成
            created_sheets = set()  # 作成したシート名を記録
            for detail in storedresults['results'][0]:
                仕入先CD = detail.get('仕入先CD')
                配送先CD = detail.get('配送先CD')

                if 仕入先CD and 配送先CD:
                    sheet_name = f"{仕入先CD}-{配送先CD}"
                    if sheet_name not in created_sheets:
                        # シートをコピーして作成
                        new_sheet = wb.copy_worksheet(template_sheet)
                        new_sheet.title = sheet_name
                        created_sheets.add(sheet_name)

                        # ページレイアウト設定の引き継ぎ
                        new_sheet.page_setup = copy(template_sheet.page_setup)
                        new_sheet.page_margins = copy(template_sheet.page_margins)  # ページ余白の設定を引き継ぐ
                        new_sheet.print_options = copy(template_sheet.print_options)  # 印刷オプションを引き継ぐ

                        # フッターコピー
                        new_sheet.oddFooter.left.text = template_sheet.oddFooter.left.text
                        new_sheet.oddFooter.center.text = template_sheet.oddFooter.center.text

                        # 印刷タイトルのコピー（必要に応じて）
                        new_sheet.print_title_rows = template_sheet.print_title_rows
                        new_sheet.print_title_cols = template_sheet.print_title_cols

                        # 作成したシートにデータを設定
                        self.populate_data_to_sheet(new_sheet, detail, request)

                        # 表データを挿入し、必要に応じて追加
                        self.populate_table_data(new_sheet, storedresults['results'][0], sheet_name, request)

                        # 改ページプレビューの設定
                        new_sheet.sheet_view.view = 'pageBreakPreview'
                        new_sheet.sheet_view.zoomScaleSheetLayoutView= 100

            # 「別紙_テンプレート」シートを削除
            del wb['別紙_テンプレート']

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

    def populate_data_to_sheet(self, ws, detail, request):
        """シートにデータを挿入すwsるメソッド"""
        # 見積番号
        ws['P3'].value = detail.get('見積番号', '')
        # 仕入先名1
        仕入先名1 = detail.get('仕入先名1', '')
        仕入先名2 = detail.get('仕入先名2', '')

        if 仕入先名2:
            # 仕入先名1 と 仕入先名2 を改行で結合
            ws['A3'].value = f"{仕入先名1}\n{仕入先名2}"
        else:
            # 仕入先名2 がない場合は 仕入先名1 のみを挿入
            ws['A3'].value = 仕入先名1
        # 宛名
        ws['A5'].value = ""
        # 宛名に様を追記
        ws['H5'].value = "様"
        # 回答期限（フォームから受取、複数フォーマット対応）
        ws['E10'].value = format_jp_date(request.params['回答期限'])
        # 作成日
        ws['A12'].value = format_jp_date() + " 出力"
        # 作成者
        ws['Q12'].value = request.user

    def populate_table_data(self, ws, results, sheet_name, request):
        """表データを仕入先CD-配送先CDの組み合わせで挿入し、必要に応じて表を追加"""
        max_rows_per_table = 20  # 1つの表に入る最大行数
        row_start = 21  # 最初の表の開始行
        filtered_data = [
            detail for detail in results if f"{detail.get('仕入先CD')}-{detail.get('配送先CD')}" == sheet_name
        ]  # 配列をの仕入先CD-配送先CDの組み合わせでグルーピング
        total_rows = len(filtered_data)  # グルーピングされた中での配列の数を取得
        total_tables = (total_rows // max_rows_per_table) + (
            1 if total_rows % max_rows_per_table > 0 else 0
        )  # 必要な表の数を取得（ループ回数の取得）

        current_row = row_start  # 現在の行位置を設定

        for table_num in range(total_tables):  # ループする回数

            table_start_row = current_row #表の開始位置を変更（ループ用）

            if table_num == 1: #ループが初夏の場合
                ws.unmerge_cells('A42:R42')


            if table_num > 0:
                # 既存の表の最後尾から9行後に新しい表を挿入
                table_start_row = current_row + 8  # 前表の終了行 + 空行数
                current_row = table_start_row

                # コピー元定義
                wb = create_excel_object('Temp見積依頼書.xlsx')
                src_sheet = wb['別紙_テンプレート']

                # コピー先ワークシート（「現在データ挿入中のワークシート」）
                dst_sheet = ws  # 目的のシートを指定

                # コピー範囲（A14からR41）
                min_col = 1  # A列
                min_row = 14  # 14行
                max_col = 18  # R列
                max_row = 41  # 41行

                # 表を挿入開始する位置（A49）
                shift_col = 0  # 列オフセット（A列に挿入するので0）
                shift_row = current_row - 21 # 新しい行の位置に挿入

                # コピー範囲と挿入
                range_copy_cell(
                    src_sheet,
                    dst_sheet,
                    min_col,
                    min_row,
                    max_col,
                    max_row,
                    shift_col,
                    shift_row,
                    True,
                )

                # 改ページ

                row_break = Break(table_start_row+20)
                ws.row_breaks.append(row_break)
                logger.info('改ページ：{}'.format(table_start_row+20))

            # === 固定データの挿入 ====

            # 表全体の開始位置から7行上にずれる場合の調整
            adjusted_start_row = table_start_row - 8  # 表の開始位置から8行上にずらす
            # 見積件名
            ws.cell(row = adjusted_start_row + 2, column=1).value = filtered_data[0].get('見積件名', '')
            # 希望納期
            kibou_nouki = request.params['希望納期']
            if kibou_nouki is not None:
                ws.cell(row=adjusted_start_row + 4, column=1).value = format_jp_date(request.params['希望納期'])
            # 回答納期（空白）
            ws.cell(row = adjusted_start_row + 4, column=8).value = ""
            # 配送先
            配送先名1 = filtered_data[0].get('配送先名1', '')
            配送先名2 = filtered_data[0].get('配送先名2', '')
            ws.cell(row=adjusted_start_row + 2, column=14).value = f"{配送先名1} {配送先名2}"

            if filtered_data[0].get('配送先CD','') == '01':
                # 配送電話番号
                ws.cell(row=adjusted_start_row + 4, column=14).value = f"TEL {filtered_data[0].get('納TEL', '')}"
                # 配送郵便番号
                ws.cell(row=adjusted_start_row + 2, column=16).value = f"〒{filtered_data[0].get('郵便番号', '')}"
                # 配送住所
                ws.cell(row=adjusted_start_row + 3, column=16).value = f"{filtered_data[0].get('住所1', '')}\n{filtered_data[0].get('住所2', '')}"
            else:
                # 配送電話番号
                ws.cell(row=adjusted_start_row + 4, column=14).value = f"TEL {filtered_data[0].get('配送電話番号', '')}"
                # 配送郵便番号
                ws.cell(row=adjusted_start_row + 2, column=16).value = f"〒{filtered_data[0].get('配送郵便番号', '')}"
                # 配送住所
                ws.cell(row=adjusted_start_row + 3, column=16).value = f"{filtered_data[0].get('配送住所1', '')}\n{filtered_data[0].get('配送住所2', '')}"

            # データの挿入
            for i in range(min(max_rows_per_table, total_rows - (table_num * max_rows_per_table))):
                row = current_row + i  # 現在の行を計算
                detail = filtered_data[table_num * max_rows_per_table + i]

                # 行番号
                if detail.get('見積区分','') in ['C','A','S']:
                    ws.cell(row=row, column=2).value = ""
                else:
                    ws.cell(row=row, column=1).value = detail.get('行番号', '')

                ws.cell(row=row, column=2).value = detail.get('製品NO', '')
                ws.cell(row=row, column=4).value = detail.get('仕様NO', '')
                
                if (detail.get('W','') == 0 and detail.get('D','') == 0 and detail.get('H','') == 0
                    and detail.get('D1','') == 0 and detail.get('D2','') == 0
                    and detail.get('H1','') == 0 and detail.get('H2','') == 0
                ):
                    if  detail.get('見積区分','') in ['C','A','S']:
                        ws.cell(row=row, column=6).value = "\n  {}".format(detail.get('漢字名称', ''))
                    else:
                        ws.cell(row=row, column=6).value = detail.get('漢字名称', '')
                else:
                    # W
                    ws.cell(row=row, column=14).value = "" if detail.get('W','') in [0,None] else detail.get('W','')

                    # D
                    if detail.get('D','') == 0:
                        if detail.get('D1') == 0 and detail.get('D2') == 0:
                            ws.cell(row=row, column=15).value = ""
                        else:
                            ws.cell(row=row, column=15).value = "{}/{}".format(detail.get('D1',''),detail.get('D2',''))
                    else:
                        ws.cell(row=row, column=15).value = "" if detail.get('D','') in [0,None] else detail.get('D','')
                    
                    # H
                    if detail.get('H','') == 0:
                        if detail.get('H1') == 0 and detail.get('H2') == 0:
                            ws.cell(row=row, column=16).value = ""
                        else:
                            ws.cell(row=row, column=16).value = "{}/{}".format(detail.get('H1',''),detail.get('H2',''))
                    else:
                        ws.cell(row=row, column=16).value = "" if detail.get('H','') in [0,None] else detail.get('H','')
                    
                    # 漢字名称
                    if detail.get('見積区分') in ['C','A','S']:
                        ws.cell(row=row, column=6).value = "  {}".format(detail.get('漢字名称', ''))
                    else:
                        ws.cell(row=row, column=6).value = detail.get('漢字名称', '')

                ws.cell(row=row, column=17).value = "" if detail.get('発注数', 0) in [0,None] else str(detail.get('発注数', 0)) + ' ' + detail.get('単位名', '')

                if detail.get('仕入単価', '') != 0:
                    ws.cell(row=row, column=18).value = detail.get('仕入単価', '')

            # 挿入した表の最後尾を更新
            current_row = table_start_row + max_rows_per_table  # 現在の行位置を更新

            if table_num == total_tables - 1:  # 最後のループの場合
                # ここに最後に実行したい処理を記述

                # コピー元定義
                wb = create_excel_object('Temp見積依頼書.xlsx')
                src_sheet = wb['別紙_テンプレート']

                # コピー先ワークシート（「現在データ挿入中のワークシート」）
                dst_sheet = ws  # 目的のシートを指定

                # コピー範囲（A42からR61）
                min_col = 1  # A列
                min_row = 42  # 39行
                max_col = 18  # R列
                max_row = 61  # 61行

                # 表を挿入開始する位置（A41）
                shift_col = 0  # 列オフセット（A列に挿入するので0）
                shift_row = current_row - 41 # 新しい行の位置に挿入

                # コピー範囲と挿入
                range_copy_cell(
                    src_sheet,
                    dst_sheet,
                    min_col,
                    min_row,
                    max_col,
                    max_row,
                    shift_col,
                    shift_row,
                    True,
                )

                # 改ページ
                row_break = Break(table_start_row+20)
                ws.row_breaks.append(row_break)

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass