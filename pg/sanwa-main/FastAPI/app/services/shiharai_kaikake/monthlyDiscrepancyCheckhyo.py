from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer, range_copy_cell
from app.utils.service_utils import ClsTokuisaki
from app.core.config import settings
from app.core.exceptions import ServiceError
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill, Font

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class monthlyDiscrepancyCheckhyoService(BaseService):
    """月ズレチェック表サービス"""

    def display(self, request, session) -> StreamingResponse:

        # 仕入先検収明細表の印刷処理
        logger.info("【START】月ズレチェック表処理")

        storedname="usp_SK0300得意先別月ズレチェック表"

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=request.params, outputparams={"@RetDcnt":"INT", "@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            # wb = create_excel_object('Temp月ズレチェック表.xlsx')
            workbook = create_excel_object('Temp月ズレチェック表.xlsx')
            template_sheet = workbook['Template']

            # 初期化
            start_row = 3
            core_columns = 15
            current_client_cd = None
            current_row = start_row  # データ書き込み開始行

            # 作成日を設定
            template_sheet["L1"] = f"作成日: {datetime.now().strftime('%Y/%m/%d')}"            

            # データのループ
            for result in storedresults['results'][0]:
                client_cd = result['得意先CD']

                # 得意先CDごとに新しいシートを作成
                if client_cd != current_client_cd:
                    # 新シートに移る前にフィルタの設定を追加する
                    if current_client_cd is not None:
                        ws.auto_filter.ref = 'A2:O{}'.format(current_row-1)
                        # ws.auto_filter.ref = ws.dimensions
                    current_client_cd = client_cd
                    sheet_name = f"{client_cd}"
                    new_sheet = workbook.copy_worksheet(template_sheet)
                    ws = new_sheet
                    ws.title = sheet_name
                    workbook.active = ws
                    current_row = start_row  # 行カウンタをリセット

                    # 仕入先情報の取得
                    tokuisaki_info = ClsTokuisaki.GetbyID(session=session, tokuisakicd=client_cd)
                    ws["A1"] = tokuisaki_info[0]['略称']

                    # 行を調整
                    rec_cnt = result['Count']
                    if rec_cnt > 2:
                        # 不足している行を挿入し、書式をコピーする
                        ws.insert_rows(start_row + 1 , rec_cnt - 2)
                        original_height = ws.row_dimensions[start_row].height
                        for row in range(start_row + 1, start_row + rec_cnt - 1):
                            ws.row_dimensions[row].height = original_height
                            for col in range(1, core_columns + 1):
                                ws.cell(row=row, column=col)._style = copy(ws.cell(row=start_row, column=col)._style)

                    elif rec_cnt < 2:
                        # ダミー行が残ってしまうので削除
                        ws.delete_rows(start_row)

                # データ挿入
                ws.cell(row=current_row, column=1, value=result['見積番号'])
                ws.cell(row=current_row, column=2, value=result['日付'])
                ws.cell(row=current_row, column=3, value=result['見積件名'])
                ws.cell(row=current_row, column=4, value=result['行番号'])
                ws.cell(row=current_row, column=5, value=result['製品NO'])
                ws.cell(row=current_row, column=6, value=result['仕様NO'])
                ws.cell(row=current_row, column=7, value=result['漢字名称'])
                ws.cell(row=current_row, column=8, value=result['仕入先CD'])
                ws.cell(row=current_row, column=9, value=result['略称'])
                ws.cell(row=current_row, column=10, value=result['仕入数量'])
                ws.cell(row=current_row, column=11, value=result['売上数量'])
                ws.cell(row=current_row, column=12, value=result['仕入金額'])
                ws.cell(row=current_row, column=13, value=result['売上金額'])
                ws.cell(row=current_row, column=14, value=result['請求済FLG'])
                ws.cell(row=current_row, column=15, value=result['U区分'])

                current_row += 1

            # 最後のシートにフィルタを設定
            ws.auto_filter.ref = 'A2:O{}'.format(current_row-1)

            # 不要なテンプレートシートの削除
            workbook.remove(template_sheet)

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, workbook, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            ) 
       
    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
