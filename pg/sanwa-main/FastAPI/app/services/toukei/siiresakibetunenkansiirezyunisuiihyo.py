import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import range_copy_cell
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.borders import Border, Side

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class siiresakibetunenkansiirezyunisuiihyoService(BaseService):
    """仕入先別年間仕入順位推移表サービス"""

    def display(self, request, session) -> Dict[str, Any]:

        # 支払予定表の印刷処理
        logger.info("【START】仕入先別年間仕入順位推移表処理")

        storedname = "usp_TK0200仕入先別年間仕入順位推移表v2_年度"

        outputparams={"@RetDcnt":"INT", "@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=request.params, outputparams=outputparams))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        yearcount = int(request.params['@iE集計日付'][:4]) - int(request.params['@iS集計日付'][:4])

        wb = create_excel_object('Template仕入先別年間仕入順位推移表.xlsx')

        # TemplateWel請求明細は リースごとにシート　物件区分ごとに表
        with get_excel_buffer() as buffer:

            ws = wb.active

            ws.conditional_formatting._cf_rules.clear()

            ws.unmerge_cells('J3:O3')
            ws.unmerge_cells('J4:K4')

            start_row = 6
            row_num = start_row

            if yearcount == 1:
                column_width = {'J': 5.0, 'K': 7.2, 'M': 7.8, 'N': 12.9, 'O': 6.9, 'P': 8.8}
                ws.delete_cols(10,6)
                for col, width in column_width.items():
                    # 行の幅を変更
                    ws.column_dimensions[col].width = width

                ws.column_dimensions['L'].hidden = False
            else:
                copy_col_num = 6
                for i in range(1,yearcount):
                    range_copy_cell(ws, ws, 4, 3, 9, 7, copy_col_num, 0, True)
                    column_width = {'4': 5.0, '5': 7.2, '7': 7.8, '8': 12.9, '9': 6.9}
                    for col, width in column_width.items():
                        # 行の幅を変更
                        ws.column_dimensions['{}'.format(get_column_letter(int(col)+copy_col_num))].width = width
                    ws.column_dimensions['{}'.format(get_column_letter(6+copy_col_num))].hidden = True
                    copy_col_num += 6


            for row_data in storedresults['results'][0]:
                ws.cell(row=row_num, column=1, value=int(row_data['仕入先CD']))._style = copy(ws.cell(row=start_row, column=1)._style)
                ws.cell(row=row_num, column=2, value=row_data['仕入先名1'])._style = copy(ws.cell(row=start_row, column=2)._style)
                ws.cell(row=row_num, column=3, value=row_data['仕入先名2'])._style = copy(ws.cell(row=start_row, column=3)._style)

                column_offset = 0
                for year_num in range(0,yearcount):
                    ws.cell(row=3, column=4 + column_offset, value="{}年".format(int(request.params['@iS集計日付'][:4]) + year_num+1))._style = copy(ws.cell(row=3, column=4)._style)
                    ws.cell(row=row_num, column=4 + column_offset, value=row_data['順位{}'.format(year_num+1)])._style = copy(ws.cell(row=start_row, column=4 + column_offset)._style)
                    ws.cell(row=row_num, column=5 + column_offset, value="=IF(ISERROR({column_num}{row_num}*100/OFFSET({column_num}{row_num},0,-6)),0,IF(OFFSET({column_num}{row_num},0,-6)=0,0,{column_num}{row_num}*100/OFFSET({column_num}{row_num},0,-6)))".format(column_num=get_column_letter(8 + column_offset), row_num=row_num))._style = copy(ws.cell(row=start_row, column=5 + column_offset)._style)
                    ws.cell(row=row_num, column=6 + column_offset, value=row_data['税抜金額{}'.format(year_num+1)])._style = copy(ws.cell(row=start_row, column=6 + column_offset)._style)
                    ws.cell(row=4, column=6 + column_offset, value="=SUM({column_num}6:{column_num}{row_num})".format(column_num=get_column_letter(6 + column_offset), row_num=row_num))
                    if row_data['税抜金額{}'.format(year_num+1)] == 0:
                        ws.cell(row=row_num, column=7 + column_offset, value="")._style = copy(ws.cell(row=start_row, column=7 + column_offset)._style)
                    else:
                        ws.cell(row=row_num, column=7 + column_offset, value=row_data['原価金額{}'.format(year_num+1)] * 100 / row_data['税抜金額{}'.format(year_num+1)])._style = copy(ws.cell(row=start_row, column=7 + column_offset)._style)
                    ws.cell(row=4, column=7 + column_offset, value="=IF(ISERROR({column_num2}4*100/{column_num1}4),0,{column_num2}4*100/{column_num1}4)".format(column_num1=get_column_letter(6 + column_offset), column_num2=get_column_letter(8 + column_offset)))
                    ws.cell(row=row_num, column=8 + column_offset, value=row_data['原価金額{}'.format(year_num+1)])._style = copy(ws.cell(row=start_row, column=8 + column_offset)._style)
                    ws.cell(row=4, column=8 + column_offset, value="=SUM({column_num}6:{column_num}{row_num})".format(column_num=get_column_letter(8 + column_offset), row_num=row_num))
                    ws.cell(row=row_num, column=9 + column_offset, value="=IF({column_num}{row_num}=0,0,{column_num}{row_num}*100/{column_num}$4)".format(column_num=get_column_letter(8 + column_offset), row_num=row_num))._style = copy(ws.cell(row=start_row, column=9 + column_offset)._style)

                    column_offset += 6
                row_num += 1

            ws['C1'] = "{}年4月～{}年3月".format(request.params['@iS集計日付'][:4], request.params['@iE集計日付'][:4])

            side = Side(style='medium', color='000000')
            border_r = Border(right=side, top=side)
            ws.cell(row=3, column=4 + column_offset -1).border = border_r

            if yearcount != 0:
                ws.column_dimensions['{}'.format(get_column_letter(10 + column_offset -6))].width = 0.64

            # 行の高さを指定
            ws.row_dimensions[2].height = 14.25
            ws.row_dimensions[3].height = 13.5
            ws.row_dimensions[4].height = 13.5
            ws.row_dimensions[5].height = 14.25

            ws.print_area = 'A1:{}{}'.format(get_column_letter(4 + column_offset ), row_num-1)


            # 条件付き書式の設定
            column_offset = 0
            for year_num in range(1,yearcount+1):

                row_num = 6

                # 原価率合計
                rule1 = FormulaRule(formula = ["={}4>=90".format(get_column_letter(7 + column_offset))], font = Font(color='FF0000'))
                ws.conditional_formatting.add('{}{}'.format(get_column_letter(7 + column_offset), start_row-2), rule1)

                for row_data in storedresults['results'][0]:
                    # 原価率
                    rule1_1 = FormulaRule(formula = ["=IF(ISERROR({column}{row}*100/OFFSET({column}{row},0,-6)),FALSE,{current_column}{row}<100)".format(column=get_column_letter(8 + column_offset),row=row_num,current_column=get_column_letter(7 + column_offset))], font = Font(color='FF0000'))
                    ws.conditional_formatting.add('{}{}'.format(get_column_letter(7 + column_offset), row_num), rule1_1)

                    # 順位
                    rule2 = FormulaRule(formula = ["={column}{row}>OFFSET({column}{row},0,-4)".format(column=get_column_letter(4 + column_offset),row=row_num)], font = Font(color='FF0000'))
                    ws.conditional_formatting.add('{}{}'.format(get_column_letter(4 + column_offset),row_num), rule2)

                    # 前対比
                    rule3 = FormulaRule(formula = ["=IF(ISERROR({column}{row}*100/OFFSET({column}{row},0,-4)),FALSE,{current_column}{row}<100)".format(column=get_column_letter(8 + column_offset),row=row_num,current_column=get_column_letter(5 + column_offset))], font = Font(color='FF0000'))
                    ws.conditional_formatting.add('{}{}'.format(get_column_letter(5 + column_offset), row_num), rule3)

                    # 純売上金額
                    rule4 = FormulaRule(formula = ["=IF(ISERROR({column}{row}*100/OFFSET({column}{row},0,-6)),FALSE,{current_column}{row}<100)".format(column=get_column_letter(8 + column_offset),row=row_num,current_column=get_column_letter(5 + column_offset))], font = Font(color='FF0000'))
                    ws.conditional_formatting.add('{}{}'.format(get_column_letter(8 + column_offset), row_num), rule4)

                    row_num += 1

                column_offset += 6

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            ws.title = "仕入先別年間仕入順位推移表"


            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, 'sanwa55')

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )



    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
