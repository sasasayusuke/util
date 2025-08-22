from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer, range_copy_cell
from app.utils.service_utils import ClsTokuisaki
from app.core.config import settings
from app.core.exceptions import ServiceError
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill, Font

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class ConstructionValueByClientService(BaseService):
    """顧客別施工金額実績表サービス"""

    def display(self, request, session) -> StreamingResponse:

        # 入庫リストの印刷処理
        logger.info("【START】顧客別施工金額実績表処理")

        storedname="usp_TK3000顧客別施工金額実績表"

        params = {
            # "@i入庫日":request.params["@i入庫日"],
            "@iS集計日付":request.params["@iS集計日付"],
            "@iE集計日付":request.params["@iE集計日付"],
            "@iS得意先CD":request.params["@iS得意先CD"],
            "@iE得意先CD":request.params["@iE得意先CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetDcnt":"INT", "@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            workbook = create_excel_object('Temp施工金額実績表.xlsx')
            template_sheet = workbook['Template']

            # 初期化
            start_row = 4
            start_core_columns = 3
            end_core_columns = 13
            current_tokuisaki_code = None
            current_mitumori_number = None

            doubleSide = Side(style='double')
            solidSide = Side(style='thin')

            doubleBorder = Border(left=solidSide,right=solidSide,bottom=solidSide,top=doubleSide)
            solidBorder = Border(left=solidSide,right=solidSide,bottom=solidSide,top=solidSide)

            # データのループ
            for result in storedresults['results'][0]:
                tokuisaki_code = result['得意先CD']
                mitumori_number = result['見積番号']

                # 得意先コードごとに新しいシートを作成
                if tokuisaki_code != current_tokuisaki_code:
                    current_tokuisaki_code = tokuisaki_code
                    sheet_name = f"{tokuisaki_code}"
                    new_sheet = workbook.copy_worksheet(template_sheet)
                    ws = new_sheet
                    ws.title = sheet_name
                    workbook.active = ws
                    current_row = start_row  # 行カウンタをリセット
                    ws.sheet_view.view = 'pageBreakPreview'

                    # 行を調整
                    rec_cnt = result['Count']
                    if rec_cnt > 2:
                        # 不足している行を挿入し、書式をコピーする
                        ws.insert_rows(start_row + 1 , rec_cnt - 2)
                        original_height = ws.row_dimensions[start_row].height
                        for row in range(start_row + 1, start_row + rec_cnt - 1):
                            ws.row_dimensions[row].height = original_height
                            for col in range(start_core_columns, end_core_columns + 1):
                                ws.cell(row=row, column=col)._style = copy(ws.cell(row=start_row, column=col)._style)

                    elif rec_cnt < 2:
                        # ダミー行が残ってしまうので削除
                        ws.delete_rows(start_row)

                    # 先頭行の関数を変更
                    end_row = start_row + rec_cnt - 1
                    ws.cell(row=3, column=11, value=f'=SUMIF(J4:J{end_row},"<>",K4:K{end_row})')
                    ws.cell(row=3, column=12, value=f'=SUMIF(J4:J{end_row},"=",L4:L{end_row})')

                # データ挿入
                # 同一見積番号ならば表示しない
                if mitumori_number != current_mitumori_number:
                    current_mitumori_number = mitumori_number
                    ws.cell(row=current_row, column=1, value=result['チームCD'])
                    ws.cell(row=current_row, column=2, value=result['売仕区分'])
                    ws.cell(row=current_row, column=3, value=result['得意先CD'])
                    ws.cell(row=current_row, column=4, value=result['見積番号'])
                    ws.cell(row=current_row, column=5, value=result['納入得意先CD'])
                    ws.cell(row=current_row, column=6, value=result['納入先CD'])
                    ws.cell(row=current_row, column=7, value=result['表示用店舗名'])
                    ws.cell(row=current_row, column=8, value=result['見積件名'])

                    # 二重線に。
                    for row in ws['A{}'.format(current_row):'M{}'.format(current_row)]:
                        for cell in row:
                            cell.border = doubleBorder
                else:
                    current_mitumori_number = mitumori_number
                    for row in ws['A{}'.format(current_row):'M{}'.format(current_row)]:
                        for cell in row:
                            cell.border = solidBorder
                
                ws.cell(row=current_row, column=9, value=result['仕入先CD'])
                ws.cell(row=current_row, column=10, value=result['仕入先名1'])
                ws.cell(row=current_row, column=11, value=result['仕入金額'])
                
                # 施工内容の場合、施工仕入金額のみ表示
                if result['売仕区分'] != 2:
                    ws.cell(row=current_row, column=12, value=result['税抜金額'])
                ws.cell(row=current_row, column=13, value=f'=IF(ISERROR((K{current_row}/L{current_row})*100),"",(K{current_row}/L{current_row})*100)')

                current_row += 1

            # 不要なテンプレートシートの削除
            workbook.remove(template_sheet)

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, workbook, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            ) 
       
    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
