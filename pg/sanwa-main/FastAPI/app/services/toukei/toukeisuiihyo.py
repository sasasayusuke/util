from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer, range_copy_cell, relative_reference_copy, relative_reference_range_copy
from app.core.config import settings
from app.core.exceptions import ServiceError
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter


# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class toukeisuiihyoService(BaseService):
    """統計推移表サービス"""

    def display(self, request, session) -> Dict[str, Any]:

        # 統計推移表の印刷処理
        logger.info("【START】統計推移表処理")

        storedname="usp_TK3200統計推移表V12"

        #24ヶ月以上の範囲を指定されている場合はエラーとする
        start_date = datetime.strptime(request.params['@iS集計日付'], '%Y/%m/%d')
        end_date = datetime.strptime(request.params['@iE集計日付'], '%Y/%m/%d')
        period_months = abs( start_date.year - end_date.year )*12 + ( end_date.month - start_date.month )
        if period_months >= 24:
            raise ServiceError('範囲が大きすぎます。（最大24ヶ月）')

        logger.info('取得開始')
        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=request.params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        logger.info('取得終了')
        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        if storedresults["output_values"]["@RetST"] == -1:
            raise ServiceError(storedresults["output_values"]["@RetMsg"])

        with get_excel_buffer() as buffer:
            wb = create_excel_object('Temp統計推移表v14.xlsx')

            # DBシートにデータを登録
            ws = wb['DB']

            # 顧客別　売上・原価
            row_num = 3
            for row_data in storedresults['results'][0]:
                ws.cell(row=row_num, column=1, value=row_data['営業担当CD'])
                ws.cell(row=row_num, column=2, value=row_data['番号'])
                ws.cell(row=row_num, column=3, value=row_data['集計CD'])
                ws.cell(row=row_num, column=4, value=row_data['年月'])
                ws.cell(row=row_num, column=5, value=row_data['作業区分CD'])
                ws.cell(row=row_num, column=6, value=row_data['売上税抜金額'])
                ws.cell(row=row_num, column=7, value=row_data['原価税抜金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 工事　売上・原価
            row_num = 3
            for row_data in storedresults['results'][1]:
                ws.cell(row=row_num, column=10, value=row_data['工事担当CD'])
                ws.cell(row=row_num, column=11, value=row_data['番号'])
                ws.cell(row=row_num, column=12, value=row_data['集計CD'])
                ws.cell(row=row_num, column=13, value=row_data['年月'])
                ws.cell(row=row_num, column=14, value=row_data['作業区分CD'])
                ws.cell(row=row_num, column=15, value=row_data['売上税抜金額'])
                ws.cell(row=row_num, column=16, value=row_data['原価税抜金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 入出庫
            row_num = 3
            for row_data in storedresults['results'][2]:
                ws.cell(row=row_num, column=25, value=row_data['集計CD'])
                ws.cell(row=row_num, column=26, value=row_data['年月'])
                ws.cell(row=row_num, column=27, value=row_data['出庫金額'])
                ws.cell(row=row_num, column=28, value=row_data['入庫金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 三研　売上（円）
            row_num = 3
            for row_data in storedresults['results'][3]:
                ws.cell(row=row_num, column=31, value=row_data['年月'])
                ws.cell(row=row_num, column=32, value=row_data['売上税抜金額_円'])
                ws.cell(row=row_num, column=33, value=row_data['原価税抜金額_円'])
                ws.cell(row=row_num, column=34, value=row_data['人件一般費_円'])
                ws.cell(row=row_num, column=35, value=row_data['営業外費_円'])

                # データを書き込んだ後、次の行に移動
                row_num += 1

            # サービス　売上（円）
            row_num = 3
            for row_data in storedresults['results'][4]:
                ws.cell(row=row_num, column=67, value=row_data['年月'])
                ws.cell(row=row_num, column=68, value=row_data['売額_貨物自_三商'])
                ws.cell(row=row_num, column=69, value=row_data['原額_貨物自_三商'])
                ws.cell(row=row_num, column=70, value=row_data['売額_貨物自_店装'])
                ws.cell(row=row_num, column=71, value=row_data['原額_貨物自_店装'])
                ws.cell(row=row_num, column=72, value=row_data['売額_貨物自_サンプラ'])
                ws.cell(row=row_num, column=73, value=row_data['原額_貨物自_サンプラ'])
                ws.cell(row=row_num, column=74, value=row_data['売額_貨物自_三シャ'])
                ws.cell(row=row_num, column=75, value=row_data['原額_貨物自_三シャ'])
                ws.cell(row=row_num, column=76, value=row_data['売額_貨物自_トラン'])
                ws.cell(row=row_num, column=77, value=row_data['原額_貨物自_トラン'])
                ws.cell(row=row_num, column=78, value=row_data['売額_貨物自_その他'])
                ws.cell(row=row_num, column=79, value=row_data['原額_貨物自_その他'])
                ws.cell(row=row_num, column=80, value=row_data['売額_貨物自_事故'])
                ws.cell(row=row_num, column=81, value=row_data['原額_貨物自_事故'])
                ws.cell(row=row_num, column=82, value=row_data['売額_貨物庸_三商'])
                ws.cell(row=row_num, column=83, value=row_data['原額_貨物庸_三商'])
                ws.cell(row=row_num, column=84, value=row_data['売額_貨物庸_店装'])
                ws.cell(row=row_num, column=85, value=row_data['原額_貨物庸_店装'])
                ws.cell(row=row_num, column=86, value=row_data['売額_貨物庸_サンプラ'])
                ws.cell(row=row_num, column=87, value=row_data['原額_貨物庸_サンプラ'])
                ws.cell(row=row_num, column=88, value=row_data['売額_貨物庸_三シャ'])
                ws.cell(row=row_num, column=89, value=row_data['原額_貨物庸_三シャ'])
                ws.cell(row=row_num, column=90, value=row_data['売額_貨物庸_トラン'])
                ws.cell(row=row_num, column=91, value=row_data['原額_貨物庸_トラン'])
                ws.cell(row=row_num, column=92, value=row_data['売額_貨物庸_その他'])
                ws.cell(row=row_num, column=93, value=row_data['原額_貨物庸_その他'])
                ws.cell(row=row_num, column=94, value=row_data['売額_貨物庸_事故'])
                ws.cell(row=row_num, column=95, value=row_data['原額_貨物庸_事故'])
                ws.cell(row=row_num, column=96, value=row_data['売額_物流_ルート'])
                ws.cell(row=row_num, column=97, value=row_data['原額_物流_ルート'])
                ws.cell(row=row_num, column=98, value=row_data['売額_物流_丸和'])
                ws.cell(row=row_num, column=99, value=row_data['原額_物流_丸和'])
                ws.cell(row=row_num, column=100, value=row_data['売額_物流_その他'])
                ws.cell(row=row_num, column=101, value=row_data['原額_物流_その他'])
                ws.cell(row=row_num, column=102, value=row_data['売額_物流_事故'])
                ws.cell(row=row_num, column=103, value=row_data['原額_物流_事故'])
                ws.cell(row=row_num, column=104, value=row_data['人件一般費_貨物'])
                ws.cell(row=row_num, column=105, value=row_data['営業外費_貨物'])
                ws.cell(row=row_num, column=106, value=row_data['人件一般費_物流'])
                ws.cell(row=row_num, column=107, value=row_data['営業外費_物流'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 営業担当別経費
            row_num = 3
            for row_data in storedresults['results'][5]:
                ws.cell(row=row_num, column=110, value=row_data['業種区分_未使用'])
                ws.cell(row=row_num, column=111, value=row_data['部CD_未使用'])
                ws.cell(row=row_num, column=112, value=row_data['年月'])
                ws.cell(row=row_num, column=113, value=row_data['人件費一般費'])
                # データを書き込んだ後、次の行に移動
                row_num += 1

            # 部署別　売上・原価
            row_num = 3
            for row_data in storedresults['results'][6]:
                ws.cell(row=row_num, column=117, value=row_data['部署CD'])
                ws.cell(row=row_num, column=118, value=row_data['年月'])
                ws.cell(row=row_num, column=119, value=row_data['作業区分CD'])
                ws.cell(row=row_num, column=120, value=row_data['売上税抜金額'])
                ws.cell(row=row_num, column=121, value=row_data['原価税抜金額'])
                # データを書き込んだ後、次の行に移動
                row_num += 1
            
            ws.print_area = 'A1:T{}'.format(len(storedresults['results'][0])+2)
            ws.sheet_view.view = 'pageBreakPreview'

            # グループシート関数追加
            ws = wb['グループ']
            current_col = 9
            current_date = end_date.replace(day=1)
            # 列を追加
            for i in range(period_months + 1):
                if i != 0:
                    # 書式・数式を反映
                    relative_reference_range_copy(ws,ws,current_col-3,3,current_col-1,35,3,0)
                    range_copy_cell(ws,ws,current_col-3,2,current_col-1,35,3,0,False)
                    # 別シート参照うまくコピーできないため別途定義する
                    ws.cell(row=16, column=current_col, value="=IFERROR(INDEX(DB!三研_原価税抜金額,MATCH({}$2,INDEX(DB!三研_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=16, column=current_col+2, value="=IFERROR(INDEX(DB!三研_売上税抜金額,MATCH({}$2,INDEX(DB!三研_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=17, column=current_col, value="={t1}16-{t2}16-(IFERROR(INDEX(DB!三研_人件一般費,MATCH({t2}$2,INDEX(DB!三研_年月,),0)),0))".format(t1=get_column_letter(current_col+2),t2=get_column_letter(current_col)))
                    # 列幅を調整
                    ws.column_dimensions["{}".format(get_column_letter(current_col))].width = 9.9
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 1))].width = 5.7
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 2))].width = 9.9
                # 日付追加
                ws['{}2'.format(get_column_letter(current_col))].number_format = 'yyyy年m月'
                ws['{}2'.format(get_column_letter(current_col))].value = current_date
                current_col += 3
                current_date = current_date + relativedelta(months=-1)
            last_col = current_col-1
            last_col_alpha = get_column_letter(last_col)
            # 合計の関数を修正
            ws.cell(row=4, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I4:{t1}4)".format(t1=last_col_alpha))
            ws.cell(row=4, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I4:{t1}4)".format(t1=last_col_alpha))
            ws.cell(row=5, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I5:{t1}5)".format(t1=last_col_alpha))
            # 株式会社　三和商研
            ws.cell(row=10, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I10:{t1}10)".format(t1=last_col_alpha))
            ws.cell(row=10, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I10:{t1}10)".format(t1=last_col_alpha))
            ws.cell(row=11, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I11:{t1}11)".format(t1=last_col_alpha))
            # 三研(上海)商貿有限公司
            ws.cell(row=16, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I16:{t1}16)".format(t1=last_col_alpha))
            ws.cell(row=16, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I16:{t1}16)".format(t1=last_col_alpha))
            ws.cell(row=17, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I17:{t1}17)".format(t1=last_col_alpha))
            # 株式会社　三和サービス
            ws.cell(row=22, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I22:{t1}22)".format(t1=last_col_alpha))
            ws.cell(row=22, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I22:{t1}22)".format(t1=last_col_alpha))
            ws.cell(row=23, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I23:{t1}23)".format(t1=last_col_alpha))
            # 株式会社　三和
            ws.cell(row=28, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I28:{t1}28)".format(t1=last_col_alpha))
            ws.cell(row=28, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I28:{t1}28)".format(t1=last_col_alpha))
            ws.cell(row=29, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I29:{t1}29)".format(t1=last_col_alpha))

            # ws.print_area = 'A1:{}35'.format(get_column_letter(last_col))
            ws.print_area = 'A1:T35'
            ws.sheet_view.view = 'pageBreakPreview'

            # 商研シート関数追加
            ws = wb['商研']
            current_col = 9
            current_date = end_date.replace(day=1)
            # 列を追加
            for i in range(period_months + 1):
                if i != 0:
                    # 書式・数式を反映
                    relative_reference_range_copy(ws,ws,current_col-3,3,current_col-1,117,3,0)
                    range_copy_cell(ws,ws,current_col-3,2,current_col-1,117,3,0,False)
                    # 列幅を調整
                    ws.column_dimensions["{}".format(get_column_letter(current_col))].width = 10.1
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 1))].width = 5.7
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 2))].width = 10.1

                # 販管費の数式代入(relative_reference_range_copyではシート参照が消えるため)
                ws['{}7'.format(get_column_letter(current_col))].value = '=IFERROR(INDEX(DB!統計集計先_人件費一般費,MATCH($C7&"-"&{}$2,INDEX(DB!統計集計先_CD&"-"&DB!統計集計先_年月,),0)),0)'.format(get_column_letter(current_col))
                # 日付追加
                ws['{}2'.format(get_column_letter(current_col))].number_format = 'yyyy年m月'
                ws['{}2'.format(get_column_letter(current_col))].value = current_date
                current_col += 3
                current_date = current_date + relativedelta(months=-1)
            last_col = current_col-1
            last_col_alpha = get_column_letter(last_col)
            # 合計の関数を修正
            ws.cell(row=4, column=6, value="=SUM(I4:{}4)".format(last_col_alpha))
            ws.cell(row=5, column=8, value="=SUM(I5:{}5)".format(last_col_alpha))
            ws.cell(row=6, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I6:{t1}6)".format(t1=last_col_alpha))
            ws.cell(row=7, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I7:{t1}7)".format(t1=last_col_alpha))
            ws.cell(row=8, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I8:{t1}8)".format(t1=last_col_alpha))
            ws.cell(row=9, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I9:{t1}9)".format(t1=last_col_alpha))
            ws.cell(row=9, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I9:{t1}9)".format(t1=last_col_alpha))
            ws.cell(row=11, column=8, value="=SUM(I11:{}11)".format(last_col_alpha))
            for i in range(13, 19):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=20, column=8, value="=SUM(I20:{}20)".format(last_col_alpha))
            for i in range(22, 28):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=29, column=8, value="=SUM(I29:{}29)".format(last_col_alpha))
            for i in range(31, 117):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))

            ws.print_area = 'A1:T117'
            # ws.print_area = 'A1:{}107'.format(get_column_letter(last_col))
            ws.sheet_view.view = 'pageBreakPreview'

            # 工事シート関数追加
            ws = wb['工事']
            current_col = 9
            current_date = end_date.replace(day=1)
            # 列を追加
            for i in range(period_months + 1):
                if i != 0:
                    # 書式・数式を反映
                    relative_reference_range_copy(ws,ws,current_col-3,3,current_col-1,75,3,0)
                    range_copy_cell(ws,ws,current_col-3,2,current_col-1,75,3,0,False)
                    # 列幅を調整
                    ws.column_dimensions["{}".format(get_column_letter(current_col))].width = 10.1
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 1))].width = 5.7
                    ws.column_dimensions["{}".format(get_column_letter(current_col + 2))].width = 10.1
                # 日付追加
                ws['{}2'.format(get_column_letter(current_col))].number_format = 'yyyy年m月'
                ws['{}2'.format(get_column_letter(current_col))].value = current_date
                current_col += 3
                current_date = current_date + relativedelta(months=-1)
            last_col = current_col-1
            last_col_alpha = get_column_letter(last_col)
            # 合計の関数を修正
            ws.cell(row=4, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I4:{t1}4)".format(t1=last_col_alpha))
            ws.cell(row=4, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I4:{t1}4)".format(t1=last_col_alpha))
            ws.cell(row=5, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I5:{t1}5)".format(t1=last_col_alpha))
            ws.cell(row=6, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I6:{t1}6)".format(t1=last_col_alpha))
            ws.cell(row=6, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I6:{t1}6)".format(t1=last_col_alpha))
            for i in range(8, 14):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=14, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I14:{t1}14)".format(t1=last_col_alpha))
            ws.cell(row=15, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I15:{t1}15)".format(t1=last_col_alpha))
            ws.cell(row=15, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I15:{t1}15)".format(t1=last_col_alpha))
            for i in range(17, 21):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=21, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I21:{t1}21)".format(t1=last_col_alpha))
            ws.cell(row=22, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I22:{t1}22)".format(t1=last_col_alpha))
            ws.cell(row=22, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I22:{t1}22)".format(t1=last_col_alpha))
            ws.cell(row=23, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I23:{t1}23)".format(t1=last_col_alpha))
            ws.cell(row=24, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I24:{t1}24)".format(t1=last_col_alpha))
            ws.cell(row=24, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I24:{t1}24)".format(t1=last_col_alpha))
            for i in range(26, 30):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=30, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I30:{t1}30)".format(t1=last_col_alpha))
            ws.cell(row=31, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I31:{t1}31)".format(t1=last_col_alpha))
            ws.cell(row=31, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I31:{t1}31)".format(t1=last_col_alpha))
            ws.cell(row=32, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I32:{t1}32)".format(t1=last_col_alpha))
            ws.cell(row=33, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I33:{t1}33)".format(t1=last_col_alpha))
            ws.cell(row=33, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I33:{t1}33)".format(t1=last_col_alpha))
            for i in range(35, 39):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=39, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I39:{t1}39)".format(t1=last_col_alpha))
            ws.cell(row=40, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I40:{t1}40)".format(t1=last_col_alpha))
            ws.cell(row=40, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I40:{t1}40)".format(t1=last_col_alpha))
            ws.cell(row=41, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I41:{t1}41)".format(t1=last_col_alpha))
            ws.cell(row=42, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I42:{t1}42)".format(t1=last_col_alpha))
            ws.cell(row=42, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I42:{t1}42)".format(t1=last_col_alpha))
            for i in range(44, 48):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=48, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I48:{t1}48)".format(t1=last_col_alpha))
            ws.cell(row=49, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I49:{t1}49)".format(t1=last_col_alpha))
            ws.cell(row=49, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I49:{t1}49)".format(t1=last_col_alpha))
            ws.cell(row=50, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I50:{t1}50)".format(t1=last_col_alpha))
            ws.cell(row=51, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I51:{t1}51)".format(t1=last_col_alpha))
            ws.cell(row=51, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I51:{t1}51)".format(t1=last_col_alpha))
            for i in range(53, 57):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=57, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I57:{t1}57)".format(t1=last_col_alpha))
            ws.cell(row=58, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I58:{t1}58)".format(t1=last_col_alpha))
            ws.cell(row=58, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I58:{t1}58)".format(t1=last_col_alpha))
            ws.cell(row=59, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I59:{t1}59)".format(t1=last_col_alpha))
            ws.cell(row=60, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I60:{t1}60)".format(t1=last_col_alpha))
            ws.cell(row=60, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I60:{t1}60)".format(t1=last_col_alpha))
            for i in range(62, 66):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=66, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I66:{t1}66)".format(t1=last_col_alpha))
            ws.cell(row=67, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I67:{t1}67)".format(t1=last_col_alpha))
            ws.cell(row=67, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I67:{t1}67)".format(t1=last_col_alpha))
            ws.cell(row=68, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I68:{t1}68)".format(t1=last_col_alpha))
            ws.cell(row=69, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I69:{t1}69)".format(t1=last_col_alpha))
            ws.cell(row=69, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I69:{t1}69)".format(t1=last_col_alpha))
            for i in range(71, 75):
                ws.cell(row=i, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
                ws.cell(row=i, column=8, value="=SUMIF(I$3:{t1}$3,$H$3,I{n1}:{t1}{n2})".format(t1=last_col_alpha,n1=i,n2=i))
            ws.cell(row=75, column=6, value="=SUMIF(I$3:{t1}$3,$F$3,I75:{t1}75)".format(t1=last_col_alpha))

            # ws.print_area = 'A1:{}75'.format(get_column_letter(last_col))
            ws.print_area = 'A1:T75'
            ws.sheet_view.view = 'pageBreakPreview'

            # 三和サービスシート関数追加
            ws = wb['三和サービス']
            current_col = 8
            current_date = end_date.replace(day=1)
            # 列を追加
            for i in range(period_months + 1):
                if i != 0:
                    # 書式・数式を反映
                    relative_reference_range_copy(ws,ws,current_col-3,3,current_col-1,56,3,0)
                    range_copy_cell(ws,ws,current_col-3,2,current_col-1,56,3,0,False)
                    # 別シート参照うまくコピーできないため別途定義する
                    ws.cell(row=22, column=current_col, value="={t1}20-{t2}20-(IFERROR(INDEX(DB!サービス_人件一般費_貨物,MATCH({t2}$2,INDEX(DB!サービス_年月,),0)),0))".format(t1=get_column_letter(current_col+2),t2=get_column_letter(current_col)))
                    ws.cell(row=24, column=current_col, value="={t1}20-{t2}20-(IFERROR(INDEX(DB!サービス_人件一般費_貨物,MATCH({t2}$2,INDEX(DB!サービス_年月,),0)),0))-(IFERROR(INDEX(DB!サービス_営業外費_貨物,MATCH({t2}$2,INDEX(DB!サービス_年月,),0)),0))".format(t1=get_column_letter(current_col+2),t2=get_column_letter(current_col)))
                    ws.cell(row=27, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_三商,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=27, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_三商,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=28, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_店装,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=28, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_店装,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=29, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_サンプラ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=29, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_サンプラ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=30, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_トラン,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=30, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_トラン,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=31, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_三シャ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=31, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_三シャ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=32, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=32, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=33, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物自_事故,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=33, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物自_事故,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=35, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_三商,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=35, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_三商,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=36, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_店装,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=36, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_店装,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=37, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_サンプラ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=37, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_サンプラ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=38, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_トラン,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=38, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_トラン,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=39, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_三シャ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=39, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_三シャ,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=40, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=40, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=41, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_事故,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=41, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_事故,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=42, column=current_col, value="=IFERROR(INDEX(DB!サービス予測_予測原価金額_貨物,MATCH({}$2,INDEX(DB!サービス予測_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=42, column=current_col+2, value="=IFERROR(INDEX(DB!サービス予測_予測売上金額_貨物,MATCH({}$2,INDEX(DB!サービス予測_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=47, column=current_col, value="={t1}45-{t2}45-(IFERROR(INDEX(DB!サービス_人件一般費_物流,MATCH({t2}$2,INDEX(DB!サービス_年月,),0)),0))".format(t1=get_column_letter(current_col+2),t2=get_column_letter(current_col)))
                    ws.cell(row=49, column=current_col, value="={t1}45-{t2}45-(IFERROR(INDEX(DB!サービス_人件一般費_物流,MATCH({t2}$2,INDEX(DB!サービス_年月,),0)),0))-(IFERROR(INDEX(DB!サービス_営業外費_物流,MATCH({t2}$2,INDEX(DB!サービス_年月,),0)),0))".format(t1=get_column_letter(current_col+2),t2=get_column_letter(current_col)))
                    # ws.cell(row=50, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_貨物庸_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    # ws.cell(row=50, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_貨物庸_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=51, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_物流_ルート,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=51, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_物流_ルート,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=52, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_物流_丸和,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=52, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_物流_丸和,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=53, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_物流_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=53, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_物流_その他,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=54, column=current_col, value="=IFERROR(INDEX(DB!サービス_原額_物流_事故,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=54, column=current_col+2, value="=IFERROR(INDEX(DB!サービス_売額_物流_事故,MATCH({}$2,INDEX(DB!サービス_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=55, column=current_col, value="=IFERROR(INDEX(DB!サービス予測_予測原価金額_物流,MATCH({}$2,INDEX(DB!サービス予測_年月,),0)),0)".format(get_column_letter(current_col)))
                    ws.cell(row=55, column=current_col+2, value="=IFERROR(INDEX(DB!サービス予測_予測売上金額_物流,MATCH({}$2,INDEX(DB!サービス予測_年月,),0)),0)".format(get_column_letter(current_col)))

                # 日付追加
                ws['{}2'.format(get_column_letter(current_col))].number_format = 'yyyy年m月'
                ws['{}2'.format(get_column_letter(current_col))].value = current_date
                current_col += 3
                current_date = current_date + relativedelta(months=-1)
            last_col = current_col-1
            last_col_alpha = get_column_letter(last_col)

            # 合計
            ws['E4'].value = '=SUMIF(H$3:{t1}$3,$E$3,H4:{t1}4)'.format(t1=last_col_alpha)
            ws['G4'].value = '=SUMIF(H$3:{t1}$3,$G$3,H4:{t1}4)'.format(t1=last_col_alpha)
            ws['E6'].value = '=SUMIF(H$3:{t1}$3,$E$3,H6:{t1}6)'.format(t1=last_col_alpha)
            ws.cell(row=8,column=5,value='=SUMIF(H$3:{t1}$3,$E$3,H8:{t1}8)'.format(t1=last_col_alpha))
            # ws['E8'].value = '=SUMIF(H$3:{t1}$3,$E$3,H8:{t1}8)'.format(t1=last_col_alpha)

            ws['E10'].value = '=SUMIF(H$3:{t1}$3,$E$3,H10:{t1}10)'.format(t1=last_col_alpha)
            ws['G10'].value = '=SUMIF(H$3:{t1}$3,$G$3,H10:{t1}10)'.format(t1=last_col_alpha)

            ws['E11'].value = '=SUMIF(H$3:{t1}$3,$E$3,H11:{t1}11)'.format(t1=last_col_alpha)
            ws['G11'].value = '=SUMIF(H$3:{t1}$3,$G$3,H11:{t1}11)'.format(t1=last_col_alpha)

            ws['E13'].value = '=SUMIF(H$3:{t1}$3,$E$3,H13:{t1}13)'.format(t1=last_col_alpha)
            ws['G13'].value = '=SUMIF(H$3:{t1}$3,$G$3,H13:{t1}13)'.format(t1=last_col_alpha)

            ws['E15'].value = '=SUMIF(H$3:{t1}$3,$E$3,H15:{t1}15)'.format(t1=last_col_alpha)
            ws['G15'].value = '=SUMIF(H$3:{t1}$3,$G$3,H15:{t1}15)'.format(t1=last_col_alpha)

            ws['E17'].value = '=SUMIF(H$3:{t1}$3,$E$3,H17:{t1}17)'.format(t1=last_col_alpha)
            ws['G17'].value = '=SUMIF(H$3:{t1}$3,$G$3,H17:{t1}17)'.format(t1=last_col_alpha)

            ws['E20'].value = '=SUMIF(H$3:{t1}$3,$E$3,H20:{t1}20)'.format(t1=last_col_alpha)
            ws['G20'].value = '=SUMIF(H$3:{t1}$3,$G$3,H20:{t1}20)'.format(t1=last_col_alpha)

            ws['E22'].value = '=SUMIF(H$3:{t1}$3,$E$3,H22:{t1}22)'.format(t1=last_col_alpha)

            ws['E24'].value = '=SUMIF(H$3:{t1}$3,$E$3,H24:{t1}24)'.format(t1=last_col_alpha)

            ws['E26'].value = '=SUMIF(H$3:{t1}$3,$E$3,H26:{t1}26)'.format(t1=last_col_alpha)
            ws['E27'].value = '=SUMIF(H$3:{t1}$3,$E$3,H27:{t1}27)'.format(t1=last_col_alpha)
            ws['E28'].value = '=SUMIF(H$3:{t1}$3,$E$3,H28:{t1}28)'.format(t1=last_col_alpha)
            ws['E29'].value = '=SUMIF(H$3:{t1}$3,$E$3,H29:{t1}29)'.format(t1=last_col_alpha)
            ws['E30'].value = '=SUMIF(H$3:{t1}$3,$E$3,H30:{t1}30)'.format(t1=last_col_alpha)
            ws['E31'].value = '=SUMIF(H$3:{t1}$3,$E$3,H31:{t1}31)'.format(t1=last_col_alpha)
            ws['E32'].value = '=SUMIF(H$3:{t1}$3,$E$3,H32:{t1}32)'.format(t1=last_col_alpha)
            ws['E33'].value = '=SUMIF(H$3:{t1}$3,$E$3,H33:{t1}33)'.format(t1=last_col_alpha)
            ws['E34'].value = '=SUMIF(H$3:{t1}$3,$E$3,H34:{t1}34)'.format(t1=last_col_alpha)
            ws['E35'].value = '=SUMIF(H$3:{t1}$3,$E$3,H35:{t1}35)'.format(t1=last_col_alpha)
            ws['E36'].value = '=SUMIF(H$3:{t1}$3,$E$3,H36:{t1}36)'.format(t1=last_col_alpha)
            ws['E37'].value = '=SUMIF(H$3:{t1}$3,$E$3,H37:{t1}37)'.format(t1=last_col_alpha)
            ws['E38'].value = '=SUMIF(H$3:{t1}$3,$E$3,H38:{t1}38)'.format(t1=last_col_alpha)
            ws['E39'].value = '=SUMIF(H$3:{t1}$3,$E$3,H39:{t1}39)'.format(t1=last_col_alpha)
            ws['E40'].value = '=SUMIF(H$3:{t1}$3,$E$3,H40:{t1}40)'.format(t1=last_col_alpha)
            ws['E41'].value = '=SUMIF(H$3:{t1}$3,$E$3,H41:{t1}41)'.format(t1=last_col_alpha)
            ws['E42'].value = '=SUMIF(H$3:{t1}$3,$E$3,H42:{t1}42)'.format(t1=last_col_alpha)

            ws['G26'].value = '=SUMIF(H$3:{t1}$3,$G$3,H26:{t1}26)'.format(t1=last_col_alpha)
            ws['G27'].value = '=SUMIF(H$3:{t1}$3,$G$3,H27:{t1}27)'.format(t1=last_col_alpha)
            ws['G28'].value = '=SUMIF(H$3:{t1}$3,$G$3,H28:{t1}28)'.format(t1=last_col_alpha)
            ws['G29'].value = '=SUMIF(H$3:{t1}$3,$G$3,H29:{t1}29)'.format(t1=last_col_alpha)
            ws['G30'].value = '=SUMIF(H$3:{t1}$3,$G$3,H30:{t1}30)'.format(t1=last_col_alpha)
            ws['G31'].value = '=SUMIF(H$3:{t1}$3,$G$3,H31:{t1}31)'.format(t1=last_col_alpha)
            ws['G32'].value = '=SUMIF(H$3:{t1}$3,$G$3,H32:{t1}32)'.format(t1=last_col_alpha)
            ws['G33'].value = '=SUMIF(H$3:{t1}$3,$G$3,H33:{t1}33)'.format(t1=last_col_alpha)
            ws['G34'].value = '=SUMIF(H$3:{t1}$3,$G$3,H34:{t1}34)'.format(t1=last_col_alpha)
            ws['G35'].value = '=SUMIF(H$3:{t1}$3,$G$3,H35:{t1}35)'.format(t1=last_col_alpha)
            ws['G36'].value = '=SUMIF(H$3:{t1}$3,$G$3,H36:{t1}36)'.format(t1=last_col_alpha)
            ws['G37'].value = '=SUMIF(H$3:{t1}$3,$G$3,H37:{t1}37)'.format(t1=last_col_alpha)
            ws['G38'].value = '=SUMIF(H$3:{t1}$3,$G$3,H38:{t1}38)'.format(t1=last_col_alpha)
            ws['G39'].value = '=SUMIF(H$3:{t1}$3,$G$3,H39:{t1}39)'.format(t1=last_col_alpha)
            ws['G40'].value = '=SUMIF(H$3:{t1}$3,$G$3,H40:{t1}40)'.format(t1=last_col_alpha)
            ws['G41'].value = '=SUMIF(H$3:{t1}$3,$G$3,H41:{t1}41)'.format(t1=last_col_alpha)
            ws['G42'].value = '=SUMIF(H$3:{t1}$3,$G$3,H42:{t1}42)'.format(t1=last_col_alpha)

            ws['E45'].value = '=SUMIF(H$3:{t1}$3,$E$3,H45:{t1}45)'.format(t1=last_col_alpha)
            ws['G45'].value = '=SUMIF(H$3:{t1}$3,$G$3,H45:{t1}45)'.format(t1=last_col_alpha)

            ws['E47'].value = '=SUMIF(H$3:{t1}$3,$E$3,H47:{t1}47)'.format(t1=last_col_alpha)

            ws['E49'].value = '=SUMIF(H$3:{t1}$3,$E$3,H49:{t1}49)'.format(t1=last_col_alpha)

            ws['E51'].value = '=SUMIF(H$3:{t1}$3,$E$3,H51:{t1}51)'.format(t1=last_col_alpha)
            ws['E52'].value = '=SUMIF(H$3:{t1}$3,$E$3,H52:{t1}52)'.format(t1=last_col_alpha)
            ws['E53'].value = '=SUMIF(H$3:{t1}$3,$E$3,H53:{t1}53)'.format(t1=last_col_alpha)
            ws['E54'].value = '=SUMIF(H$3:{t1}$3,$E$3,H54:{t1}54)'.format(t1=last_col_alpha)
            ws['E55'].value = '=SUMIF(H$3:{t1}$3,$E$3,H55:{t1}55)'.format(t1=last_col_alpha)

            ws['G51'].value = '=SUMIF(H$3:{t1}$3,$G$3,H51:{t1}51)'.format(t1=last_col_alpha)
            ws['G52'].value = '=SUMIF(H$3:{t1}$3,$G$3,H52:{t1}52)'.format(t1=last_col_alpha)
            ws['G53'].value = '=SUMIF(H$3:{t1}$3,$G$3,H53:{t1}53)'.format(t1=last_col_alpha)
            ws['G54'].value = '=SUMIF(H$3:{t1}$3,$G$3,H54:{t1}54)'.format(t1=last_col_alpha)
            ws['G55'].value = '=SUMIF(H$3:{t1}$3,$G$3,H55:{t1}55)'.format(t1=last_col_alpha)


            ws.print_area = 'A1:S56'
            # ws.print_area = 'A1:{}56'.format(get_column_letter(last_col))
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_state = 'hidden'

            # グループシートをアクティブにする
            wb.active = wb['グループ']

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, 'sanwa55')

            logger.debug("Excel file generated successfully")

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
