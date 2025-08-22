import datetime
from typing import Dict, Any
import io
import logging
from copy import copy
from fastapi.responses import StreamingResponse
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer
from app.utils.service_utils import ClsToukeiSyukei, ModKubuns
from app.core.config import settings
from app.core.exceptions import ServiceError
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill, Font

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class teamtsumiagecheckhyoService(BaseService):
    """チーム別積上チェック表サービス"""

    def display(self, request, session) -> Dict[str, Any]:

        # 支払予定表の印刷処理
        logger.info("【START】チーム別積上チェック表処理")

        if request.params['@i物件集計'] == "物件集計":
            storedname = "usp_TK2601チーム別積上チェック表_物件集計v2"
        else:
            storedname = "usp_TK2600チーム別積上チェック表_個別集計v2"

        params = {
            "@iS集計日付": request.params['@iS集計日付'],
            "@iE集計日付": request.params['@iE集計日付'],
            "@i業種区分": request.params['@i業種区分']
        }

        outputparams={"@RetDcnt":"INT", "@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams=outputparams))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        wb = create_excel_object('Temp積上チェック表.xlsx')

        # TemplateWel請求明細は リースごとにシート　物件区分ごとに表
        with get_excel_buffer() as buffer:

            ws = wb.active
            temp_ws = wb['Template']

            # シートコピー用前回処理納入先CD
            cp_last_syukei_cd = None

            # 罫線情報
            side1 = Side(style='double', color='000000')
            side2 = Side(style='thin', color='000000')
            side3 = Side(style='medium', color='000000')
            border = Border(top=side1, bottom=side2, right=side2, left=side2)
            first_border = Border(top=side1, bottom=side2, right=side2, left=side3)
            last_border = Border(top=side1, bottom=side2, right=side3, left=side2)

            # 初期列
            start_row = 4
            # Templateシートをコピー
            for result in storedresults['results'][0]:
                if cp_last_syukei_cd != result['統計集計先CD']:
                    ws = wb.copy_worksheet(temp_ws)

                    clsresult = ClsToukeiSyukei.GetbyID(session, result['統計集計先CD'])

                    ws['A1'] = result['統計集計先CD']

                    if clsresult:
                        ws['G1'] = clsresult[0]['統計集計先名']
                        ws.title = clsresult[0]['統計集計先名']
                    else:
                        ws['G1'] = "その他"
                        ws.title = "その他"
                    # 前回処理ウエルシアリース区分の更新
                    cp_last_syukei_cd = result['統計集計先CD']

                else:
                    # 前回処理ウエルシアリース区分の更新
                    cp_last_syukei_cd = result['統計集計先CD']

            # Templateシートの削除
            wb.remove(wb['Template'])

            for ws in wb.worksheets:

                # 各シートの処理開始時にrow_numを初期化
                row_num = start_row

                # シートの最初にする処理
                firstexec = False

                for row_data in storedresults['results'][0]:
                    # シートが納入先で分かれているときの処理
                    if ws['A1'].value == row_data['統計集計先CD']:
                        if firstexec == False:
                            last_tokuisaki_cd = row_data['納入得意先CD']
                            firstexec = True
                        ws.cell(row=row_num, column=1, value=row_data['見積番号'])._style = copy(ws.cell(row=start_row, column=1)._style)
                        if row_data['作業区分CD'] == 99999:
                            ws.cell(row=row_num, column=2, value="合計")._style = copy(ws.cell(row=start_row, column=2)._style)
                        elif row_data['作業区分CD'] == 0:
                            ws.cell(row=row_num, column=2, value="営業")._style = copy(ws.cell(row=start_row, column=2)._style)
                        else:
                            ws.cell(row=row_num, column=2, value=ModKubuns.GetSagyouKubunmei(row_data['作業区分CD']))._style = copy(ws.cell(row=start_row, column=2)._style)
                        ws.cell(row=row_num, column=3, value=row_data['見積件名'])._style = copy(ws.cell(row=start_row, column=3)._style)
                        ws.cell(row=row_num, column=4, value=row_data['得意先CD'])._style = copy(ws.cell(row=start_row, column=4)._style)
                        ws.cell(row=row_num, column=5, value=row_data['日付'])._style = copy(ws.cell(row=start_row, column=5)._style)
                        ws.cell(row=row_num, column=6, value=ModKubuns.GetBukkenSyubetumei(row_data['物件種別']))._style = copy(ws.cell(row=start_row, column=6)._style)
                        ws.cell(row=row_num, column=7, value=row_data['税抜金額'])._style = copy(ws.cell(row=start_row, column=7)._style)
                        ws.cell(row=row_num, column=8, value=row_data['仕入金額'])._style = copy(ws.cell(row=start_row, column=8)._style)
                        # 原価率
                        ws.cell(row=row_num, column=9, value='=IF(ISERROR(H{row_num}/G{row_num}),"",H{row_num}/G{row_num})'.format(row_num=row_num))._style = copy(ws.cell(row=start_row, column=9)._style)
                        ws.cell(row=row_num, column=10, value=row_data['請求日付'])._style = copy(ws.cell(row=start_row, column=10)._style)
                        ws.cell(row=row_num, column=11, value=row_data['請求予定日'])._style = copy(ws.cell(row=start_row, column=11)._style)

                        # 高さ調整
                        ws.row_dimensions[row_num].height = ws.row_dimensions[start_row].height

                        # 得意先CDが前回と違う場合は二重線を設定
                        if row_data['納入得意先CD'] + "" != last_tokuisaki_cd:
                            for column_num in range(1,12):
                                if column_num == 1:
                                    ws.cell(row=row_num, column=column_num).border = first_border
                                elif column_num == 11:
                                    ws.cell(row=row_num, column=column_num).border = last_border
                                else:
                                    ws.cell(row=row_num, column=column_num).border = border

                        last_tokuisaki_cd = row_data['納入得意先CD'] + ""

                        row_num += 1


                # 日付
                ws['H1'] = "作成日： {}".format("{0:%Y/%m/%d}".format(datetime.date.today()))
                # 金額合計
                ws['G3'] = '=SUMIF(B4:B{row_num},"<>合計",G4:G{row_num})'.format(row_num=row_num -1)
                # 原価額合計
                ws['H3'] = '=SUMIF(B4:B{row_num},"<>合計",H4:H{row_num})'.format(row_num=row_num -1)
                # 統計集計先CD削除
                ws['A1'] = ""

                # 改ページプレビューの設定
                ws.sheet_view.view = 'pageBreakPreview'
                ws.sheet_view.zoomScaleSheetLayoutView= 100

                # ウィンドウ枠の固定
                ws.freeze_panes = 'A3'

                # 条件式書式
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                rule1 = FormulaRule(formula=['=#REF! = "原価"'], fill=yellow_fill)
                rule2 = CellIsRule(operator="greaterThanOrEqual", formula=["0.83"], font = Font(color='FF0000'))
                rule3 = CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], font = Font(color='FF0000'))
                ws.conditional_formatting.add(f"H4:H{row_num -1}", rule1)
                ws.conditional_formatting.add(f"I4:I{row_num -1}", rule2)
                ws.conditional_formatting.add("I3", rule3)

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, 'sanwa55')

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )


    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass