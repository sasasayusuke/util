from datetime import datetime
from time import strftime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font, PatternFill, Protection
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class RetractedListService(BaseService):
    """消込済一覧表"""

    def display(self, request, session) -> Dict[str, Any]:

        # 消込済一覧表の印刷処理
        logger.info("【START】消込済一覧表")

        storedname="usp_ND0800消込済一覧表抽出"
        params = {
            "@iDateFrom":request.params["@iDateFrom"],
            "@iDateTo":request.params["@iDateTo"],
            "@is得意先CD":request.params["@is得意先CD"],
            "@ie得意先CD":request.params["@ie得意先CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_消込済一覧表.xlsx")
            ws = wb["Template"]
            ts = wb["Copy"]

            requestDateFrom = request.params["@iDateFrom"]
            requestDateTo =  request.params["@iDateTo"]
            custmerCodeFrom = request.params["@is得意先CD"]
            custmerCodeTo = request.params["@ie得意先CD"]

            # ヘッダー設定
            ws["F1"] = datetime.strptime(requestDateFrom, "%Y/%m/%d")
            ws["M1"] = datetime.strptime(requestDateTo, "%Y/%m/%d")
            ws["F2"] = custmerCodeFrom if custmerCodeFrom is not None else "最初"
            ws["K2"] = custmerCodeTo if custmerCodeTo is not None else "最後"
            ws["AP1"] = datetime.now()

            # 初期化
            start_row = 3
            current_nyukin_no = None
            current_row = start_row

            # データのループ
            for result in storedresults['results'][0]:
                nyukin_no = result['入金番号']

                # 入金番号ごとに表を分割
                if nyukin_no != current_nyukin_no:
                    current_nyukin_no = nyukin_no

                    # 初回行でない場合は1行開ける
                    if (current_row != start_row):
                        ws.row_dimensions[current_row].height = 28.5
                        current_row += 1

                    # サブヘッダーをコピー
                    range_copy_cell_by_address(ts,ws,'A1','AV3','A{}'.format(current_row),True)

                    # サブヘッダーを設定
                    ws.cell(row=current_row + 1, column=1, value=result['入金日付'])
                    ws.cell(row=current_row + 1, column=7, value=result['入金番号'])
                    ws.cell(row=current_row + 1, column=11, value=result['得意先CD'])
                    ws.cell(row=current_row + 1, column=15, value=result['得意先名1'])
                    ws.cell(row=current_row + 1, column=26, value=result['得意先名2'])
                    ws.cell(row=current_row + 1, column=39, value=result['入金金額'])
                    ws.cell(row=current_row + 1, column=44, value=result['入金済金額'])

                    # コピーした行数分カウントアップ
                    current_row = current_row + 3

                # 詳細行をコピー
                if result['税込金額'] > result['売上消込合計']:
                    # 税込金額が赤字の書式をコピー
                    range_copy_cell_by_address(ts,ws,'A5','AV5','A{}'.format(current_row),True)
                elif result['税込金額'] < result['売上消込合計']:
                    # 売上消込合計が赤字の書式をコピー
                    range_copy_cell_by_address(ts,ws,'A6','AV6','A{}'.format(current_row),True)
                else:
                    range_copy_cell_by_address(ts,ws,'A4','AV4','A{}'.format(current_row),True)

                # データ挿入
                ws.cell(row=current_row, column=3, value=result['売上日付'])
                ws.cell(row=current_row, column=9, value=result['請求日付'])
                ws.cell(row=current_row, column=14, value=result['見積番号'])
                ws.cell(row=current_row, column=20, value=result['税込金額'])
                ws.cell(row=current_row, column=25, value=result['売上消込合計'])

                current_row += 1

            # シート名の変更
            ws.title = '消込済一覧表'

            # コピー用シートを削除
            wb.remove(ts)

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="AV{}".format(current_row)
                )
                res_obj = converter.generate_pdf()

            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
