from datetime import datetime
from time import strftime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font, PatternFill, Protection
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class DailySalesLedgerService(BaseService):
    """売上日計表"""

    def display(self, request, session) -> Dict[str, Any]:

        # 売上日計表の印刷処理
        logger.info("【START】売上日計表")

        dict_denpyou_kubun = {
            '1': "売上",
            '2': "返品",
            '3': "単価訂正"
        }

        storedname="usp_HD0900売上日計表"
        type = request.params["@条件"]
        params = {
            "@is売上日付":request.params["@iDateFrom"] if type == '1' else None,
            "@ie売上日付":request.params["@iDateTo"] if type == '1' else None,
            "@is請求日付":None if type == '1' else request.params["@iDateFrom"],
            "@ie請求日付":None if type == '1' else request.params["@iDateTo"],
            "@is得意先CD":request.params["@is得意先CD"],
            "@ie得意先CD":request.params["@ie得意先CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        # raise ServiceError('該当データなし')

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')
        
        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_売上日計表.xlsx")
            ws = wb["Template"]
            ts = wb["Copy"]

            requestDateFrom = request.params["@iDateFrom"]
            requestDateTo =  request.params["@iDateTo"]
            custmerCodeFrom = request.params["@is得意先CD"]
            custmerCodeTo = request.params["@ie得意先CD"]

            # ヘッダー設定
            ws["F1"] = custmerCodeFrom if custmerCodeFrom is not None else "最初"
            ws["K1"] = custmerCodeTo if custmerCodeTo is not None else "最後"
            ws["F2"] = datetime.strptime(requestDateFrom, "%Y/%m/%d")
            ws["M2"] = datetime.strptime(requestDateTo, "%Y/%m/%d")

            ws["BD1"] = datetime.now() 
            
            # 初期化
            start_row = 4
            current_row = start_row
            old_custmer_code = None
            old_nounyuu_code = None
            old_uriage_date = None
            old_seikyu_date = None
            total_amount = 0

            # データのループ
            for result in storedresults['results'][0]:

                if (result['区分'] == 1):
                    # 詳細行1,2をコピー
                    range_copy_cell_by_address(ts,ws,'A1','BM2','A{}'.format(current_row),False)

                    custmer_code = result['納入得意先CD']
                    nounyuu_code = result['納入先CD']
                    uriage_date = result['売上日付']
                    seikyu_date = result['請求日付']
                    nounyuu_name1 = result['納入先名1']
                    nounyuu_name2 = result['納入先名2']

                    # 請求日付、売上日付、得意先CD、納入先CD、納入先1、納入先1の表示制御
                    if (custmer_code == old_custmer_code):
                        if (nounyuu_code == old_nounyuu_code):
                            if (uriage_date == old_uriage_date):
                                if (seikyu_date == old_seikyu_date):
                                    # 何も表示しない
                                    pass
                                else:
                                    # 請求日付を表示
                                    ws.cell(row=current_row + 1, column=19, value=seikyu_date)
                            else:
                                # 請求日付、売上日付を表示
                                ws.cell(row=current_row + 1, column=15, value=uriage_date)
                                ws.cell(row=current_row + 1, column=19, value=seikyu_date)
                        else:
                            # 請求日付、売上日付、得意先CD、納入先CD、納入先1、納入先2を表示
                            ws.cell(row=current_row + 1, column=1, value=custmer_code)
                            ws.cell(row=current_row + 1, column=3, value=nounyuu_code)
                            ws.cell(row=current_row, column=5, value=nounyuu_name1 if nounyuu_name2 else None)
                            ws.cell(row=current_row + 1, column=5, value=nounyuu_name2 if nounyuu_name2 else nounyuu_name1)
                            ws.cell(row=current_row + 1, column=15, value=uriage_date)
                            ws.cell(row=current_row + 1, column=19, value=seikyu_date)
                    else:
                        # 請求日付、売上日付、得意先CD、納入先CD、納入先1、納入先2を表示
                        ws.cell(row=current_row + 1, column=1, value=custmer_code)
                        ws.cell(row=current_row + 1, column=3, value=nounyuu_code)
                        ws.cell(row=current_row, column=5, value=nounyuu_name1 if nounyuu_name2 else None)
                        ws.cell(row=current_row + 1, column=5, value=nounyuu_name2 if nounyuu_name2 else nounyuu_name1)
                        ws.cell(row=current_row + 1, column=15, value=uriage_date)
                        ws.cell(row=current_row + 1, column=19, value=seikyu_date)

                    old_custmer_code = custmer_code
                    old_nounyuu_code = nounyuu_code
                    old_uriage_date = uriage_date
                    old_seikyu_date = seikyu_date

                    w = result['W']
                    d = result['D']
                    d1 = result['D1']
                    d2 = result['D2']
                    h = result['H']
                    h1 = result['H1']
                    h2 = result['H2']
                    unit_price = result['売上単価']
                    amount = result['売上金額']

                    # その他のデータ挿入                    
                    ws.cell(row=current_row, column=29, value=result['製品NO'])
                    ws.cell(row=current_row, column=32, value=result['仕様NO'])
                    if (w != 0):
                        ws.cell(row=current_row, column=36, value=result['W'])
                    # D
                    if (d != 0):
                        ws.cell(row=current_row, column=40, value=d)
                    else:
                        if (d1 == 0 and d2 == 0):
                            ws.cell(row=current_row, column=40, value="")
                        else:
                            ws.cell(row=current_row, column=40, value=d1/d2)
                    # H
                    if (h != 0):
                        ws.cell(row=current_row, column=44, value=h)
                    else:
                        if (h1 == 0 and h2 == 0):
                            ws.cell(row=current_row, column=44, value="")
                        else:
                            ws.cell(row=current_row, column=44, value=h1/h2)

                    ws.cell(row=current_row + 1, column=24, value=f"{result['見積番号']}-{result['売上明細行番号']}")
                    ws.cell(row=current_row + 1, column=29, value=result['漢字名称'])
                    ws.cell(row=current_row + 1, column=47, value=result['売上数量'])
                    ws.cell(row=current_row + 1, column=51, value=result['単位名'])
                    if (unit_price != 0):
                        ws.cell(row=current_row + 1, column=53, value=unit_price)
                    if (amount != 0):
                        ws.cell(row=current_row + 1, column=58, value=amount)
                    ws.cell(row=current_row + 1, column=63, value=dict_denpyou_kubun.get(result['伝票区分'], ""))
                    
                    total_amount = total_amount + amount

                elif(result['区分'] == 2):
                    # 消費税行,合計行をコピー
                    range_copy_cell_by_address(ts,ws,'A3','BM4','A{}'.format(current_row),True)
                    ws.cell(row=current_row, column=58, value=result['売上金額'])
                    ws.cell(row=current_row + 1, column=58, value=total_amount)

                    total_amount = 0

                    # 1行改行
                    current_row = current_row + 1
                else:
                    pass
                
                current_row = current_row + 2

            # シート名の変更
            ws.title = '売上日計表'

            # コピー用シートを削除
            wb.remove(ts)

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="BM{}".format(current_row)
                )
                res_obj = converter.generate_pdf()

                logger.debug("Pdf file generated successfully")
            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

                logger.debug("Excel file generated successfully")

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
