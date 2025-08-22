from datetime import datetime
from time import strftime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font, PatternFill, Protection
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class ShiharainikkeihyoService(BaseService):
    """支払日計表"""

    def display(self, request, session) -> Dict[str, Any]:

        # 支払日計表の印刷処理
        logger.info("支払日計表")

        requestDateFrom = request.params["@is支払日付"]
        requestDateTo =  request.params["@ie支払日付"]
        custmerCodeFrom = request.params["@is仕入先CD"]
        custmerCodeTo = request.params["@ie仕入先CD"]

        storedname="usp_SD0500支払日計表_明細行抽出"
        params = {
            "@is支払日付":request.params["@is支払日付"],
            "@ie支払日付":request.params["@ie支払日付"],
            "@is仕入先CD":request.params["@is仕入先CD"],
            "@ie仕入先CD":request.params["@ie仕入先CD"],
        }

        # SQL実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        # raise ServiceError('該当データなし')

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')
        
        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_支払日計表.xlsx")
            ws = wb["Template"]
            ts = wb["Copy"]

            # ヘッダー設定
            ws["F2"] = datetime.strptime(requestDateFrom, "%Y/%m/%d")
            ws["M2"] = datetime.strptime(requestDateTo, "%Y/%m/%d")
            ws["F1"] = '[' + (custmerCodeFrom if custmerCodeFrom is not None else "最初") + ']'
            ws["K1"] = '[' + (custmerCodeTo if custmerCodeTo is not None else "最後") + ']'
            ws["BD1"] = datetime.now() 
            
            # 初期化
            start_row = 4
            current_row = start_row

            current_siharai_bangou = None
            total_amount = 0
            tantousya_name = None
            init_flag = True

            # SQL1データのループ
            for result in storedresults['results'][0]:
                siharai_bangou = result['支払番号']
                if current_row == start_row:
                   current_siharai_bangou = siharai_bangou

                if current_siharai_bangou != siharai_bangou:
                    # データ行2をコピー
                    range_copy_cell_by_address(ts,ws,'A2','BO2','A{}'.format(current_row),True)
                    # データ挿入
                    ws.cell(row=current_row, column=38, value=total_amount) 
                    ws.cell(row=current_row, column=47, value=tantousya_name) 

                    total_amount = 0
                    current_siharai_bangou = siharai_bangou
                    init_flag = True
                    current_row += 1
                
                total_amount = total_amount + result['支払金額']
                tantousya_name = result['担当者名']
                # データ行1をコピー
                range_copy_cell_by_address(ts,ws,'A1','BO1','A{}'.format(current_row),True)
                # データ挿入
                if (init_flag):
                    ws.cell(row=current_row, column=1, value=f"[{result['仕入先CD']}]") 
                    ws.cell(row=current_row, column=4, value=result['仕入先名1']) 
                    ws.cell(row=current_row, column=14, value=result['仕入先名2']) 
                    ws.cell(row=current_row, column=24, value=result['支払日付']) 
                    ws.cell(row=current_row, column=29, value=siharai_bangou) 
                    
                ws.cell(row=current_row, column=33, value=result['支払区分名']) 
                ws.cell(row=current_row, column=38, value=result['支払金額']) 

                if (result['支払種別'] == 2):
                    tagata_str_date = result['手形期日'].strftime("%Y/%m/%d")
                    ws.cell(row=current_row, column=44, value=f"手形期日:{tagata_str_date}") 
                    ws.cell(row=current_row, column=51, value=f"手形番号:{result['手形番号']}") 
                ws.cell(row=current_row, column=58, value=result['摘要名'])

                init_flag = False
                current_row += 1
            
            # 最後の支払番号の合計を出力
            # データ行2をコピー
            range_copy_cell_by_address(ts,ws,'A2','BO2','A{}'.format(current_row),True)
            # データ挿入
            ws.cell(row=current_row, column=38, value=total_amount) 
            ws.cell(row=current_row, column=47, value=tantousya_name) 
            
            current_row += 1

            # 合計行1をコピー
            range_copy_cell_by_address(ts,ws,'A3','BO3','A{}'.format(current_row),True)
            current_row += 1
            
            total_nikkei = 0
            total_ruikei = 0
            genkin_nikkei = 0
            genkin_ruikei = 0

            # 集計用SQL実行
            storedname="usp_SD0200支払日計表抽出"
            params = {
                "@iSDATE":request.params["@is支払日付"],
                "@iFDATE":request.params["@ie支払日付"],
                "@iSSIIRE":request.params["@is仕入先CD"],
                "@iFSIIRE":request.params["@ie仕入先CD"],
            }

            # STORED実行
            storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

            # SQL2データのループ
            for result in storedresults['results'][0]:
                siharai_kubun = result["支払区分名"]
                nikkei = result["支払日計"]
                ruikei = result["支払累計"]
                
                # 合計行2をコピー
                range_copy_cell_by_address(ts,ws,'A4','BO4','A{}'.format(current_row),True)
                # データ挿入
                ws.cell(row=current_row, column=5, value=siharai_kubun) 
                ws.cell(row=current_row, column=10, value=nikkei) 
                ws.cell(row=current_row, column=16, value=ruikei) 

                total_nikkei = total_nikkei + nikkei
                total_ruikei = total_ruikei + ruikei
                if (siharai_kubun == '現金'):
                    genkin_nikkei = nikkei
                    genkin_ruikei = ruikei
                
                current_row += 1

            # 合計行3をコピー
            range_copy_cell_by_address(ts,ws,'A5','BO5','A{}'.format(current_row),True)
            # 合計出力
            ws.cell(row=current_row, column=10, value=total_nikkei) 
            ws.cell(row=current_row, column=16, value=total_ruikei) 
            current_row += 1

            # 合計行4をコピー
            range_copy_cell_by_address(ts,ws,'A6','BO6','A{}'.format(current_row),True)
            # 現金率出力
            ws.cell(row=current_row, column=10, value=(genkin_nikkei / total_nikkei * 100)) 
            ws.cell(row=current_row, column=16, value=(genkin_ruikei / total_ruikei * 100)) 
            
            # シート名の変更
            ws.title = '支払日計表'

            # コピー用シートを削除
            wb.remove(ts)

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="BO{}".format(current_row)
                )
                res_obj = converter.generate_pdf()

                logger.debug("Pdf file generated successfully")
            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

                logger.debug("Excel file generated successfully")

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )
        
    def display_siwake(self, request, session) -> StreamingResponse:

        # 支払日計表の印刷処理
        logger.info("【START】支払日計表仕分処理")

        storedname="usp_SD0100支払仕訳出力"

        params = {
            "@isDATE":request.params["@is支払日付"],
            "@ieDATE":request.params["@ie支払日付"],
            "@i開始仕入先":request.params["@is仕入先CD"],
            "@i終了仕入先":request.params["@ie仕入先CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')
        
        csv_data = ''
        first_row_bool = True

        # データのループ
        for row_data in storedresults['results'][0]:
            if not first_row_bool:
                csv_data += '\r\n'
            else:
                first_row_bool = False
            if row_data['年'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['年']) + ','
            if row_data['月'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['月']) + ','
            if row_data['日'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['日']) + ','
            if row_data['伝票番号'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['伝票番号']) + ','
            if row_data['借方金額'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方金額']) + ','
            if row_data['借方勘定科目CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方勘定科目CD']) + ','
            if row_data['借方勘定科目名'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方勘定科目名']) + ','
            if row_data['借方消費税CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方消費税CD']) + ','
            if row_data['借方消費税率'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方消費税率']) + ','
            if row_data['借方業種'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方業種']) + ','
            if row_data['借方補助科目CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方補助科目CD']) + ','
            if row_data['借方補助科目名'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方補助科目名']) + ','
            if row_data['借方部門CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方部門CD']) + ','
            if row_data['借方部門名'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['借方部門名']) + ','
            if row_data['貸方勘定科目CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方勘定科目CD']) + ','
            if row_data['貸方勘定科目名'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方勘定科目名']) + ','
            if row_data['貸方消費税CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方消費税CD']) + ','
            if row_data['貸方消費税率'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方消費税率']) + ','
            if row_data['貸方業種'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方業種']) + ','
            if row_data['貸方補助科目CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方補助科目CD']) + ','
            if row_data['貸方補助科目名'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方補助科目名']) + ','
            if row_data['貸方部門CD'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方部門CD']) + ','
            if row_data['貸方部門名'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方部門名']) + ','
            if row_data['貸方金額'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['貸方金額']) + ','
            if row_data['摘要'] is None:
                csv_data += '"",'
            else:
                csv_data = csv_data + '"' + str(row_data['摘要']) + '",'
            if row_data['期日_年'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['期日_年']) + ','
            if row_data['期日_月'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['期日_月']) + ','
            if row_data['期日_日'] is None:
                csv_data += ','
            else:
                csv_data = csv_data + str(row_data['期日_日']) + ','
            if row_data['手形番号'] is None:
                csv_data += ''
            else:
                csv_data = csv_data + str(row_data['手形番号'])

        csv_data = csv_data.encode('cp932')
        # csv_data = csv_data.encode('shift-jis')

        return StreamingResponse(
            io.BytesIO(csv_data),
            media_type="text/csv; charset=Shift-JIS",
            headers={
                'Content-Disposition': 'attachment; filename="sample.csv"'
            }
        )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
