from datetime import datetime
from time import strftime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font, PatternFill, Protection
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class ShiirenikkeihyoService(BaseService):
    """仕入日計表"""

    def display(self, request, session) -> Dict[str, Any]:

        # 仕入日計表の印刷処理
        logger.info("【START】仕入日計表")

        dict_denpyou_kubun = {
            '1': "仕入",
            '2': "返品",
            '3': "単価訂正"
        }

        storedname="usp_SD0400仕入日計表"
        type = request.params["@条件"]
        params = {
            "@is仕入日付":request.params["@iDateFrom"] if type == '1' else None,
            "@ie仕入日付":request.params["@iDateTo"] if type == '1' else None,
            "@is支払日付":request.params["@iDateFrom"] if type == '2' else None,
            "@ie支払日付":request.params["@iDateTo"] if type == '2' else None,
            "@is仕入先CD":request.params["@is仕入先CD"],
            "@ie仕入先CD":request.params["@ie仕入先CD"],
        }

        # SQL実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        # raise ServiceError('該当データなし')

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')
        
        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_仕入日計表.xlsx")
            ws = wb["Template"]
            ts = wb["Copy"]

            requestDateFrom = request.params["@iDateFrom"]
            requestDateTo =  request.params["@iDateTo"]
            custmerCodeFrom = request.params["@is仕入先CD"]
            custmerCodeTo = request.params["@ie仕入先CD"]

            # ヘッダー設定
            ws["F1"] = '[' + (custmerCodeFrom if custmerCodeFrom is not None else "最初") + ']'
            ws["K1"] = '[' + (custmerCodeTo if custmerCodeTo is not None else "最後") + ']'
            ws["A2"] = "≪仕入日付≫" if type == 1 else "≪支払日付≫"
            ws["F2"] = datetime.strptime(requestDateFrom, "%Y/%m/%d")
            ws["M2"] = datetime.strptime(requestDateTo, "%Y/%m/%d")

            ws["BD1"] = datetime.now() 
            
            # 初期化
            start_row = 4
            current_row = start_row
            old_siiresaki_code = None
            old_siire_date = None
            old_siharai_date = None
            old_mitumori_name = None
            old_siire_bangou = None
            total_amount = 0
            siire_total_amount = 0

            # データのループ
            for result in storedresults['results'][0]:

                if (old_siire_bangou is not None) and (old_siire_bangou != result['仕入番号']):
                    range_copy_cell_by_address(ts,ws,'A5','BM6','A{}'.format(current_row),True)
                    ws.cell(row=current_row, column=18, value=old_mitumori_name)
                    ws.cell(row=current_row, column=55, value=total_amount)
                    total_amount = 0
                    current_row = current_row + 2

                if (old_siiresaki_code is not None) and (result['仕入先CD'] != old_siiresaki_code):
                    # 仕入先計を表示
                    range_copy_cell_by_address(ts,ws,'A7','BM8','A{}'.format(current_row),True)
                    ws.cell(row=current_row, column=55, value=siire_total_amount)
                    siire_total_amount = 0
                    # 1行改行
                    current_row = current_row + 2

                if (result['区分'] == 1):
                    # 詳細行1,2をコピー
                    range_copy_cell_by_address(ts,ws,'A1','BM2','A{}'.format(current_row),False)

                    siiresaki_code = result['仕入先CD']
                    siire_date = result['仕入日付']
                    siharai_date = result['支払日付']
                    siire_name1 = result['仕入先名1']
                    siire_name2 = result['仕入先名2']

                    # 請求日付、売上日付、得意先CD、納入先CD、納入先1、納入先1の表示制御
                    if (siiresaki_code == old_siiresaki_code):
                        if (siire_date == old_siire_date):
                            if (siharai_date == old_siharai_date):
                                # 何も表示しない
                                pass
                            else:
                                # 支払日付を表示
                                ws.cell(row=current_row + 1, column=18, value=siharai_date)
                        else:
                            # 請求日付、売上日付を表示
                            ws.cell(row=current_row + 1, column=14, value=siire_date)
                            ws.cell(row=current_row + 1, column=18, value=siharai_date)
                    else:
                        # 請求日付、売上日付、得意先CD、納入先CD、納入先1、納入先2を表示
                        ws.cell(row=current_row + 1, column=1, value=siiresaki_code)
                        ws.cell(row=current_row, column=4, value=siire_name1 if siire_name2 else None)
                        ws.cell(row=current_row + 1, column=4, value=siire_name2 if siire_name2 else siire_name1)
                        ws.cell(row=current_row + 1, column=14, value=siire_date)
                        ws.cell(row=current_row + 1, column=18, value=siharai_date)

                    old_siiresaki_code = siiresaki_code
                    old_siire_date = siire_date
                    old_siharai_date = siharai_date

                    w = result['W']
                    d = result['D']
                    d1 = result['D1']
                    d2 = result['D2']
                    h = result['H']
                    h1 = result['H1']
                    h2 = result['H2']
                    unit_price = result['仕入単価']
                    amount = result['仕入金額']

                    # その他のデータ挿入                    
                    ws.cell(row=current_row+1, column=23, value=str(result['見積番号'])+'-'+str(result['仕入明細行番号']).zfill(3))
                    ws.cell(row=current_row, column=27, value=result['製品NO'])
                    ws.cell(row=current_row+1, column=27, value=result['漢字名称'])
                    ws.cell(row=current_row, column=30, value=result['仕様NO'])
                    if (w != 0):
                        ws.cell(row=current_row, column=36, value=result['W'])
                    # D
                    if (d != 0):
                        ws.cell(row=current_row, column=39, value=d)
                    else:
                        if (d1 == 0 and d2 == 0):
                            ws.cell(row=current_row, column=39, value="")
                        else:
                            ws.cell(row=current_row, column=39, value=d1/d2)
                    # H
                    if (h != 0):
                        ws.cell(row=current_row, column=42, value=h)
                    else:
                        if (h1 == 0 and h2 == 0):
                            ws.cell(row=current_row, column=42, value="")
                        else:
                            ws.cell(row=current_row, column=42, value=h1/h2)

                    ws.cell(row=current_row + 1, column=45, value=result['仕入数量'])
                    ws.cell(row=current_row + 1, column=48, value=result['単位名'])
                    if (unit_price != 0):
                        ws.cell(row=current_row + 1, column=51, value=unit_price)
                    ws.cell(row=current_row + 1, column=55, value=amount)
                    ws.cell(row=current_row + 1, column=59, value=dict_denpyou_kubun.get(result['伝票区分'], ""))
                    ws.cell(row=current_row + 1, column=64, value=result['支払締日'])

                    total_amount = total_amount + amount
                    siire_total_amount = siire_total_amount + amount
                    current_row = current_row + 2
                    old_mitumori_name = result['見積件名']
                    old_siire_bangou = result['仕入番号']

                elif(result['区分'] == 2):
                    # 消費税行をコピー
                    range_copy_cell_by_address(ts,ws,'A3','BM4','A{}'.format(current_row),True)
                    ws.cell(row=current_row+1, column=55, value=result['仕入金額'])
                    total_amount = total_amount + result['仕入金額']
                    siire_total_amount = siire_total_amount + result['仕入金額']

                    # 改行
                    current_row = current_row + 2
                    old_mitumori_name = result['見積件名']
                    old_siire_bangou = result['仕入番号']
                else:
                    pass

            # 最終行に合計金額と仕分先計を表示
            range_copy_cell_by_address(ts,ws,'A5','BM6','A{}'.format(current_row),True)
            ws.cell(row=current_row, column=18, value=old_mitumori_name)
            ws.cell(row=current_row, column=55, value=total_amount)
            total_amount = 0
            current_row = current_row + 2

            # 仕入先計を表示
            range_copy_cell_by_address(ts,ws,'A7','BM8','A{}'.format(current_row),True)
            ws.cell(row=current_row, column=55, value=siire_total_amount)
            siire_total_amount = 0
            # 1行改行
            current_row = current_row + 2
                
            # シート名の変更
            ws.title = '仕入日計表'

            # コピー用シートを削除
            wb.remove(ts)

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="BM{}".format(current_row)
                )
                res_obj = converter.generate_pdf()

                logger.debug("Pdf file generated successfully")
            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

                logger.debug("Excel file generated successfully")

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
