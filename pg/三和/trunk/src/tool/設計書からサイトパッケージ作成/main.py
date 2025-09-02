import util
import config
import openpyxl
import os
import re
import requests
import copy
import traceback
import datetime
import json5
import logging
import database

class CustomError(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)

def setup_logging():
    log_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'log')
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)

    log_file = os.path.join(log_folder, f'application_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
        ]
    )


def save_setting(app):
    setting_data = {
        "API_KEY": app.api_key_value,
        "SERVER_ADDRESS": app.server_value,
        "SITE_ID": app.site_value,
        "INPUT_DIRECTORY": app.input_dir_value,
        "PACKAGE_FILE": app.package_file_value,
        # "DATA_OPTION": app.radio_api_value,
        "ENTRY_OPTION": app.radio_action_value,
    }

    try:
        util.save_json(setting_data, 'setting.json')
        app.show_message("設定保存", "設定が正常に保存されました。")
    except Exception as e:
        app.show_message("エラー", f"設定の保存中にエラーが発生しました: {str(e)}")

def main(app):
    setup_logging()
    logging.info("Starting application")

    try:
        server_address = app.server_value
        site_id = app.site_value
        api_key = app.api_key_value
        api_ver = 1.1
        input_dir = app.input_dir_value
        package_file = app.package_file_value
        checked_sheets = app.checkbox_values
        is_create = app.is_create
        is_copy = app.is_copy

        logging.info(f"Server Address: {server_address}")
        logging.info(f"Site ID: {site_id}")
        logging.info(f"Input Directory: {input_dir}")
        logging.info(f"PACKAGE FILE: {package_file}")

        # チェックボックスの選択状態を確認
        if not checked_sheets[config.SHEETS["詳細_編集要素"]["key"]] and (
            checked_sheets[config.SHEETS["一覧_画面レイアウト"]["key"]]
            or checked_sheets[config.SHEETS["詳細_画面レイアウト"]["key"]]
        ):
            raise CustomError("詳細_編集要素 が選択されていない場合、一覧_画面レイアウト または 詳細_画面レイアウト は選択できません。")
        elif not any(checked_sheets.values()):
            raise CustomError("少なくとも1つの項目を選択してください。")

        # 読み込み設計書ファイル名が '~$' で始まるものを除外し、正しい形式のものだけを処理
        params = {}
        site_info = {}
        pattern = r'([^_]+)_(.*)_画面設計書\.xlsx'
        for file_name in os.listdir(input_dir):
            if file_name.startswith('~$'):
                # バックアップファイルは読み込みスキップ
                continue
            matched = re.match(pattern, file_name)
            if not matched:
                # パターンにマッチしなかった場合
                continue

            # パターンにマッチした場合
            site_update_id = matched.group(1)
            site_name = matched.group(2)

            site_info[site_name] = {
                "id": site_update_id,
                "file_name": file_name,
            }

        if is_copy:
            package_json = {
                "TargetSiteId": site_id,
                "SiteTitle": "Test",
                "SelectedSites": util.read_json(package_file)["HeaderInfo"]["Convertors"]
            }

            app.log_output(f'{server_address}/items/{site_id}/index へ {package_file}を移植します。')

            # APIキーとバージョンを追加
            package_json["ApiVersion"] = api_ver
            package_json["ApiKey"] = api_key

            # サイトコピーPOSTリクエスト
            response = requests.post(f"{server_address}/api/items/{site_id}/copysitepackage", json=package_json)

            # APIキーとバージョンを削除
            del package_json["ApiVersion"]
            del package_json["ApiKey"]

            app.log_output(f'Status Code: {response.status_code}')
            app.log_output(f'Response: {response.json()}')

            site_json = copy.deepcopy(config.post_json)

            # 移植データの取得
            transplant_data = json5.loads(response.json()["Response"]["Data"])
            for transplant_site in transplant_data:
                newId = transplant_site["NewSiteId"]
                title = transplant_site["Title"]
                refType = transplant_site["ReferenceType"]
                if refType != "Results":
                    continue
                if title in site_info:
                    site_info[title]["id"] = newId
                    site_info[title]["update_url"] = f"{server_address}/api/items/{newId}/updatesite"
                    site_info[title]["site_json"] = site_json


        if is_create:
            for site_name, site in site_info.items():
                site_json = copy.deepcopy(config.post_json)

                # APIキーとバージョンを追加
                site_json["ApiVersion"] = api_ver
                site_json["ApiKey"] = api_key

                # JSONデータを用いてPOSTリクエストを実行
                response = requests.post(f"{server_address}/api/items/{site_id}/createsite", json=site_json)

                # APIキーとバージョンを削除
                del site_json["ApiVersion"]
                del site_json["ApiKey"]


                app.log_output(f'Status Code: {response.status_code}')
                app.log_output(f'Response: {response.json()}')

                site_update_id = response.json()["Id"]
                app.log_output(f'{server_address}/items/{site_update_id}/index を作成しました。')

                site_json["Title"] = site_name

                site["id"] = site_update_id
                site["update_url"] = f"{server_address}/api/items/{site_update_id}/updatesite"
                site["site_json"] = site_json



        for site_name, site in site_info.items():
            app.log_output(f"# {site['file_name']}")
        app.log_output(f"## 読み取り件数: {len(site_info)}")

        for site_name, site in site_info.items():
            if "site_json" not in site:
                app.log_output(f"#### {site_name} は名称が照合しなかったためスキップされました", "warning")
                site['skipped'] = True
                continue
            else:
                site['skipped'] = False

            file_path = os.path.join(input_dir, site["file_name"])
            # ワークブックを読み込みを開始
            app.log_output(f"### {file_path} の Book読み込みを開始")
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            item_dict = {}
            for name in ["詳細_編集要素", "スクリプト要素"]:
                # 読み込み対象のシートがチェックされているか確認
                if checked_sheets[config.SHEETS[name]["key"]]:
                    app.log_output(f"#### {name} の Sheet読み込みを開始")
                else:
                    app.log_output(f"#### {name} は読み込みスキップされました")
                    continue

                sht = workbook[name]

                # 型情報データを取得する
                type_cells = [cell for cell in sht[config.EDIT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
                type_cells.append(sht[util.get_address(config.EDIT_ROW_INDEX_TYPE, sht.max_column + 1)])
                for cur_cell, next_cell in util.pairwise(type_cells):
                    app.log_output(f"{cur_cell} 読み込みを開始", "debug")

                    start_row = config.EDIT_ROW_INDEX_GRID
                    end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column)) + 1
                    start_col = cur_cell.column
                    end_col = next_cell.column - 1

                    for i in range(start_row, end_row):
                        edit_obj = {}
                        for j in range(start_col, end_col):

                            app.log_output(f"row: {i} col: {j} 読み込みを開始", "debug")
                            item_cell = sht.cell(row=config.EDIT_ROW_INDEX_ITEM, column=j)
                            target_cell = sht.cell(row=i, column=j)

                            # 値チェック
                            if util.is_empty(target_cell.value):
                                continue
                            elif item_cell.value not in config.PARAMETERS:
                                app.log_output(f"{util.get_address(item_cell.row, item_cell.column)}: PARAMETERS に登録されていないのでスキップします。: {item_cell.value}", "warning")
                                continue
                            # 型チェック
                            if config.PARAMETERS[item_cell.value]["type"] == "float":
                                if util.is_numeric(target_cell.value):
                                    edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value
                                else:
                                    app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がfloatではないためスキップします。: {target_cell.value}", "warning")
                                    continue
                            elif config.PARAMETERS[item_cell.value]["type"] == "bool":
                                if target_cell.value == config.MARK_OK or target_cell.value == "":
                                    edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value == config.MARK_OK
                                else:
                                    app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がboolではないためスキップします。: {target_cell.value}", "warning")
                                    continue
                            elif config.PARAMETERS[item_cell.value]["type"] == "array":
                                if target_cell.value in config.PARAMETERS[item_cell.value]["values"]:
                                    edit_obj[config.PARAMETERS[item_cell.value]["key"]] = config.PARAMETERS[item_cell.value]["values"][target_cell.value]
                                else:
                                    app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の値が不正のためスキップします。: {target_cell.value}", "warning")
                                    continue
                            elif config.PARAMETERS[item_cell.value]["type"] == "invalid":
                                continue
                            else:
                                item_value = target_cell.value
                                if item_cell.value == "項目名":
                                    if util.is_numeric(target_cell.value):
                                        item_value = config.TYPES[cur_cell.value]["key"] + util.to_N_digits(target_cell.value, 3)
                                    elif len(target_cell.value) == 1 and target_cell.value.isalpha():
                                        item_value = config.TYPES[cur_cell.value]["key"] + target_cell.value.upper()
                                    else:
                                        app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 不正な値なのでこの行は読み込みスキップします。: {target_cell.value}", "warning")
                                        break

                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = item_value

                        # 詳細_編集要素 必須チェック
                        if name == "詳細_編集要素" and config.PARAMETERS["項目名"]["key"] in edit_obj and config.PARAMETERS["表示名"]["key"] in edit_obj:
                            # 重複チェック
                            if config.PARAMETERS["項目名"]["key"] in item_dict.values():
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 項目名 が重複しているのでこの行は読み込みスキップします。: {config.PARAMETERS['項目名']['key']}", "warning")
                                continue
                            if config.PARAMETERS["表示名"]["key"] in item_dict.keys():
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 表示名 が重複しているのでこの行は読み込みスキップします。: {config.PARAMETERS['表示名']['key']}", "warning")
                                continue
                            item_dict[edit_obj[config.PARAMETERS["表示名"]["key"]]] = edit_obj[config.PARAMETERS["項目名"]["key"]]

                            site["site_json"]["SiteSettings"]["Columns"].append(edit_obj)

                        # スクリプト要素 必須チェック
                        elif name == "スクリプト要素" and config.PARAMETERS["タイトル"]["key"] in edit_obj:
                            site["site_json"]["SiteSettings"][config.TABS[cur_cell.value]["key"]].append(edit_obj)


                site["site_json"]["SiteSettings"]["EditorColumnHash"]["General"] = item_dict.values()

            for name in ["一覧_画面レイアウト", "詳細_画面レイアウト"]:
                # 読み込み対象のシートがチェックされているか確認
                if checked_sheets[config.SHEETS[name]["key"]]:
                    app.log_output(f"#### {name} の Sheet読み込みを開始")
                else:
                    app.log_output(f"#### {name} は読み込みスキップされました")
                    continue

                packed_item_dict = {**item_dict, **config.LAYOUTS[name]}

                sht = workbook[name]
                app.log_output(f"#### {name} の Sheet読み込みを開始")
                # 型情報データを取得する
                type_cells = [cell for cell in sht[config.LAYOUT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
                type_cells.append(sht[util.get_address(config.LAYOUT_ROW_INDEX_TYPE, sht.max_column + 1)])
                for cur_cell, next_cell in util.pairwise(type_cells):
                    app.log_output(f"{cur_cell} 読み込みを開始", "debug")
                    start_row = config.LAYOUT_ROW_INDEX_GRID
                    end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column))
                    start_col = cur_cell.column
                    end_col = next_cell.column - 1
                    list = []
                    for i in range(start_row, end_row):
                        for j in range(start_col + 1, end_col):
                            app.log_output(f"row: {i} col: {j} 読み込みを開始", "debug")
                            target_cell = sht.cell(row=i, column=j)
                            if util.is_empty(target_cell.value):
                                continue
                            if target_cell.value not in packed_item_dict.keys():
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 詳細_編集要素 に登録されていないのでスキップします。: {target_cell.value} ", "warning")
                                continue
                            item_name = packed_item_dict[target_cell.value]
                            if item_name in list:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: ２重に登録されているのでスキップします。: {target_cell.value} ", "warning")
                                continue
                            list.append(item_name)
                    if cur_cell.value in config.TABS:
                        if cur_cell.value == "編集要素":
                            site["site_json"]["SiteSettings"]["EditorColumnHash"]["General"] = list
                        elif cur_cell.value == "集計要素":
                            site["site_json"]["SiteSettings"]["Aggregations"] = list
                        else:
                            site["site_json"]["SiteSettings"][config.TABS[cur_cell.value]["key"]] = list

            # 画面設計書の読み込みを終了し、取得データを変数ファイル用に更新
            for key, value in item_dict.items():
                if site_name not in params:
                    params[site_name] = {"ID": site_info[site_name]["id"]}
                params[site_name][value] = key
            for key, value in item_dict.items():
                params[site_name][key] = value


        # ループ終了後に削除を実行
        for site_name in [site_name for site_name, site in site_info.items() if site.get('skipped', False)]:
            del site_info[site_name]

        # 変数ファイルを保存
        app.log_output(f"#### 読込したJSONデータを変数ファイルを{file_path}へ保存")
        file_path = util.save_js_object(params, directory=r"C:\Users\NT-210174\Desktop\work\svn\YJY_三和商研受発注システム\trunk\src\frontend\js")
        # サイトパッケージJSONを保存
        directory = os.path.join(r"C:\Users\NT-210174\Desktop\work\svn\YJY_三和商研受発注システム\trunk\src\tool\設計書からサイトパッケージ作成\サイトパッケージ", datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        for site_name, site in site_info.items():
            app.log_output(f"#### {site_name} の 読込したJSONデータを比較用に保存")

            file_name = util.save_json(site["site_json"], site_name, directory)
            app.log_output(f"{file_name} を作成しました")


        # JSONデータを用いてPOSTリクエストを実行
        db_factory = database.DatabaseManagerFactory.get_instance()
        session = db_factory.get_pleasanter_db()

        # 読込したデータを上書き更新
        for site_name, site in site_info.items():
            app.log_output(f"#### {site_name} の 読込したデータへ上書き更新を開始")

            # # 選択肢一覧のLookup対応更新
            # key_name = config.PARAMETERS["選択肢一覧"]["key"]
            # for column in site["site_json"]["SiteSettings"]["Columns"]:
            #     if key_name in column:
            #         # 選択肢一覧が存在するもののみ抽出
            #         try:
            #             column[key_name] = util.process_site_info(column[key_name], site_info)
            #             update_flag = True
            #         except (json5.JSONDecodeError, ValueError):
            #             app.log_output(f"Invalid JSON: {column[key_name]}", "warning")


            # APIキーとバージョンを追加
            site["site_json"]["ApiVersion"] = api_ver
            site["site_json"]["ApiKey"] = api_key

            response = requests.post(site["update_url"], json=site["site_json"])

            # APIキーとバージョンを削除
            del site["site_json"]["ApiVersion"]
            del site["site_json"]["ApiKey"]

            app.log_output(f'Status Code: {response.status_code}')
            app.log_output(f'Response: {response.json()}')

            site_update_id = response.json()["Id"]

            app.log_output(f'{server_address}/items/{site_update_id}/index にアクセスしてください')

            # ビューの作成
            database.create_view(session, site_name, f"SELECT * FROM Results WHERE SiteId = {site_update_id}")

            app.log_output(f"ビュー '{site_name}' が正常に作成されました。")

        app.log_output(f'全ての処理が正常終了しました。')


    except CustomError as e:
        app.show_message("エラー", f"{str(e)}")
    except Exception:
        error_info = traceback.format_exc()
        logging.error(f"Detailed traceback information:\n{error_info}")
        app.show_message("エラー", f"処理の実行中にエラーが発生しました。")
        app.log_output(f"処理の実行中にエラーが発生しました。\n{error_info}", "error")


if __name__ == "__main__":
    import gui

    app = gui.Gui("設計書からサイトパッケージ作成",execute_callback=main, save_callback=save_setting)
    app.mainloop()