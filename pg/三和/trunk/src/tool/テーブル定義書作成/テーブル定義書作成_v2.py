# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
# ##使い方
# 1.pythonをインストール
# 2.必要ライブラリをpipでインストール（多分sysとopenpyxl）
# 3.このpythonファイル、テンプレートを同フォルダ内に格納する
# 4.39行目にjsonファイル名を記載
# 5.52行目にテンプレートファイル名を記載
# 6.実行してみる（エラーが出たら教えて）
# 7.各シートのO列が改行されていなかったら一回ダブルクリック or F2キーを押す
# 8.「成形マクロ.txt」を全コピーして、VBEの標準モジュールに張り付ける　→　実行する
# 
# 
# ##注意事項
# ・不備があるかもしれないから目視確認して
# ・フィルタは当てにならないから画面と見比べて
# ・B列が赤いセルは手作業で。
# ・1個も項目が有効化されていないテーブル・ダッシュボードはスキップされる
# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

import sys,os,json,openpyxl as excel
from openpyxl.styles import Font, PatternFill
import re


def choiceSelect(choice):
    if re.match(r"^([1-9]\d*|0)$",choice.replace("[[","").replace("]]","")) != None:
        choice = int(choice.replace("[[","").replace("]]",""))
        choice = list(filter(lambda e : e["SiteId"] == choice,convertors))
        choice = choice[0]["SiteTitle"]
    elif choice.replace('\n','') == '[[Users]]':
        choice = "ユーザー"
    elif choice.replace('\n','') == '[[Depts]]':
        choice = '組織'
    elif choice.replace('\n','') == '[[Groups]]':
        choice = 'グループ'
    return choice

# 現在のディレクトリからjsonファイルを取得
fileName = 'bk_2024_07_02 09_07_29.json'
filePath = os.getcwd() + "\\" + fileName
if not os.path.isfile(filePath):
    print('ファイルが存在しない')
    sys.exit()
else:
    print("処理を開始します。")

json_open = open(filePath,'r',encoding="utf-8_sig")
json_load = json.load(json_open)

# Excelの準備
# テンプレートをロード
wb = excel.load_workbook("テーブル定義書.xlsx")
template_ws = wb["Sheet1"]

convertors = json_load["HeaderInfo"]["Convertors"]
sites = json_load["Sites"]


for site in sites: 

    # フォルダの場合はスキップ
    if site["ReferenceType"] == "Sites" or site["ReferenceType"] == "Dashboards" or not 'EditorColumnHash' in site["SiteSettings"]:
        continue

    # シートをコピー
    ws = wb.copy_worksheet(template_ws)
    ws.title = site["Title"]
    ws.cell(1,2).value = site["SiteId"]
    ws.cell(1,4).value = site["Title"]

    data = site["SiteSettings"]
    columns = data["Columns"] if 'Columns' in data else [] # 各項目の情報
    editItem = data["EditorColumnHash"]["General"] if 'General' in data["EditorColumnHash"] else [] # エディタで有効化されている項目
    linkColumns = data["LinkColumns"] if 'LinkColumns' in data else [] # リンクタブで有効化されている項目
    filterColumns = data["FilterColumns"] if 'FilterColumns' in data else [] # フィルタで有効化されている項目
    gridColumns = data["GridColumns"] if 'GridColumns' in data else [] # 一覧で有効化されている項目
    TitleColumns = data["TitleColumns"] if 'TitleColumns' in data else [] # タイトル項目の情報
    
    i = 5
    font_red = Font(color='FF0000')

    # 項目ごとに繰り返す
    for item in editItem:
        # section　と　Linkを除外
        if item[:1] == "_":
            continue

        # 項目の詳細情報を取得
        column = list(filter(lambda x : x["ColumnName"] == item,columns))
        
        # タイトルだけは特別扱い
        if item == "Title":
            description = ''
            if len(TitleColumns) > 0:
                description = ','.join(TitleColumns)
            
            # columnsにTitle項目が存在しない場合
            if len(column) == 0:
                ws.cell(i,1).value = 'Title'
                ws.cell(i,2).value = 'タイトル'
                ws.cell(i,3).value = 'タイトル'
                ws.cell(i,4).value = '〇'
                ws.cell(i,5).value = '〇' if "Title" in linkColumns else ''
                ws.cell(i,6).value = '〇' if "Title" in filterColumns else ''
                ws.cell(i,13).value = '〇'
                ws.cell(i,18).value = description
                i += 1
                continue
            
            itemName = ''
            # 項目名
            if 'LabelText' in column[0]:
                itemName = column[0]['LabelText']
            else:
                itemName = 'タイトル'
            
            # 選択肢
            choice = column[0]["ChoicesText"] if "ChoicesText" in column[0] else ''
            choice = choiceSelect(choice)
            
            ws.cell(i,1).value = 'Title'  #採番
            ws.cell(i,2).value = itemName    #項目名
            ws.cell(i,3).value = 'タイトル'    #型
            ws.cell(i,4).value = '○'     #エディタ
            ws.cell(i,5).value = '〇' if "Title" in linkColumns else ''
            ws.cell(i,6).value = '〇' if "Title" in filterColumns else ''
            ws.cell(i,7).value = '〇' if "Title" in gridColumns else ''
            ws.cell(i,11).value = column[0]["DefaultInput"] if 'DefaultInput' in column[0] else '' #規定値
            ws.cell(i,13).value = "〇" if "ValidateRequired" in column[0] and column[0]["ValidateRequired"] else ""  #必須
            ws.cell(i,14).value = "〇" if "NoDuplication" in column[0] and column[0]["NoDuplication"] else ""  #重複禁止
            ws.cell(i,15).value = choice   #選択肢
            ws.cell(i,16).value = "〇" if "MultipleSelections" in column[0] and column[0]["MultipleSelections"] else "" #複数選択可
            ws.cell(i,17).value = "〇" if "EditorReadOnly" in column[0] and column[0]["EditorReadOnly"] else ''#読み取り専用
            ws.cell(i,18).value =  description # 備考
            i += 1
            continue

        # columnが見つからなかった場合は赤文字にして次の繰り返しへ
        # if len(column) == 0:
        #     ws.cell(i,1).value = item
        #     ws.cell(i,1).font = font_red
        #     i += 1
        #     continue

        # データ型
        dataType = ''
        match item[:3]:
            case 'Cla':
                dataType = '分類項目'
            case 'Num':
                dataType = '数値項目'
            case 'Dat':
                dataType = '日付項目'
            case 'Des':
                dataType = '説明項目'
            case 'Che':
                dataType = 'チェック項目'
            case 'Att':
                dataType = '添付ファイル項目'
            case 'Tit':
                dataType = 'タイトル'
            case 'Sta':
                dataType = '状況'
            case 'Com':
                dataType = 'コメント'
            case 'Iss':
                dataType = 'ID'
            case 'Res':
                dataType = 'ID'
            case 'Ver':
                dataType = 'バージョン'
            case 'Bod':
                dataType = '内容'
            
        itemName = ''
        # 項目名
        if len(column) == 0 or not 'LabelText' in column[0]:
        # 項目名を置換する
            #　項目の末尾を取得（ClassAとかのAの部分）
            tmp_itemNo = ''
            tmp_itemName = ''
            tmp = item[:3]
            if tmp == 'Cla' or tmp == 'Num' or tmp == 'Dat' or tmp == 'Des' or tmp == 'Che' or tmp == 'Att':
                if re.match(r"/[0-9]$/",item[-1]):
                    tmp_itemNo = item[-3:]
                    tmp_itemName = item[:-3]
                else:
                    tmp_itemNo = item[-1:]
                    tmp_itemName = item[:-3]
            match tmp:
                case 'Cla':
                    itemName = '分類' + tmp_itemNo
                case 'Num':
                    itemName = '数値' + tmp_itemNo
                case 'Dat':
                    itemName = '日付' + tmp_itemNo
                case 'Des':
                    itemName = '説明' + tmp_itemNo
                case 'Che':
                    itemName = 'チェック' + tmp_itemNo
                case 'Att':
                    itemName = '添付ファイル' + tmp_itemNo
                case 'Tit':
                    itemName = 'タイトル'
                case 'Sta':
                    itemName = '状況'
                case 'Com':
                    itemName = 'コメント'
                case 'Iss':
                    itemName = 'ID'
                case 'Res':
                    itemName = 'ID'
                case 'Ver':
                    itemName = 'バージョン'
                case 'Bod':
                    itemName = '内容' 
                case _:
                    itemName = ''
                    ws.cell(i,2).fill = PatternFill(patternType='solid', fgColor='FF0000')
        else:
            itemName = column[0]["LabelText"]
        
        # 選択肢のマスタ名
        choice = ''
        if not len(column) == 0:
            choice = column[0]["ChoicesText"] if 'ChoicesText' in column[0] else ''
            choice = choiceSelect(choice)

        
        ws.cell(i,1).value = item
        ws.cell(i,2).value = itemName
        ws.cell(i,3).value = dataType
        ws.cell(i,4).value = '〇'
        ws.cell(i,5).value = '〇' if item in linkColumns else ''
        ws.cell(i,6).value = '〇' if item in filterColumns else ''
        ws.cell(i,7).value = '〇' if item in gridColumns else ''

        if len(column) > 0:
            column = column[0]
            ws.cell(i,8).value = column["DecimalPlaces"] if "DecimalPlaces" in column else ''
            ws.cell(i,9).value = column["Min"] if "Min" in column else ('-999999999999999' if item[:3] == "Num" else '')
            ws.cell(i,10).value = column["Max"] if "Max" in column else ('999999999999999' if item[:3] == "Num" else '')
            ws.cell(i,11).value = column["DefaultInput"] if "DefaultInput" in column else ''
            ws.cell(i,12).value = column["Unit"] if "Unit" in column else ''
            ws.cell(i,13).value = "〇" if "ValidateRequired" in column and column["ValidateRequired"] == True else ''
            ws.cell(i,14).value = "〇" if "NoDuplication" in column and column["NoDuplication"] else ''
            ws.cell(i,15).value = choice
            ws.cell(i,16).value = "〇" if "MultipleSelections" in column and column["MultipleSelections"] else ''
            ws.cell(i,17).value = "〇" if "EditorReadOnly" in column and column["EditorReadOnly"] else ''
        else:
            ws.cell(i,9).value = '-999999999999999' if item[:3] == "Num" else ''
            ws.cell(i,10).value = '999999999999999' if item[:3] == "Num" else ''
        i += 1

wb.remove(wb['Sheet1'])
wb.save("tetete.xlsx")
print("終了")