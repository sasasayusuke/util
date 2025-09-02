from datetime import datetime
from time import strftime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font, PatternFill, Protection
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class UncollectedListService(BaseService):
    """未回収一覧表"""

    def display(self, request, session) -> Dict[str, Any]:

        # 未回収一覧表の印刷処理
        logger.info("【START】未回収一覧表")

        storedname="usp_ND0700未回収一覧表抽出"
        params = {
            "@iDateFrom":request.params["@iDateFrom"],
            "@iDateTo":request.params["@iDateTo"],
            "@is得意先CD":request.params["@is得意先CD"],
            "@ie得意先CD":request.params["@ie得意先CD"],
            "@i全額含む":request.params["@i全額含む"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_未回収一覧表.xlsx")
            ws = wb["Template"]
            ts = wb["Copy"]

            requestDateFrom = request.params["@iDateFrom"]
            requestDateTo =  request.params["@iDateTo"]
            custmerCodeFrom = request.params["@is得意先CD"]
            custmerCodeTo = request.params["@ie得意先CD"]

            # ヘッダー設定
            ws["F1"] = datetime.strptime(requestDateFrom, "%Y/%m/%d")
            ws["M1"] = datetime.strptime(requestDateTo, "%Y/%m/%d")
            ws["F2"] = custmerCodeFrom if custmerCodeFrom is not None else "最初"
            ws["K2"] = custmerCodeTo if custmerCodeTo is not None else "最後"
            ws["AU1"] = datetime.now()

            # 初期化
            start_row = 4
            current_row = start_row
            total_zeikomi_kingaku = 0
            total_nyukin_zumi_kingaku = 0
            total_mi_nyukin_gaku = 0

            # データのループ
            for result in storedresults['results'][0]:
                # 詳細行をコピー
                range_copy_cell_by_address(ts,ws,'A1','BQ2','A{}'.format(current_row),True)

                # データ挿入
                ws.cell(row=current_row + 1, column=1, value=result['入金状況'])
                ws.cell(row=current_row + 1, column=4, value=result['売上日付'])
                ws.cell(row=current_row + 1, column=9, value=result['請求日付'])
                ws.cell(row=current_row + 1, column=14, value=result['見積番号'])
                ws.cell(row=current_row + 1, column=18, value=result['得意先CD'])

                custmer_name1 = result['得意先名1']
                custmer_name2 = result['得意先名2']
                if (custmer_name2):
                    ws.cell(row=current_row, column=23, value=custmer_name1)         #上段
                    ws.cell(row=current_row + 1, column=23, value=custmer_name2)     #下段
                else:
                    ws.cell(row=current_row + 1, column=23, value=custmer_name1)     #下段

                ws.cell(row=current_row, column=33, value=result['見積件名'])
                ws.cell(row=current_row + 1, column=47, value=result['決算月'])
                ws.cell(row=current_row + 1, column=50, value=result['回収予定日'])

                zeikomi_kingaku = result['税込金額']
                nyukin_zumi_kingaku = result['入金済金額']
                mi_nyukin_gaku = result['未入金額']
                ws.cell(row=current_row + 1, column=55, value=zeikomi_kingaku)
                ws.cell(row=current_row + 1, column=60, value=nyukin_zumi_kingaku)
                ws.cell(row=current_row + 1, column=65, value=mi_nyukin_gaku)

                total_zeikomi_kingaku = total_zeikomi_kingaku + zeikomi_kingaku
                total_nyukin_zumi_kingaku = total_nyukin_zumi_kingaku + nyukin_zumi_kingaku
                total_mi_nyukin_gaku = total_mi_nyukin_gaku + mi_nyukin_gaku

                current_row = current_row + 2

            # 合計行を追加
            range_copy_cell_by_address(ts,ws,'A3','BQ4','A{}'.format(current_row),True)

            # データ挿入
            ws.cell(row=current_row + 1, column=55, value=total_zeikomi_kingaku)
            ws.cell(row=current_row + 1, column=60, value=total_nyukin_zumi_kingaku)
            ws.cell(row=current_row + 1, column=65, value=total_mi_nyukin_gaku)

            current_row = current_row + 1

            # シート名の変更
            ws.title = '未回収一覧表'

            # コピー用シートを削除
            wb.remove(ts)

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="BQ{}".format(current_row)
                )
                res_obj = converter.generate_pdf()

            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
