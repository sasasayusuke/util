from datetime import datetime
from time import strftime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import null
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell,
    range_copy_cell_by_address,
    insert_image,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsDates, ClsMitumoriH, ClsTanto,ClsKaisya, ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font, PatternFill, Protection
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class DailyDepositService(BaseService):
    """入金日計表"""

    def display(self, request, session) -> Dict[str, Any]:

        # 入金日計表の印刷処理
        logger.info("入金日計表")

        requestDateFrom = request.params["@iDateFrom"]
        requestDateTo =  request.params["@iDateTo"]
        custmerCodeFrom = request.params["@is得意先CD"]
        custmerCodeTo = request.params["@ie得意先CD"]

        sql_query = f"""
SELECT
    ND.入金番号
    , ND.枝番
    , ND.入金日付
    , ND.得意先CD
    , ND.得意先名1
    , ND.得意先名2
    , ND.入金区分CD
    , ND.入金区分名
    , ND.入金種別
    , ND.入金金額
    , ND.手形期日
    , ND.手形番号
    , ND.摘要名
    , TN.担当者名 
FROM
    TD入金 AS ND 
    LEFT JOIN TM得意先 AS TK 
        ON ND.得意先CD = TK.得意先CD 
    LEFT JOIN TM担当者 AS TN 
        ON TK.担当者CD = TN.担当者CD 
WHERE
    ND.入金日付 >= '{requestDateFrom}'
    AND ND.入金日付 <= '{requestDateTo}'
    AND ND.得意先CD >= COALESCE({f"'{custmerCodeFrom}'" if custmerCodeFrom else 'NULL'}, ND.得意先CD)
    AND ND.得意先CD <= COALESCE({f"'{custmerCodeTo}'" if custmerCodeTo else 'NULL'}, ND.得意先CD)
ORDER BY
    ND.得意先CD
    , ND.入金日付
    , ND.入金番号
    , ND.枝番;
    """

        sql_result=dict(SQLExecutor(session).execute_query(query=sql_query))

        if sql_result["count"] < 1:
            raise ServiceError('該当データなし')

        storedname="usp_ND0200入金日計表抽出"
        params = {
            "@iSDATE":request.params["@iDateFrom"],
            "@iFDATE":request.params["@iDateTo"],
            "@iSTOKUI":request.params["@is得意先CD"],
            "@iFTOKUI":request.params["@ie得意先CD"],
        }

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=params, outputparams={"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}))
        
        with get_excel_buffer() as buffer:
            wb = create_excel_object("Template_入金日計表.xlsx")
            ws = wb["Template"]
            ts = wb["Copy"]

            # ヘッダー設定
            ws["F2"] = datetime.strptime(requestDateFrom, "%Y/%m/%d")
            ws["M2"] = datetime.strptime(requestDateTo, "%Y/%m/%d")
            ws["F1"] = custmerCodeFrom if custmerCodeFrom is not None else "最初"
            ws["K1"] = custmerCodeTo if custmerCodeTo is not None else "最後"
            ws["BD1"] = datetime.now() 
            
            # 初期化
            start_row = 4
            current_row = start_row

            current_nyuukin_bangou = None
            total_amount = 0
            tantousya_name = None
            init_flag = True

            # SQL1データのループ
            for result in sql_result['results']:
                nyuukin_bangou = result['入金番号']
                if current_row == start_row:
                   current_nyuukin_bangou = nyuukin_bangou

                if current_nyuukin_bangou != nyuukin_bangou:
                    # データ行2をコピー
                    range_copy_cell_by_address(ts,ws,'A2','BO2','A{}'.format(current_row),True)
                    # データ挿入
                    ws.cell(row=current_row, column=38, value=total_amount) 
                    ws.cell(row=current_row, column=47, value=tantousya_name) 

                    total_amount = 0
                    current_nyuukin_bangou = nyuukin_bangou
                    init_flag = True
                    current_row += 1
                
                total_amount = total_amount + result['入金金額']
                tantousya_name = result['担当者名']
                # データ行1をコピー
                range_copy_cell_by_address(ts,ws,'A1','BO1','A{}'.format(current_row),True)
                # データ挿入
                if (init_flag):
                    ws.cell(row=current_row, column=1, value=f"[{result['得意先CD']}]") 
                    ws.cell(row=current_row, column=4, value=result['得意先名1']) 
                    ws.cell(row=current_row, column=14, value=result['得意先名2']) 
                    ws.cell(row=current_row, column=24, value=result['入金日付']) 
                    ws.cell(row=current_row, column=29, value=nyuukin_bangou) 
                    
                ws.cell(row=current_row, column=33, value=result['入金区分名']) 
                ws.cell(row=current_row, column=38, value=result['入金金額']) 

                if (result['入金種別'] == 2):
                    tagata_str_date = result['手形期日'].strftime("%Y/%m/%d")
                    ws.cell(row=current_row, column=44, value=f"手形期日:{tagata_str_date}") 
                    ws.cell(row=current_row, column=51, value=f"手形番号:{result['手形番号']}") 
                ws.cell(row=current_row, column=58, value=result['摘要名'])

                init_flag = False
                current_row += 1
            
            # 最後の入金番号の合計を出力
            # データ行2をコピー
            range_copy_cell_by_address(ts,ws,'A2','BO2','A{}'.format(current_row),True)
            # データ挿入
            ws.cell(row=current_row, column=38, value=total_amount) 
            ws.cell(row=current_row, column=47, value=tantousya_name) 
            
            current_row += 1

            # 合計行1をコピー
            range_copy_cell_by_address(ts,ws,'A3','BO3','A{}'.format(current_row),True)
            current_row += 1
            
            total_nikkei = 0
            total_ruikei = 0
            genkin_nikkei = 0
            genkin_ruikei = 0

            # SQL2データのループ
            for result in storedresults['results'][0]:
                nyuukin_kubun = result["入金区分名"]
                nikkei = result["入金日計"]
                ruikei = result["入金累計"]
                
                # 合計行2をコピー
                range_copy_cell_by_address(ts,ws,'A4','BO4','A{}'.format(current_row),True)
                # データ挿入
                ws.cell(row=current_row, column=5, value=nyuukin_kubun) 
                ws.cell(row=current_row, column=10, value=nikkei) 
                ws.cell(row=current_row, column=16, value=ruikei) 

                total_nikkei = total_nikkei + nikkei
                total_ruikei = total_ruikei + ruikei
                if (nyuukin_kubun == '現金'):
                    genkin_nikkei = nikkei
                    genkin_ruikei = ruikei
                
                current_row += 1

            # 合計行3をコピー
            range_copy_cell_by_address(ts,ws,'A5','BO5','A{}'.format(current_row),True)
            # 合計出力
            ws.cell(row=current_row, column=10, value=total_nikkei) 
            ws.cell(row=current_row, column=16, value=total_ruikei) 
            current_row += 1

            # 合計行4をコピー
            range_copy_cell_by_address(ts,ws,'A6','BO6','A{}'.format(current_row),True)
            # 現金率出力
            ws.cell(row=current_row, column=10, value=(genkin_nikkei / total_nikkei * 100)) 
            ws.cell(row=current_row, column=16, value=(genkin_ruikei / total_ruikei * 100)) 
            
            # シート名の変更
            ws.title = '入金日計表'

            # コピー用シートを削除
            wb.remove(ts)

            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell="BO{}".format(current_row)
                )
                res_obj = converter.generate_pdf()

                logger.debug("Pdf file generated successfully")
            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

                logger.debug("Excel file generated successfully")

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
