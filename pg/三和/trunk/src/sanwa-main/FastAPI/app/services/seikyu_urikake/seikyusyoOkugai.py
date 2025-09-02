from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi import requests
from fastapi.responses import StreamingResponse
from sqlalchemy import false
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.core.config import settings
from app.core.exceptions import ServiceError
from app.utils.excel_utils import (
    get_excel_buffer,
    create_excel_object,
    save_excel_to_buffer,
    range_copy_cell_by_address,
    ExcelPDFConverter
)
from app.utils.service_utils import ClsTanto,ClsKaisya,ClsMitumoriH,ClsZeiritsu,ClsOutputRireki
from app.utils.string_utils import null_to_zero
from openpyxl.styles.borders import Border,Side
from openpyxl.styles.alignment import Alignment


# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

# class OkugaiSeikyusyoService(BaseService):
#     """請求書発行　屋外広告物専用"""

#     def display(self, request, session) -> Dict[str, Any]:

#         # 【START】請求書発行　屋外広告物専用の印刷処理
#         logger.info("【START】請求書発行　屋外広告物専用")

        
#         # 見積番号チェック
#         cSeikyuH = ClsMitumoriH(request.params["@i見積番号"],0,0)
#         b_category = cSeikyuH.IsUCategory(session,'B')
#         u_category = cSeikyuH.IsUCategory(session,'U')
        
#         if b_category and u_category:
#             raise ServiceError('原価・売価未確定項目があります。')
#         if b_category:
#             raise ServiceError('売価未確定項目があります。')
#         if u_category:
#             raise ServiceError('原価未確定項目があります。')
        
#         # ロックのチェック
#         # /////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        
#         # 担当者チェック
#         cTanto = ClsTanto()
#         tanto_results = cTanto.GetbyID(session,request.params["@i担当者CD"])
#         if len(tanto_results) == 0:
#             raise ServiceError('指定の担当者は存在しません。')
        
#         # 売上の請求できるかのチェックと請求用ヘッダー作成処理
#         self.download_kakutei(session,request.params["@i見積番号"],request.params["@i請求日付"])
        
#         stored_results = self.download(session,request.params["@i見積番号"],request.params["@i請求書発行日付"])

#         logger.info('storedresults:{}'.format(stored_results['results']))

#         with get_excel_buffer() as buffer:
#             wb = create_excel_object('Temp_請求書_屋外広告物.xlsx')
#             ws = wb["Template"]
            
#             grs_head = stored_results["results"][0][0]
#             grs_detail = stored_results["results"][2]
            
#             # 請求日付と請求書発行日の日付データを用意
#             billing_date = datetime.strptime(request.params["@i請求日付"],'%Y-%m-%d')
#             billing_output_date = datetime.strptime(request.params["@i請求書発行日付"],'%Y-%m-%d')
            
#             # ヘッダー入力
#             ws['A1'].value = "№{}".format(grs_head["見積番号"])
#             ws['O1'].value = billing_output_date.strftime('%Y{}%m{}%d{}').format('年','月','日')
#             if grs_head["得意先名2"] == "":
#                 ws['A3'].value = ""
#                 ws['A4'].value = grs_head["得意先名1"]
#             else:
#                 ws['A3'].value = grs_head["得意先名1"]
#                 ws['A4'].value = grs_head["得意先名2"]
#             ws['D8'].value = grs_head["合計金額"]
            
#             ClsTax = ClsZeiritsu()
#             ws['A10'].value = "消費税等({}%)".format(ClsTax.GetbyDate(session,grs_head["請求日付"]))
            
#             ws['D10'].value = grs_head["外税額"]
#             ws['D13'].value = grs_head["見積件名"]
            
#             if grs_head["納期S"] != None and grs_head["納期E"] != None:
#                 ws['D15'].value = "{}～{}".format(
#                     datetime.strftime(grs_head["納期S"],'%Y{}%m{}%d{}'.format('年','月','日')),
#                     datetime.strftime(grs_head["納期E"],'%Y{}%m{}%d{}'.format('年','月','日')),
#                 )
#             elif grs_head["納期S"] == None:
#                 ws['D15'].value = datetime.strftime(grs_head["納期E"],'%Y{}%m{}%d{}'.format('年','月','日'))
#             else:
#                 datetime.strftime(grs_head["納期S"],'%Y{}%m{}%d{}'.format('年','月','日'))
            
#             if grs_head["納入得意先CD"] in ['2401','8801']:
#                 ws['D16'].value = grs_head["現場名"].replace(grs_head["納入先CD"],"")
#                 ws['F16'].value = grs_head["納入先CD"]
#             else:
#                 ws['D16'].value = grs_head['現場名']
#                 ws['F16'].value = ""
            
#             if grs_head['支払条件'] == 0:
#                 ws['D17'].value = "別途御打ち合せによる"
#             elif grs_head['支払条件'] == 1:
#                 ws['D17'].value = grs_head['支払条件その他']
            
#             ws['D18'].value = grs_head['備考']
            
#             ws['K16'].value = "担当：{}".format(tanto_results[0]["問い合わせ先"])
#             ws['K17'].value = ""
            
#             if grs_head['ウエルシア売場面積'] != 0:
#                 ws['A18'].value = '売 場 面 積 ：'
#                 ws['D18'].value = "{}坪".format(grs_head['ウエルシア売場面積'])
#                 tmp = Side(style='thin',color='000000')
#                 ws['A18:D18'].border = Border(bottom=tmp)
            
#             # インボイス番号
#             cKaisya = ClsKaisya().GetData(session)
#             if len(cKaisya) == 1 and cKaisya[0]["インボイス登録番号"] != "":
#                 ws['K9'].value = "登録番号 {}".format(cKaisya[0]["インボイス登録番号"])
#             else:
#                 ws['K9'].value = ""
            
#             # 改ページプレビューの設定
#             ws.sheet_view.view = 'pageBreakPreview'
#             ws.sheet_view.zoomScaleSheetLayoutView= 100

#             # 明細書込み
#             row = 21
#             siwake_no = ""

#             # 改ページ処理
#             def insert_detail(ws,copy_ws,row):
#                 row = ((page_count-1) * 35) + 39
#                 range_copy_cell_by_address(copy_ws,ws,'A1','O35','A{}'.format(row),True)
#                 return [ws,row]

            
            
#             for row_data in grs_detail:
                
#                 if siwake_no != row_data["仕分番号"]:
                    
#                     ws.cell(row=row,column=1,value=row_data["仕分番号"])
#                     ws.cell(row=row,column=3,value=null_to_zero(row_data["仕分名称"],""))

#                     siwake_no = row_data["仕分番号"]
                    
#                     row = row + 1
                
#                 # 見積シート情報セット
#                 ws.cell(row=row,column=1,value=row_data["仕分番号"])
#                 if row_data["伝票区分"] in ['C','A','S']:
#                     ws.cell(row=row,column=2,value="")
#                     ws.cell(row=row,column=3,value="  {}".format(row_data["漢字名称"]))
#                 else:
#                     ws.cell(row=row,column=2,value=row_data["行番号"])
#                     ws.cell(row=row,column=3,value="  {}".format(null_to_zero(row_data["漢字名称"],"")))
#                 ws.cell(row=row,column=5,value="{}{}".format(row_data["PC区分"] , row_data["製品NO"]))
#                 ws.cell(row=row,column=6,value=row_data["仕様NO"])
#                 ws.cell(row=row,column=7,value=row_data["W"])
                
#                 if row_data['D'] == 0:
#                     if row_data['D1'] == 0 and row_data['D2'] == 0:
#                         ws.cell(row=row,column=8,value="")
#                     else:
#                         ws.cell(row=row,column=8,value="{}/{}".format(row_data['D1'],row_data['D2']))
#                 else:
#                     ws.cell(row=row,column=8,value="d")
                
#                 if row_data['H'] == 0:
#                     if row_data['H1'] == 0 and row_data['H2'] == 0:
#                         ws.cell(row=row,column=9,value="")
#                     else:
#                         ws.cell(row=row,column=9,value="{}/{}".format(row_data['H1'],row_data['H2']))
#                 else:
#                     ws.cell(row=row,column=9,value=row_data['H'])
                
#                 ws.cell(row=row,column=10,value=row_data['売上数量'])
#                 ws.cell(row=row,column=12,value=row_data['単位名'])
#                 ws.cell(row=row,column=13,value=row_data['売上単価'])
                
#                 if not row_data["伝票区分"] in ['A','C','S']:
#                     ws.cell(row=row,column=15,value=row_data['売上金額'])
#                 else:
#                     if null_to_zero(row_data['漢字名称'],"") in "\\":
#                         ws.cell(row=row,column=10,value=1)
#                         ws.cell(row=row,column=12,value="式")
#                         tmp = null_to_zero(row_data["漢字名称"],"")
#                         position = tmp.find("\\")
#                         tmp = tmp[position:]
#                         ws.cell(row=row,column=13,value=tmp)
#                         ws.cell(row=row,column=15,value=tmp)
#                         ws['D9'].value = tmp
#                     else:
#                         ws.cell(row=row,column=15,value="" if row_data["売上金額"] else row_data["売上金額"])
                
#                 row = row + 1

#             # 合計
#             row = row + 1
#             ws.cell(row=row,column=3,value="【合　　計】")
#             ws["C{0}".format(row)].alignment = Alignment(horizontal='center')
#             ws["D{0}".format(row)].alignment = Alignment(horizontal='center')
#             ws.cell(row=row,column=15,value="=SUM(O{}:O{})".format(23,row-2))
            
#             # 印刷範囲の設定
#             # ws.print_area = 'A1:T' + str(len(storedresults["results"][0])+2)

#             # シート名の変更
#             ws.title = '請求書'
            
#             res_obj = None
#             res_type = request.params["@type"]
#             if res_type == 'pdf':
#                 # PDFファイルの返却
#                 converter = ExcelPDFConverter(
#                     wb=wb,
#                     ws=ws,
#                     start_cell="A1",
#                     end_cell="O{}".format(row)
#                 )
#                 res_obj = converter.generate_pdf()
#             else:
#                 # Excelオブジェクトの保存
#                 save_excel_to_buffer(buffer, wb, None)
#                 res_obj = buffer.getvalue()

#                 logger.debug("Excel file generated successfully")
                
#             # 履歴情報セット
#             cOutputRireki = ClsOutputRireki()
#             cOutputRireki.SetOutputLog(session,'請求書',request.params["@i見積番号"])

#             media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

#             return StreamingResponse(
#                 io.BytesIO(res_obj),
#                 media_type=media_type,
#                 headers={
#                     'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
#                 }
#             )


#     def execute(self, request, session) -> Dict[str, Any]:
#         """実行処理を行うメソッド"""
#         pass

#     def download_kakutei(self,session,estimate_no,billing_date) -> bool:
#         res_bool = False
        
#         stored_name = "usp_SE0801請求確定処理"
        
#         params = {
#             "@i見積番号":estimate_no,
#             "@i請求日付":billing_date
#         }
        
#         outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

#         # STORED実行
#         storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=params,outputparams=outputparams))

#         logger.info('storedresults:{}'.format(storedresults['results']))
        
#         result = storedresults["output_values"]
#         if result["@RetST"] != 0:
#             res_bool = False
#             raise ServiceError("{}：{}".format(result["@RetST"],result["@RetMsg"]))
#         else:
#             res_bool = True
            
#         return res_bool
    
#     def download(self,session,estimate_no,billing_output_date):
        
#         params = {
#             "@i見積番号":estimate_no,
#             "@i請求書発行日付":billing_output_date
#         }
        
#         stored_name = "usp_SE0801請求書発行_客在"
#         outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}
        
#         storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=params,outputparams=outputparams))
        
#         result = storedresults["output_values"]
#         if storedresults["output_values"]["@RetST"] != 0:
#             raise ServiceError("{}：{}".format(result["@RetST"],result["@RetMsg"]))
#         elif len(storedresults["results"][0]) == 0:
#             raise ServiceError("該当データ無し")
        
#         if len(storedresults["results"][1]) == 0:
#             raise ServiceError("該当データ無し")
#         if len(storedresults["results"][2]) == 0:
#             raise ServiceError("該当データ無し")
#         return storedresults




class OkugaiSeikyusyoService(BaseService):
    """請求書発行　屋外広告物専用"""

    def display(self, request, session) -> Dict[str, Any]:

        # 【START】請求書発行　屋外広告物専用の印刷処理
        logger.info("【START】請求書発行　屋外広告物専用")

        
        # 見積番号チェック
        cSeikyuH = ClsMitumoriH(request.params["@i見積番号"],0,0)
        b_category = cSeikyuH.IsUCategory(session,'B')
        u_category = cSeikyuH.IsUCategory(session,'U')
        
        if b_category and u_category:
            raise ServiceError('原価・売価未確定項目があります。')
        if b_category:
            raise ServiceError('売価未確定項目があります。')
        if u_category:
            raise ServiceError('原価未確定項目があります。')
        
        # ロックのチェック
        # /////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        
        # 担当者チェック
        cTanto = ClsTanto()
        tanto_results = cTanto.GetbyID(session,request.params["@i担当者CD"])
        if len(tanto_results) == 0:
            raise ServiceError('指定の担当者は存在しません。')
        
        # 売上の請求できるかのチェックと請求用ヘッダー作成処理
        self.download_kakutei(session,request.params["@i見積番号"],request.params["@i請求日付"])
        
        stored_results = self.download(session,request.params["@i見積番号"],request.params["@i請求書発行日付"])

        logger.info('storedresults:{}'.format(stored_results['results']))

        with get_excel_buffer() as buffer:
            wb = create_excel_object('Temp_請求書_屋外広告物.xlsx')
            ws = wb["Template"]
            copy_ws = wb["copy_sheet"]
            
            grs_head = stored_results["results"][0][0]
            grs_detail = stored_results["results"][2]
            
            # 請求日付と請求書発行日の日付データを用意
            billing_date = datetime.strptime(request.params["@i請求日付"],'%Y-%m-%d')
            billing_output_date = datetime.strptime(request.params["@i請求書発行日付"],'%Y-%m-%d')
            
            # ヘッダー入力
            ws['A1'].value = "№{}".format(grs_head["見積番号"])
            ws['O1'].value = billing_output_date.strftime('%Y{}%m{}%d{}').format('年','月','日')
            if grs_head["得意先名2"] == "":
                ws['A3'].value = ""
                ws['A4'].value = grs_head["得意先名1"]
            else:
                ws['A3'].value = grs_head["得意先名1"]
                ws['A4'].value = grs_head["得意先名2"]
            ws['D8'].value = grs_head["合計金額"]
            
            ClsTax = ClsZeiritsu()
            ws['A10'].value = "消費税等({}%)".format(ClsTax.GetbyDate(session,grs_head["請求日付"]))
            
            ws['D10'].value = grs_head["外税額"]
            ws['D13'].value = grs_head["見積件名"]
            
            if grs_head["納期S"] != None and grs_head["納期E"] != None:
                ws['D15'].value = "{}～{}".format(
                    datetime.strftime(grs_head["納期S"],'%Y{}%m{}%d{}'.format('年','月','日')),
                    datetime.strftime(grs_head["納期E"],'%Y{}%m{}%d{}'.format('年','月','日')),
                )
            elif grs_head["納期S"] == None and grs_head["納期E"] != None:
                ws['D15'].value = datetime.strftime(grs_head["納期E"],'%Y{}%m{}%d{}'.format('年','月','日'))

            elif grs_head["納期S"] != None:
                ws['D15'].value = datetime.strftime(grs_head["納期S"],'%Y{}%m{}%d{}'.format('年','月','日'))
            
            if grs_head["納入得意先CD"] in ['2401','8801']:
                ws['D16'].value = grs_head["現場名"].replace(grs_head["納入先CD"],"")
                ws['F16'].value = grs_head["納入先CD"]
            else:
                ws['D16'].value = grs_head['現場名']
                ws['F16'].value = ""
            
            if grs_head['支払条件'] == 0:
                ws['D17'].value = "別途御打ち合せによる"
            elif grs_head['支払条件'] == 1:
                ws['D17'].value = grs_head['支払条件その他']
            
            ws['D18'].value = grs_head['備考']
            
            ws['K16'].value = "担当：{}".format(tanto_results[0]["問い合わせ先"])
            ws['K17'].value = ""
            
            if grs_head['ウエルシア売場面積'] != 0:
                ws['A18'].value = '売 場 面 積 ：'
                ws['D18'].value = "{}坪".format(grs_head['ウエルシア売場面積'])
                tmp = Side(style='thin',color='000000')
                ws['A18:D18'].border = Border(bottom=tmp)
            
            # インボイス番号
            cKaisya = ClsKaisya().GetData(session)
            if len(cKaisya) == 1 and cKaisya[0]["インボイス登録番号"] != "":
                ws['K9'].value = "登録番号 {}".format(cKaisya[0]["インボイス登録番号"])
                ws['K11'].value = "〒{}".format(cKaisya[0]["郵便番号"])
                ws['K12'].value = cKaisya[0]["住所1"]
                ws['K13'].value = "TEL {}(代)".format(cKaisya[0]["電話番号"])
                ws['K14'].value = "FAX {}".format(cKaisya[0]["FAX番号"])
            else:
                ws['K9'].value = ""
            
            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # 明細書込み
            row = 22
            row_count_for_page = 1
            page_count = 1
            siwake_no = ""



            # 明細ページ挿入関数
            def insert_detail(ws,copy_ws,row):
                row = ((page_count-1) * 35) + 38
                range_copy_cell_by_address(copy_ws,ws,'A1','O35','A{}'.format(row),True)
                return [ws,row]
            
            for row_data in grs_detail:

                if row == 38:
                    res = insert_detail(ws,copy_ws,row-1)
                    ws = res[0]
                    row = 42
                    page_count = 2
                    row_count_for_page = 1
                
                if row_count_for_page == 33:
                    res = insert_detail(ws,copy_ws,row)
                    ws=res[0]
                    row = res[1]
                    row = row + 3
                    page_count = page_count + 1
                    row_count_for_page = 1

                
                
                if siwake_no != row_data["仕分番号"]:
                    
                    ws.cell(row=row,column=1,value=row_data["仕分番号"])
                    ws.cell(row=row,column=3,value=null_to_zero(row_data["仕分名称"],""))

                    siwake_no = row_data["仕分番号"]
                    
                    row = row + 1
                
                # 見積シート情報セット
                ws.cell(row=row,column=1,value=row_data["仕分番号"])
                if row_data["伝票区分"] in ['C','A','S']:
                    ws.cell(row=row,column=2,value="")
                    ws.cell(row=row,column=3,value="  {}".format(row_data["漢字名称"]))
                else:
                    ws.cell(row=row,column=2,value=row_data["行番号"])
                    ws.cell(row=row,column=3,value="  {}".format(null_to_zero(row_data["漢字名称"],"")))
                ws.cell(row=row,column=5,value="{}{}".format(row_data["PC区分"] , row_data["製品NO"]))
                ws.cell(row=row,column=6,value=row_data["仕様NO"])
                ws.cell(row=row,column=7,value=row_data["W"])
                
                if row_data['D'] == 0:
                    if row_data['D1'] == 0 and row_data['D2'] == 0:
                        ws.cell(row=row,column=8,value="")
                    else:
                        ws.cell(row=row,column=8,value="{}/{}".format(row_data['D1'],row_data['D2']))
                else:
                    ws.cell(row=row,column=8,value=row_data["D"])
                
                if row_data['H'] == 0:
                    if row_data['H1'] == 0 and row_data['H2'] == 0:
                        ws.cell(row=row,column=9,value="")
                    else:
                        ws.cell(row=row,column=9,value="{}/{}".format(row_data['H1'],row_data['H2']))
                else:
                    ws.cell(row=row,column=9,value=row_data['H'])
                
                ws.cell(row=row,column=10,value=row_data['売上数量'])
                ws.cell(row=row,column=11,value='=IF(MOD(J{0},1)=0,TEXT(J{0},"#,###"),TEXT(J{0},"#,##0.##"))'.format(row))
                ws.cell(row=row,column=12,value=row_data['単位名'])
                ws.cell(row=row,column=13,value=row_data['売上単価'])
                ws.cell(row=row,column=14,value='=IF(MOD(M{0},1)=0,TEXT(M{0},"#,###"),TEXT(M{0},"#,##0.##"))'.format(row))
                
                if not row_data["伝票区分"] in ['A','C','S']:
                    ws.cell(row=row,column=15,value=row_data['売上金額'])
                else:
                    if "\\" in null_to_zero(row_data['漢字名称'],""):
                        ws.cell(row=row,column=10,value=1)
                        ws.cell(row=row,column=12,value="式")
                        tmp = null_to_zero(row_data["漢字名称"],"")
                        position = tmp.find("\\")
                        tmp = tmp[position:]
                        ws.cell(row=row,column=13,value=int(tmp.replace("\\","").replace(",","")))
                        ws.cell(row=row,column=15,value=int(tmp.replace("\\","").replace(",","")))
                        ws['D9'].value = int(tmp.replace("\\","").replace(",",""))
                    else:
                        ws.cell(row=row,column=15,value="" if row_data["売上金額"] == 0 else row_data["売上金額"])
                
                row = row + 1
                row_count_for_page = row_count_for_page + 1

            # 合計
            row = row + 1
            ws.cell(row=row,column=3,value="【合　　計】")
            ws["C{0}".format(row)].alignment = Alignment(horizontal='center')
            ws["D{0}".format(row)].alignment = Alignment(horizontal='center')
            ws.cell(row=row,column=15,value="=SUM(O{}:O{})".format(23,row-2))
            
            # 印刷範囲の設定
            ws.print_area = 'A1:O{}'.format(((page_count-1) * 35) + 37)
            # ws.print_area = 'A1:O37'


            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # コピー用シートを削除
            wb.remove(copy_ws)

            # シート名の変更
            ws.title = '請求書'
            
            res_obj = None
            res_type = request.params["@type"]
            if res_type == 'pdf':
                # PDFファイルの返却
                converter = ExcelPDFConverter(
                    wb=wb,
                    ws=ws,
                    start_cell="A1",
                    end_cell='O{}'.format(((page_count-1) * 35) + 37)
                )
                res_obj = converter.generate_pdf()
            else:
                # Excelオブジェクトの保存
                save_excel_to_buffer(buffer, wb, None)
                res_obj = buffer.getvalue()

                logger.debug("Excel file generated successfully")
                
            # 履歴情報セット
            cOutputRireki = ClsOutputRireki()
            cOutputRireki.SetOutputLog(session,'請求書',request.params["@i見積番号"],request.params["@pc_name"])

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if res_type == "pdf" else "application/pdf"

            return StreamingResponse(
                io.BytesIO(res_obj),
                media_type=media_type,
                headers={
                    'Content-Disposition': 'attachment; filename="sample.{}"'.format(res_type)
                }
            )


    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass

    def download_kakutei(self,session,estimate_no,billing_date) -> bool:
        res_bool = False
        
        stored_name = "usp_SE0801請求確定処理"
        
        params = {
            "@i見積番号":estimate_no,
            "@i請求日付":billing_date
        }
        
        outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}

        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=params,outputparams=outputparams))

        logger.info('storedresults:{}'.format(storedresults['results']))
        
        result = storedresults["output_values"]
        if result["@RetST"] != 0:
            res_bool = False
            raise ServiceError("{}：{}".format(result["@RetST"],result["@RetMsg"]))
        else:
            res_bool = True
            
        return res_bool
    
    def download(self,session,estimate_no,billing_output_date):
        
        params = {
            "@i見積番号":estimate_no,
            "@i請求書発行日付":billing_output_date
        }
        
        stored_name = "usp_SE0801請求書発行_客在"
        outputparams = {"@RetST":"INT", "@RetMsg":"VARCHAR(255)"}
        
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=stored_name, params=params,outputparams=outputparams))
        
        result = storedresults["output_values"]
        if storedresults["output_values"]["@RetST"] != 0:
            raise ServiceError("{}：{}".format(result["@RetST"],result["@RetMsg"]))
        elif len(storedresults["results"][0]) == 0:
            raise ServiceError("該当データ無し")
        
        if len(storedresults["results"][1]) == 0:
            raise ServiceError("該当データ無し")
        if len(storedresults["results"][2]) == 0:
            raise ServiceError("該当データ無し")
        return storedresults
