from datetime import datetime
from typing import Dict, Any
import io
import logging
from fastapi.responses import StreamingResponse
from copy import copy
from app.core.exceptions import ServiceError
from app.services.base_service import BaseService
from app.utils.db_utils import SQLExecutor
from app.utils.excel_utils import get_excel_buffer, create_excel_object, save_excel_to_buffer
from app.utils.string_utils import format_jp_date
from app.core.config import settings
import locale
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break

# 名前を指定してロガーを取得する
logger = logging.getLogger(settings.APP_NAME)

class SekouiraisyoService(BaseService):
    """施工依頼書サービス"""

    def display(self, request, session) -> StreamingResponse:
        """印刷処理を行うメソッド"""

        # 施工依頼書出力処理
        logger.info("【START】施工依頼書出力処理")

        storedname="usp_MT1000施工依頼書出力"

        mitsumori_no = {
            "@i見積番号":request.params['@i見積番号']
        }


        # STORED実行
        storedresults = dict(SQLExecutor(session).execute_stored_procedure(storedname=storedname, params=mitsumori_no))

        if storedresults["count"] < 1:
            raise ServiceError('該当データなし')

        with get_excel_buffer() as buffer:
            # Excelオブジェクトの作成
            wb = create_excel_object('Temp施工依頼書.xlsx')
            ws = wb.active

            # EXCELに情報を追加
            # 依頼日
            ws['F1'] = format_jp_date()
            # 依頼者
            ws['J3'] = request.user
            # 店舗NO
            ws['A6'] = storedresults['results'][0][0]['納入先CD']
            # 物件名
            if storedresults['results'][0][0]['納入先名2'] == "":
                ws['C6'] = storedresults['results'][0][0]['納入先名1']
            else:
                ws['C6'] = "{}\n{}".format(storedresults['results'][0][0]['納入先名1'],storedresults['results'][0][0]['納入先名2'])
            # 見積りNO
            ws['A8'] = int(storedresults['results'][0][0]['見積番号'])
            # 施工日
            if request.params['施工日'] != "":
                ws['C8'] = format_jp_date(request.params['施工日'])
            else:
                ws['C8'] = ""
            # 住所
            if storedresults['results'][0][0]['住所2'] == "":
                ws['C11'] = storedresults['results'][0][0]['住所1']
            else:
                ws['C11'] = "{}\n{}".format(storedresults['results'][0][0]['住所1'],storedresults['results'][0][0]['住所2'])
            # 電話
            ws['C12'] = storedresults['results'][0][0]['納TEL']
            # 備考
            ws['A16'] = storedresults['results'][0][0]['備考']
            # 作業内容
            ws['A19'] = storedresults['results'][0][0]['作業内容']

            # 改ページプレビューの設定
            ws.sheet_view.view = 'pageBreakPreview'
            ws.sheet_view.zoomScaleSheetLayoutView= 100

            # 罫線の設定
            ws['L5'].border = Border(
                 right=Side(style='medium', color='000000'),
                 top=Side(style='medium', color='000000')
                 )
            ws['L7'].border = Border(
                 right=Side(style='medium', color='000000'),
                 top=Side(style='medium', color='000000')
                 )
            ws['L10'].border = Border(
                 right=Side(style='medium', color='000000'),
                 top=Side(style='medium', color='000000')
                 )
            ws['A11'].border = Border(
                 left=Side(style='medium', color='000000'),
                 bottom=Side(style='thin', color='000000')
                 )
            ws['A12'].border = Border(
                 left=Side(style='medium', color='000000'),
                 bottom=Side(style='thin', color='000000')
                 )
            
            ws.row_breaks.append(Break(id=37))
            # ws.row_breaks.append(37)

            # シート名の変更
            ws.title = '依頼書'

            # Excelオブジェクトの保存
            save_excel_to_buffer(buffer, wb, None)

            return StreamingResponse(
                io.BytesIO(buffer.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    'Content-Disposition': 'attachment; filename="sample.xlsx"'
                }
            )

    def execute(self, request, session) -> Dict[str, Any]:
        """実行処理を行うメソッド"""
        pass
